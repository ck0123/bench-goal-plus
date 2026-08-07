from __future__ import annotations

import json
import os
import stat
import subprocess
import tempfile
import threading
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from bench_goal_plus.application import BenchmarkAgent
from bench_goal_plus.catalog import Catalog
from bench_goal_plus.errors import ContractError
from bench_goal_plus.loopback_bridge import bridged_url, loopback_target
from bench_goal_plus.runners.factory import create_runner
from bench_runtime_paths import ensure_temp_root
from experiments.swe_bench_verified import (
    checkout,
    environment,
    goal_plus_evidence,
    reporting,
    runtime,
)
from experiments.swe_bench_verified.config import (
    SweBenchContractError,
    load_profile,
    read_json,
    validate_profile,
    write_json,
)
from scripts import benchmark_report


class SweBenchVerifiedContractTest(unittest.TestCase):
    def temporary_directory(self) -> tempfile.TemporaryDirectory[str]:
        return tempfile.TemporaryDirectory(dir=ensure_temp_root("test-swe-bench-verified"))

    def profile(self, profile_id: str) -> dict:
        return load_profile(profile_id)[1]

    def checkout_probe_output(
        self,
        base_commit: str,
        *,
        observed_head: str,
        parents: list[str] | None = None,
        author_email: str = checkout.OFFICIAL_SETUP_EMAIL,
        committer_email: str = checkout.OFFICIAL_SETUP_EMAIL,
        subject: str = checkout.OFFICIAL_SETUP_SUBJECT,
        clean: bool = True,
        base_tree: str = "6" * 40,
        observed_tree: str = "7" * 40,
    ) -> str:
        return "\n".join(
            [
                f"observed_head={observed_head}",
                f"parents={' '.join(parents if parents is not None else [base_commit])}",
                f"author_email={author_email}",
                f"committer_email={committer_email}",
                f"subject={subject}",
                f"observed_tree={observed_tree}",
                f"base_tree={base_tree}",
                f"clean={'true' if clean else 'false'}",
            ]
        )

    def test_goal_plus_installer_uses_the_bind_cache_owner(self) -> None:
        script = environment.goal_plus_install_script()

        self.assertIn("os.stat('/opt/pip-cache')", script)
        self.assertIn("os.setgid(cache.st_gid)", script)
        self.assertIn("os.setuid(cache.st_uid)", script)
        self.assertIn("'/opt/goal-plus-runtime-requirements.lock'", script)
        self.assertIn("PATH=/opt/goal-plus-bin:/opt/node/bin:$PATH", script)
        self.assertNotIn("PATH", environment.goal_plus_runtime_environment())
        self.assertEqual(
            environment.goal_plus_runtime_environment()["HOME"],
            "/opt/agent-tmp",
        )

    def write_goal_plus_state(
        self,
        root: Path,
        *,
        max_parallel: int = 1,
        session_count: int = 1,
        verifier_runs: int = 1,
    ) -> None:
        run_id = "run_test"
        candidate_id = "c001"
        write_json(
            root / "goal-plus/gp_test/goal.json",
            {
                "goal_plus_id": "gp_test",
                "status": "complete",
                "phase": "done",
                "linked_search": {"run_id": run_id},
            },
        )
        write_json(
            root / f"runs/{run_id}/run.json",
            {
                "run_id": run_id,
                "state": "promoted",
                "frozen_spec_id": "spec_test",
                "selected_candidate_id": candidate_id,
            },
        )
        write_json(
            root / "specs/spec_test/frozen_spec.json",
            {
                "spec": {
                    "budget": {"max_parallel": max_parallel},
                    "strategy": {
                        "worker_host": "pi-rpc",
                        "orchestration_mode": "parallel_loops",
                        "worker_budget": {"max_runtime_seconds": 1500},
                        "config": {"closeout_reserve_seconds": 300},
                    },
                    "process_verifiers": [
                        {
                            "name": "visible",
                            "role": "ranking_signal",
                            "command": [
                                "python",
                                ".goal-plus-verifiers/visible_test_verifier.py",
                                "--timeout-seconds",
                                "300",
                                "--",
                                "python",
                                "-m",
                                "pytest",
                            ],
                        }
                    ],
                    "promotion_verifiers": [
                        {
                            "name": "visible-promotion",
                            "role": "promotion_gate",
                            "command": [
                                "python",
                                ".goal-plus-verifiers/visible_test_verifier.py",
                                "--timeout-seconds",
                                "300",
                                "--",
                                "python",
                                "-m",
                                "pytest",
                            ],
                        }
                    ],
                }
            },
        )
        write_json(
            root / f"runs/{run_id}/candidates/{candidate_id}/candidate.json",
            {"candidate_id": candidate_id, "iterations": [{"score": 1.0}]},
        )
        for index in range(session_count):
            write_json(
                root / f"runs/{run_id}/agent_sessions/agent_{index}.json",
                {
                    "host": "pi-rpc",
                    "candidate_id": candidate_id,
                    "host_handle": {
                        "external_id": f"agent_{index}",
                        "metadata": {
                            "pi_metrics": {
                                "usage_total": {"input_tokens": 10 + index}
                            }
                        },
                    },
                    "counters": {"verifier_runs": verifier_runs},
                },
            )
        promotion = root / f"runs/{run_id}/promotion/{candidate_id}.patch"
        promotion.parent.mkdir(parents=True, exist_ok=True)
        promotion.write_text("diff --git a/a b/a\n", encoding="utf-8")

    def test_catalog_presets_freeze_methods_and_tkcr(self) -> None:
        agent = BenchmarkAgent(catalog=Catalog())
        codex = agent.resolve_spec(
            preset_id="swe-bench-verified-sympy-16886-codex-smoke"
        )
        pi = agent.resolve_spec(
            preset_id="swe-bench-verified-sympy-16886-pi-smoke"
        )
        goal_plus_pi = agent.resolve_spec(
            preset_id="swe-bench-verified-sympy-16886-goal-plus-pi-smoke"
        )
        luna_goal_plus_pi = agent.resolve_spec(
            preset_id=(
                "swe-bench-verified-sympy-16886-goal-plus-pi-luna-high-smoke"
            )
        )
        codex_c2 = agent.resolve_spec(
            preset_id="swe-bench-verified-sympy-codex-terra-c2-smoke"
        )

        self.assertEqual(codex.runner.runner_id, "swe-bench-native")
        self.assertEqual(codex.methods, ("plain-codex",))
        self.assertEqual(codex.concurrency(), {"T": 1800, "K": 1, "C": 1, "R": 1})
        self.assertEqual(pi.methods, ("plain-pi",))
        self.assertEqual(pi.model, "zai/glm-5.2")
        self.assertEqual(pi.concurrency(), {"T": 1800, "K": 1, "C": 1, "R": 1})
        self.assertEqual(goal_plus_pi.methods, ("goal-plus-pi",))
        self.assertEqual(goal_plus_pi.model, "zai/glm-5.2")
        self.assertEqual(
            goal_plus_pi.concurrency(), {"T": 1800, "K": 1, "C": 1, "R": 1}
        )
        self.assertEqual(luna_goal_plus_pi.methods, ("goal-plus-pi",))
        self.assertEqual(luna_goal_plus_pi.model, "bench-openai/gpt-5.6-luna")
        self.assertEqual(luna_goal_plus_pi.reasoning_effort, "high")
        self.assertEqual(
            luna_goal_plus_pi.concurrency(),
            {"T": 1800, "K": 1, "C": 1, "R": 1},
        )
        self.assertEqual(codex_c2.methods, ("plain-codex",))
        self.assertEqual(codex_c2.model, "gpt-5.6-terra")
        self.assertEqual(
            codex_c2.concurrency(), {"T": 1800, "K": 1, "C": 2, "R": 1}
        )
        self.assertTrue(codex_c2.runner.capabilities.cell_concurrency)
        self.assertTrue(codex.runner.capabilities.official_evaluator)
        self.assertTrue(codex.runner.capabilities.retain_containers)
        self.assertTrue(codex.runner.capabilities.detach)
        self.assertEqual(codex.runner.evidence_filename, "campaign-summary.json")

    def test_profiles_pin_the_official_base_commit(self) -> None:
        expected = "c50643a49811e9fe2f4851adff4313ad46f7325e"

        self.assertEqual(
            self.profile("sympy-16886-codex-smoke")["tasks"][0]["base_commit"],
            expected,
        )
        self.assertEqual(
            self.profile("sympy-16886-pi-smoke")["tasks"][0]["base_commit"],
            expected,
        )
        self.assertEqual(
            self.profile("sympy-16886-goal-plus-pi-smoke")["tasks"][0][
                "base_commit"
            ],
            expected,
        )
        self.assertEqual(
            self.profile("sympy-16886-goal-plus-pi-luna-high-smoke")["tasks"][0][
                "base_commit"
            ],
            expected,
        )

    def test_goal_plus_profile_freezes_worker_and_closeout_budget(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-smoke")
        self.assertEqual(
            profile["goal_plus"],
            {
                "worker_runtime_seconds": 1500,
                "closeout_reserve_seconds": 300,
                "visible_verifier_timeout_seconds": 300,
                "evidence_annotator": "disabled",
            },
        )

        invalid = json.loads(json.dumps(profile))
        invalid["goal_plus"]["worker_runtime_seconds"] = 1700
        with self.assertRaisesRegex(SweBenchContractError, "must fit T"):
            validate_profile(str(invalid["id"]), invalid)

        invalid = json.loads(json.dumps(profile))
        invalid["goal_plus"]["evidence_annotator"] = "enabled"
        with self.assertRaisesRegex(SweBenchContractError, "disabled"):
            validate_profile(str(invalid["id"]), invalid)

    def test_codex_profile_freezes_custom_responses_auth(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        self.assertEqual(
            profile["agent_provider"],
            {
                "id": "bench_proxy",
                "name": "Benchmark OpenAI-compatible proxy",
                "auth_mode": "openai-compatible",
                "base_url_env": "OPENAI_BASE_URL",
                "api_key_env": "OPENAI_API_KEY",
                "wire_api": "responses",
            },
        )

        invalid = json.loads(json.dumps(profile))
        invalid["agent_provider"]["auth_mode"] = "oauth"
        with self.assertRaisesRegex(SweBenchContractError, "openai-compatible"):
            validate_profile(str(invalid["id"]), invalid)

    def test_codex_runtime_ignores_sforge_anthropic_environment(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        values = {
            "OPENAI_BASE_URL": "http://127.0.0.1:3788/v1",
            "OPENAI_API_KEY": "openai-secret",
            "SFORGE_AGENT_API_BASE_URL": "https://api.z.ai/api/anthropic",
            "SFORGE_AGENT_API_KEY": "anthropic-secret",
        }
        with mock.patch.dict(os.environ, values, clear=False):
            resolved = environment.resolve_codex_runtime(profile)

        self.assertEqual(resolved["base_url_env"], "OPENAI_BASE_URL")
        self.assertEqual(resolved["api_key_env"], "OPENAI_API_KEY")
        self.assertEqual(resolved["api_base_url"], values["OPENAI_BASE_URL"])
        self.assertNotIn("api_key", resolved)

    def test_luna_pi_profile_freezes_custom_responses_provider(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-luna-high-smoke")
        self.assertEqual(profile["model"], "bench-openai/gpt-5.6-luna")
        self.assertEqual(profile["reasoning_effort"], "high")
        self.assertEqual(
            profile["agent_provider"],
            {
                "id": "bench-openai",
                "name": "Benchmark OpenAI-compatible proxy",
                "auth_mode": "openai-compatible",
                "base_url_env": "OPENAI_BASE_URL",
                "api_key_env": "OPENAI_API_KEY",
                "wire_api": "responses",
            },
        )

        invalid = json.loads(json.dumps(profile))
        invalid["agent_provider"]["id"] = "other-provider"
        with self.assertRaisesRegex(SweBenchContractError, "must match"):
            validate_profile(str(invalid["id"]), invalid)

    def test_luna_pi_runtime_writes_environment_reference_not_secret(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-luna-high-smoke")
        secret = "not-for-provider-config"
        values = {
            "OPENAI_BASE_URL": "http://127.0.0.1:3788/v1",
            "OPENAI_API_KEY": secret,
        }
        with mock.patch.dict(os.environ, values, clear=False):
            runtime_info = environment.resolve_pi_runtime(profile)

        self.assertTrue(runtime_info["custom_provider"])
        self.assertEqual(runtime_info["provider"], "bench-openai")
        self.assertEqual(runtime_info["model_id"], "gpt-5.6-luna")
        self.assertEqual(runtime_info["credential_env"], "OPENAI_API_KEY")
        self.assertNotIn("api_key", runtime_info)

        runtime_info["runtime_api_base_url"] = "http://192.0.2.10:45678/v1"
        with self.temporary_directory() as temporary:
            models_file = Path(temporary) / "provider-runtime/models.json"
            environment.write_pi_models_config(
                models_file,
                runtime_info,
                reasoning_effort=profile["reasoning_effort"],
            )
            raw = models_file.read_text(encoding="utf-8")
            payload = json.loads(raw)

        provider = payload["providers"]["bench-openai"]
        self.assertEqual(provider["baseUrl"], "http://192.0.2.10:45678/v1")
        self.assertEqual(provider["api"], "openai-responses")
        self.assertEqual(provider["apiKey"], "$OPENAI_API_KEY")
        self.assertEqual(
            provider["models"][0]["thinkingLevelMap"], {"high": "high"}
        )
        self.assertNotIn(secret, raw)

    def test_loopback_bridge_preserves_the_responses_base_path(self) -> None:
        self.assertEqual(
            loopback_target("http://127.0.0.1:3788/v1"),
            ("127.0.0.1", 3788),
        )
        self.assertEqual(
            bridged_url("http://127.0.0.1:3788/v1", "192.0.2.10", 45678),
            "http://192.0.2.10:45678/v1",
        )

    def test_default_huggingface_cache_is_repository_local(self) -> None:
        expected = ensure_temp_root("huggingface")
        with mock.patch.dict(
            os.environ,
            {"XDG_CACHE_HOME": "/outside/cache"},
            clear=False,
        ):
            os.environ.pop("HF_HOME", None)
            observed = runtime._configure_huggingface_cache()

        self.assertEqual(observed, expected)

    def test_official_setup_commit_is_preserved_as_the_patch_base(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        base_commit = profile["tasks"][0]["base_commit"]
        observed_head = "5" * 40
        commands: list[list[str]] = []

        def docker_checked(command: list[str], *, timeout: int = 120) -> str:
            del timeout
            commands.append(command)
            if checkout.CHECKOUT_PROBE_SCRIPT in command:
                return self.checkout_probe_output(
                    base_commit,
                    observed_head=observed_head,
                )
            return ""

        with mock.patch.object(runtime, "_docker_checked", side_effect=docker_checked):
            checkout_info = runtime._initialize_agent_container(
                "container-id", profile, {}
            )

        self.assertTrue(checkout_info["synthetic_head"])
        self.assertEqual(checkout_info["observed_head"], observed_head)
        self.assertEqual(checkout_info["base_commit"], base_commit)
        self.assertEqual(checkout_info["baseline_kind"], "official-setup")
        self.assertTrue(checkout_info["official_setup_commit"])
        self.assertFalse(checkout_info["tree_matches_base"])
        self.assertEqual(checkout_info["patch_base_commit"], observed_head)
        self.assertFalse(any("reset" in command for command in commands))
        self.assertFalse(any("clean" in command for command in commands))

    def test_dataset_base_or_exact_official_setup_commit_is_required(self) -> None:
        base_commit = "1" * 40
        at_base = checkout.validate_checkout_probe(
            self.checkout_probe_output(
                base_commit,
                observed_head=base_commit,
                parents=["0" * 40],
                base_tree="2" * 40,
                observed_tree="2" * 40,
            ),
            base_commit,
        )
        self.assertTrue(at_base["passed"])
        self.assertEqual(at_base["patch_base_commit"], base_commit)

        cases = {
            "wrong parent": {"parents": ["3" * 40]},
            "wrong author": {"author_email": "other@example.com"},
            "wrong committer": {"committer_email": "other@example.com"},
            "wrong subject": {"subject": "not SWE-bench"},
            "dirty tree": {"clean": False},
        }
        for label, overrides in cases.items():
            with self.subTest(label=label):
                result = checkout.validate_checkout_probe(
                    self.checkout_probe_output(
                        base_commit,
                        observed_head="4" * 40,
                        **overrides,
                    ),
                    base_commit,
                )
                self.assertFalse(result["passed"])
                self.assertIsNone(result["patch_base_commit"])

    def test_plain_pi_c2_and_invalid_methods_resolve_before_prepare(self) -> None:
        agent = BenchmarkAgent(catalog=Catalog())
        with self.assertRaisesRegex(ContractError, "does not support.*goal-plus-codex"):
            agent.resolve_spec(
                target_ids=("swe-bench-verified",),
                profile="sympy-16886-codex-smoke",
                methods=("goal-plus-codex",),
                model="gpt-5.6-sol",
            )
        with self.assertRaisesRegex(ContractError, "PROVIDER/MODEL"):
            agent.resolve_spec(
                target_ids=("swe-bench-verified",),
                profile="sympy-16886-pi-smoke",
                methods=("plain-pi",),
                model="glm-5.2",
            )
        pi_c2 = agent.resolve_spec(
            target_ids=("swe-bench-verified",),
            profile="sympy-16886-19346-pi-glm-5-2-c2",
            methods=("plain-pi",),
            model="zai/glm-5.2",
            wall_time_seconds=1800,
            live_search_concurrency=1,
            cell_concurrency=2,
        )
        self.assertEqual(
            pi_c2.concurrency(), {"T": 1800, "K": 1, "C": 2, "R": 1}
        )

        with self.assertRaisesRegex(ContractError, "supports at most C=1"):
            agent.resolve_spec(
                target_ids=("swe-bench-verified",),
                profile="sympy-16886-goal-plus-pi-smoke",
                methods=("goal-plus-pi",),
                model="zai/glm-5.2",
                cell_concurrency=2,
            )

    def test_native_plan_uses_doctor_prepare_and_detached_run(self) -> None:
        agent = BenchmarkAgent(catalog=Catalog())
        spec = agent.resolve_spec(
            preset_id="swe-bench-verified-sympy-16886-codex-smoke",
            retain_containers=True,
        )
        runner = create_runner(spec.runner)

        doctor = runner.provision_commands(spec, skip_provision=True)
        prepare, campaign = runner.prepare_commands(spec)
        run = runner.start_command(spec, campaign, detach=True)

        self.assertEqual(len(doctor), 1)
        self.assertIn("doctor", doctor[0])
        self.assertIn("--method", doctor[0])
        self.assertIn("prepare", prepare[0])
        self.assertIn("--wall-time-seconds", prepare[0])
        self.assertIn("--retain-containers", prepare[0])
        self.assertTrue(spec.as_dict()["retain_containers"])
        self.assertEqual(run[-3:-1], ["--campaign", campaign.campaign_id])
        self.assertEqual(run[-1], "--detach")

    def test_unproven_runner_rejects_retained_containers_during_plan(self) -> None:
        agent = BenchmarkAgent(catalog=Catalog())
        with self.assertRaisesRegex(ContractError, "does not support retained"):
            agent.resolve_spec(
                preset_id="edgebench-vliw-codex-local-smoke",
                retain_containers=True,
            )

    def test_prepare_separates_agent_task_from_hidden_evaluator_fields(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        instance = {
            "instance_id": "sympy__sympy-16886",
            "repo": "sympy/sympy",
            "base_commit": profile["tasks"][0]["base_commit"],
            "problem_statement": "Public issue text",
            "version": "1.5",
            "patch": "gold patch",
            "test_patch": "hidden tests",
            "FAIL_TO_PASS": ["hidden-fail"],
            "PASS_TO_PASS": ["hidden-pass"],
        }
        with self.temporary_directory() as temporary:
            campaign = Path(temporary) / "campaign"

            def git_value(_path: Path, *args: str) -> str:
                return "" if args[:2] == ("status", "--porcelain") else "a" * 40

            with (
                mock.patch.object(runtime, "campaign_dir", return_value=campaign),
                mock.patch.object(runtime, "preserve_conflict", return_value=None),
                mock.patch.object(runtime, "_load_pinned_instance", return_value=instance),
                mock.patch.object(runtime, "_validate_instance_image"),
                mock.patch.object(runtime, "_git_value", side_effect=git_value),
            ):
                runtime.prepare("test-campaign", profile)

            manifest = json.loads((campaign / "campaign.json").read_text())
            controller = read_json(campaign / "controller.json")
            task = json.loads((campaign / manifest["cells"][0]["task_file"]).read_text())
            evaluator_path = campaign / manifest["dataset"]["evaluator_instances_file"]
            evaluator = json.loads(evaluator_path.read_text())

            self.assertTrue(runtime.HIDDEN_INSTANCE_FIELDS.isdisjoint(task))
            self.assertEqual(task["problem_statement"], "Public issue text")
            self.assertEqual(len(evaluator), 1)
            self.assertEqual(evaluator[0]["patch"], "gold patch")
            self.assertEqual(evaluator[0]["test_patch"], "hidden tests")
            self.assertEqual(stat.S_IMODE(evaluator_path.stat().st_mode), 0o600)
            self.assertFalse(manifest["container_retention"]["requested"])
            self.assertEqual(manifest["container_retention"]["scope"], "agent")
            self.assertEqual(controller["state"], "prepared")
            self.assertIsNone(controller["pid"])
            self.assertEqual(controller["log_file"], "controller.log")

            from swebench.harness.utils import load_swebench_dataset

            loaded = load_swebench_dataset(str(evaluator_path), "test")
            self.assertEqual(len(loaded), 1)
            self.assertEqual(loaded[0]["instance_id"], instance["instance_id"])

    def test_detached_launch_records_controller_without_credentials(self) -> None:
        secret = "must-not-enter-controller-json"
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            write_json(
                campaign / "campaign.json",
                {
                    "schema_version": 1,
                    "campaign_id": "detached-test",
                    "benchmark_id": "swe-bench-verified",
                    "state": "prepared",
                    "methods": ["plain-codex"],
                    "model": "gpt-5.6-luna",
                    "budget": {"cell_concurrency": 1},
                    "cells": [],
                },
            )
            write_json(
                campaign / "controller.json",
                {
                    "schema_version": 1,
                    "campaign_id": "detached-test",
                    "state": "prepared",
                    "pid": None,
                    "pgid": None,
                },
            )
            child = mock.Mock(pid=4242)
            with (
                mock.patch.dict(os.environ, {"OPENAI_API_KEY": secret}, clear=False),
                mock.patch.object(runtime.subprocess, "Popen", return_value=child) as popen,
            ):
                self.assertEqual(runtime.launch(campaign, detach=True), 0)

            saved = read_json(campaign / "controller.json")
            raw = (campaign / "controller.json").read_text(encoding="utf-8")
            self.assertEqual(saved["state"], "launching")
            self.assertEqual(saved["pid"], 4242)
            self.assertEqual(saved["pgid"], 4242)
            self.assertEqual(saved["command"][-2:], ["--campaign", "detached-test"])
            self.assertNotIn(secret, raw)
            launch_kwargs = popen.call_args.kwargs
            self.assertTrue(launch_kwargs["start_new_session"])
            self.assertIs(launch_kwargs["stdin"], subprocess.DEVNULL)
            self.assertEqual(launch_kwargs["env"]["OPENAI_API_KEY"], secret)

    def test_detached_launch_rejects_a_live_controller(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            write_json(
                campaign / "campaign.json",
                {
                    "schema_version": 1,
                    "campaign_id": "already-running",
                    "benchmark_id": "swe-bench-verified",
                    "state": "prepared",
                    "methods": ["plain-codex"],
                    "model": "gpt-5.6-luna",
                    "budget": {"cell_concurrency": 1},
                    "cells": [],
                },
            )
            write_json(
                campaign / "controller.json",
                {
                    "schema_version": 1,
                    "campaign_id": "already-running",
                    "state": "launching",
                    "pid": 4242,
                },
            )
            with (
                mock.patch.object(runtime, "process_alive", return_value=True),
                mock.patch.object(runtime.subprocess, "Popen") as popen,
                self.assertRaisesRegex(SweBenchContractError, "already running"),
            ):
                runtime.launch(campaign, detach=True)
            popen.assert_not_called()

    def test_two_task_prepare_isolates_cells_and_hidden_evaluator_data(self) -> None:
        profile = self.profile("sympy-16886-19346-codex-terra-c2")
        instances = [
            {
                "instance_id": task["instance_id"],
                "repo": task["repo"],
                "base_commit": task["base_commit"],
                "problem_statement": f"Issue for {task['instance_id']}",
                "version": "1",
                "patch": f"hidden patch {index}",
                "test_patch": f"hidden test {index}",
                "FAIL_TO_PASS": [f"hidden-fail-{index}"],
                "PASS_TO_PASS": [f"hidden-pass-{index}"],
            }
            for index, task in enumerate(profile["tasks"])
        ]
        with self.temporary_directory() as temporary:
            campaign = Path(temporary) / "campaign"

            def git_value(_path: Path, *args: str) -> str:
                return "" if args[:2] == ("status", "--porcelain") else "a" * 40

            with (
                mock.patch.object(runtime, "campaign_dir", return_value=campaign),
                mock.patch.object(runtime, "preserve_conflict", return_value=None),
                mock.patch.object(
                    runtime, "_load_pinned_instances", return_value=instances
                ) as load_instances,
                mock.patch.object(runtime, "_validate_instance_image") as validate,
                mock.patch.object(runtime, "_git_value", side_effect=git_value),
            ):
                runtime.prepare("two-task-campaign", profile)

            manifest = read_json(campaign / "campaign.json")
            self.assertEqual(manifest["budget"]["cell_concurrency"], 2)
            self.assertEqual(len(manifest["cells"]), 2)
            self.assertEqual(load_instances.call_count, 1)
            self.assertEqual(validate.call_count, 2)
            self.assertEqual(
                [cell["task_id"] for cell in manifest["cells"]],
                profile["task_ids"],
            )
            self.assertEqual(
                [cell["image"] for cell in manifest["cells"]],
                [task["image"] for task in profile["tasks"]],
            )
            self.assertEqual(
                len({cell["cell_file"] for cell in manifest["cells"]}), 2
            )
            self.assertEqual(
                len({cell["evaluator_dir"] for cell in manifest["cells"]}), 2
            )
            for cell in manifest["cells"]:
                task = read_json(campaign / cell["task_file"])
                durable = read_json(campaign / cell["cell_file"])
                self.assertTrue(runtime.HIDDEN_INSTANCE_FIELDS.isdisjoint(task))
                self.assertEqual(durable["cell_id"], cell["cell_id"])
            evaluator_path = campaign / manifest["dataset"][
                "evaluator_instances_file"
            ]
            self.assertEqual(len(json.loads(evaluator_path.read_text())), 2)
            self.assertEqual(stat.S_IMODE(evaluator_path.stat().st_mode), 0o600)

    def test_cell_specific_container_uses_its_image_and_base_commit(self) -> None:
        profile = self.profile("sympy-16886-19346-codex-terra-c2")
        cell = {
            "cell_id": "plain-codex--sympy__sympy-19346",
            **profile["tasks"][1],
        }
        runtime_info = {
            "archive": Path("/runtime/codex.tgz"),
            "archive_present": True,
            "credential_present": True,
            "api_base_url": "http://127.0.0.1:3788/v1",
            "runtime_api_base_url": "http://192.0.2.1:3788/v1",
        }
        commands: list[list[str]] = []
        observed_head = "e" * 40

        def docker_checked(command: list[str], *, timeout: int = 120) -> str:
            del timeout
            commands.append(command)
            if command[:2] == ["docker", "create"]:
                return "container-id"
            if checkout.CHECKOUT_PROBE_SCRIPT in command:
                return self.checkout_probe_output(
                    cell["base_commit"],
                    observed_head=observed_head,
                )
            return ""

        with mock.patch.object(runtime, "_docker_checked", side_effect=docker_checked):
            runtime._create_agent_container(
                "campaign", profile, runtime_info, cell=cell
            )
            checkout_info = runtime._initialize_agent_container(
                "container-id", profile, runtime_info, cell=cell
            )

        create = next(command for command in commands if command[:2] == ["docker", "create"])
        self.assertEqual(create[-3], cell["image"])
        self.assertIn(f"bench-goal-plus.cell={cell['cell_id']}", create)
        self.assertEqual(checkout_info["base_commit"], cell["base_commit"])
        self.assertEqual(checkout_info["patch_base_commit"], observed_head)
        probe = next(
            command for command in commands if checkout.CHECKOUT_PROBE_SCRIPT in command
        )
        self.assertEqual(probe[-1], cell["base_commit"])
        self.assertFalse(any("reset" in command for command in commands))
        self.assertFalse(any("clean" in command for command in commands))

    def test_plain_pi_c2_cells_use_unique_containers_and_pi_homes(self) -> None:
        profile = self.profile("sympy-16886-19346-pi-glm-5-2-c2")
        runtime_info = {
            "credential_present": True,
            "credential_env": "ZAI_API_KEY",
            "provider": "zai",
            "model_id": "glm-5.2",
            "node_root": Path("/opt/node-host"),
            "package_root": Path("/opt/pi-host"),
        }
        commands: list[list[str]] = []

        def docker_checked(command: list[str], *, timeout: int = 120) -> str:
            del timeout
            commands.append(command)
            if command[:2] == ["docker", "create"]:
                return f"container-{len(commands)}"
            return ""

        with mock.patch.object(runtime, "_docker_checked", side_effect=docker_checked):
            for task in profile["tasks"]:
                runtime._create_agent_container(
                    "pi-c2",
                    profile,
                    runtime_info,
                    cell={
                        "cell_id": f"plain-pi--{task['instance_id']}",
                        **task,
                    },
                )

        creates = [command for command in commands if command[:2] == ["docker", "create"]]
        self.assertEqual(len(creates), 2)
        self.assertEqual(len({command[command.index("--name") + 1] for command in creates}), 2)
        self.assertEqual({command[-3] for command in creates}, {task["image"] for task in profile["tasks"]})
        for command in creates:
            self.assertIn("/opt/pi-home:rw,nosuid,nodev,size=128m", command)

    def test_prompt_rejects_hidden_fields(self) -> None:
        with self.assertRaisesRegex(SweBenchContractError, "hidden fields"):
            runtime.build_agent_prompt(
                {"problem_statement": "issue", "test_patch": "do not expose"}
            )

    def test_pi_commands_inherit_only_the_credential_variable_name(self) -> None:
        profile = self.profile("sympy-16886-pi-smoke")
        secret = "not-for-command-lines"
        runtime_info = {
            "credential_env": "ZAI_API_KEY",
            "provider": "zai",
            "model_id": "glm-5.2",
            "node_root": Path("/opt/node-host"),
            "package_root": Path("/opt/pi-host"),
        }
        with mock.patch.dict(os.environ, {"ZAI_API_KEY": secret}, clear=False):
            command = runtime._agent_command("container-id", profile, runtime_info)

        self.assertIn("ZAI_API_KEY", command)
        self.assertFalse(any(secret in argument for argument in command))
        self.assertNotIn(f"ZAI_API_KEY={secret}", command)

        completed = subprocess.CompletedProcess([], 0, "glm-5.2", "")
        with mock.patch.object(environment, "run_capture", return_value=completed) as capture:
            environment._pi_container_probe(profile["tasks"][0]["image"], runtime_info)
        probe = capture.call_args.args[0]
        self.assertIn("ZAI_API_KEY", probe)
        self.assertFalse(any(secret in argument for argument in probe))

    def test_custom_pi_container_probe_mounts_generated_provider_config(self) -> None:
        secret = "not-for-command-lines"
        with self.temporary_directory() as temporary:
            models_file = Path(temporary) / "provider-runtime/models.json"
            models_file.parent.mkdir(parents=True)
            models_file.write_text("{\"providers\": {}}\n", encoding="utf-8")
            runtime_info = {
                "credential_env": "OPENAI_API_KEY",
                "provider": "bench-openai",
                "model_id": "gpt-5.6-luna",
                "node_root": Path("/opt/node-host"),
                "package_root": Path("/opt/pi-host"),
                "models_file": models_file,
                "bridge_host": "192.0.2.10",
            }
            completed = subprocess.CompletedProcess([], 0, "gpt-5.6-luna", "")
            with (
                mock.patch.dict(
                    os.environ, {"OPENAI_API_KEY": secret}, clear=False
                ),
                mock.patch.object(
                    environment, "run_capture", return_value=completed
                ) as capture,
            ):
                environment._pi_container_probe("image:latest", runtime_info)

        command = capture.call_args.args[0]
        joined = " ".join(str(item) for item in command)
        self.assertEqual(command[:4], ["docker", "run", "--pull", "never"])
        self.assertIn("OPENAI_API_KEY", command)
        self.assertIn("dst=/opt/provider,readonly", joined)
        self.assertIn("PI_CODING_AGENT_DIR=/opt/pi-home/.pi/agent", command)
        self.assertIn("NO_PROXY=192.0.2.10", command)
        self.assertNotIn(secret, joined)

    def test_goal_plus_container_mounts_and_outer_pi_command_are_explicit(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-smoke")
        secret = "not-for-command-lines"
        with self.temporary_directory() as temporary:
            assets = Path(temporary)
            goal_plus_root = assets / "goal-plus"
            goal_plus_root.mkdir()
            dependency_lock = assets / "requirements.lock"
            dependency_lock.write_text("pydantic==2.13.4\n", encoding="utf-8")
            verifier = assets / "visible.py"
            verifier.write_text("print('{}')\n", encoding="utf-8")
            controller = assets / "controller.py"
            controller.write_text("print('{}')\n", encoding="utf-8")
            pip_cache = assets / "pip-cache"
            pip_cache.mkdir()
            runtime_info = {
                "credential_env": "ZAI_API_KEY",
                "credential_present": True,
                "provider": "zai",
                "model_id": "glm-5.2",
                "node_root": Path("/host/node"),
                "package_root": Path("/host/pi"),
                "goal_plus_root": goal_plus_root,
                "goal_plus_dependency_lock": dependency_lock,
                "goal_plus_visible_verifier": verifier,
                "goal_plus_controller": controller,
                "goal_plus_pip_cache": pip_cache,
            }
            docker_commands: list[list[str]] = []

            def docker_checked(command: list[str], *, timeout: int = 120) -> str:
                del timeout
                docker_commands.append(command)
                return "container-id" if command[:2] == ["docker", "create"] else ""

            with mock.patch.object(
                runtime, "_docker_checked", side_effect=docker_checked
            ):
                runtime._create_agent_container(
                    "goal-plus-campaign", profile, runtime_info
                )

            create = docker_commands[0]
            joined = " ".join(create)
            self.assertEqual(create[:4], ["docker", "create", "--pull", "never"])
            self.assertIn("dst=/opt/goal-plus,readonly", joined)
            self.assertIn("dst=/opt/pi,readonly", joined)
            self.assertIn("dst=/opt/node,readonly", joined)
            self.assertIn("dst=/opt/pip-cache", joined)
            self.assertNotIn("dst=/opt/pip-cache,readonly", joined)
            self.assertIn(
                "/opt/goal-plus-runtime:rw,exec,nosuid,nodev,size=512m", create
            )

            prompt = runtime.build_goal_plus_prompt(
                {"problem_statement": "Public issue text"}, profile
            )
            runtime_info.update(
                {
                    "outer_deadline_at": "2026-08-03T12:00:00+00:00",
                    "main_session_id": "swe-bench-main-test",
                    "goal_prompt": prompt,
                }
            )
            with mock.patch.dict(os.environ, {"ZAI_API_KEY": secret}, clear=False):
                command = runtime._agent_command(
                    "container-id", profile, runtime_info
                )

            self.assertEqual(command[-1], prompt)
            self.assertTrue(prompt.startswith("/goal-plus mode=autonomous"))
            self.assertIn("budget.max_parallel=1", prompt)
            self.assertIn("do not set the deprecated max_candidates", prompt)
            self.assertIn("Public issue text", prompt)
            self.assertIn("/opt/goal-plus/.pi/extensions/goal-plus.ts", command)
            self.assertIn("/opt/goal-plus/.pi/skills/goal-plus/SKILL.md", command)
            self.assertIn("GOAL_PLUS_ROOT=/testbed/.gp", command)
            self.assertIn("GOAL_PLUS_EVIDENCE_ANNOTATOR_DISABLED=1", command)
            self.assertIn(
                'export PATH=/opt/goal-plus-bin:/opt/node/bin:$PATH; exec "$@"',
                command,
            )
            self.assertIn("ZAI_API_KEY", command)
            self.assertFalse(any(secret in argument for argument in command))

    def test_goal_plus_durable_evidence_enforces_k_and_verifier_contract(self) -> None:
        with self.temporary_directory() as temporary:
            root = Path(temporary) / "passing"
            self.write_goal_plus_state(root)
            state = goal_plus_evidence.collect_goal_plus_state(
                root,
                expected_k=1,
                expected_worker_runtime_seconds=1500,
                expected_closeout_reserve_seconds=300,
                expected_visible_verifier_timeout_seconds=300,
            )

            self.assertTrue(state["completion"]["passed"])
            self.assertEqual(state["actual_subagent_count"], 1)
            self.assertEqual(state["worker_usage"]["input_tokens"], 10)
            self.assertEqual(state["worker_usage"]["sessions_covered"], 1)
            visible = state["completion"]["checks"]["visible_verifiers"]["actual"]
            self.assertTrue(visible["process"]["passed"])
            self.assertTrue(visible["promotion"]["passed"])
            self.assertEqual(
                visible["promotion"]["verifiers"][0]["role"], "promotion_gate"
            )

            mismatched = Path(temporary) / "mismatched"
            self.write_goal_plus_state(mismatched, session_count=0)
            mismatch = goal_plus_evidence.collect_goal_plus_state(
                mismatched,
                expected_k=1,
                expected_worker_runtime_seconds=1500,
                expected_closeout_reserve_seconds=300,
                expected_visible_verifier_timeout_seconds=300,
            )
            self.assertFalse(mismatch["completion"]["passed"])
            self.assertEqual(mismatch["actual_subagent_count"], 0)
            self.assertIn(
                "bound_pi_worker_sessions", mismatch["completion"]["reason"]
            )

    def test_codex_runtime_probe_and_agent_use_the_same_bounded_tmpfs(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        runtime_info = {
            "archive": Path("/runtime/codex.tgz"),
            "archive_present": True,
            "provider_id": "bench_proxy",
            "provider_name": "Benchmark OpenAI-compatible proxy",
            "auth_mode": "openai-compatible",
            "wire_api": "responses",
            "base_url_env": "OPENAI_BASE_URL",
            "api_key_env": "OPENAI_API_KEY",
            "api_base_url": "http://127.0.0.1:3788/v1",
            "runtime_api_base_url": "http://192.0.2.10:45678/v1",
            "credential_present": True,
            "bridge_host": "192.0.2.10",
            "bridge": None,
        }
        completed = subprocess.CompletedProcess([], 0, "codex-cli 0.144.1", "")
        with mock.patch.object(environment, "run_capture", return_value=completed) as capture:
            environment._codex_container_probe(
                profile["tasks"][0]["image"], runtime_info["archive"]
            )
        probe = capture.call_args.args[0]

        docker_commands: list[list[str]] = []

        def docker_checked(command: list[str], *, timeout: int = 120) -> str:
            del timeout
            docker_commands.append(command)
            return "container-id"

        with (
            mock.patch.object(runtime, "resolve_codex_runtime", return_value=runtime_info),
            mock.patch.object(runtime, "_docker_checked", side_effect=docker_checked),
        ):
            runtime._create_agent_container("campaign", profile)
        create = docker_commands[0]

        self.assertIn(environment.CODEX_RUNTIME_TMPFS, probe)
        self.assertIn(environment.CODEX_RUNTIME_TMPFS, create)
        self.assertIn(":rw,exec,nosuid,nodev,", environment.CODEX_RUNTIME_TMPFS)
        self.assertEqual(environment.CODEX_RUNTIME_TMPFS.rsplit("=", 1)[1], "512m")
        self.assertNotIn("/opt/host-auth", " ".join(create))

        secret = "not-for-command-lines"
        with mock.patch.dict(os.environ, {"OPENAI_API_KEY": secret}, clear=False):
            agent_command = runtime._agent_command(
                "container-id", profile, runtime_info
            )
        joined = " ".join(agent_command)
        self.assertIn('model_provider="bench_proxy"', joined)
        self.assertIn('wire_api="responses"', joined)
        self.assertIn("http://192.0.2.10:45678/v1", joined)
        self.assertIn("OPENAI_API_KEY", agent_command)
        self.assertNotIn(secret, joined)
        self.assertNotIn("chatgpt.com", joined)

    def test_container_responses_probe_inherits_key_by_name(self) -> None:
        secret = "not-for-command-lines"
        runtime_info = {
            "runtime_api_base_url": "http://192.0.2.10:45678/v1",
            "api_key_env": "OPENAI_API_KEY",
        }
        completed = subprocess.CompletedProcess(
            [],
            0,
            json.dumps(
                {
                    "passed": True,
                    "http_status": 200,
                    "object": "response",
                    "model": "gpt-5.6-sol",
                    "response_status": "completed",
                }
            ),
            "",
        )
        with (
            mock.patch.dict(os.environ, {"OPENAI_API_KEY": secret}, clear=False),
            mock.patch.object(
                environment, "run_capture", return_value=completed
            ) as capture,
        ):
            result = environment.codex_container_responses_probe(
                "image:latest", runtime_info, model="gpt-5.6-sol"
            )

        command = capture.call_args.args[0]
        self.assertTrue(result["passed"])
        self.assertEqual(command[:4], ["docker", "run", "--pull", "never"])
        self.assertIn("OPENAI_API_KEY", command)
        self.assertFalse(any(secret in argument for argument in command))
        self.assertNotIn(f"OPENAI_API_KEY={secret}", command)

    def test_timeout_stops_container_before_exporting_patch(self) -> None:
        profile = self.profile("sympy-16886-pi-smoke")
        patch_base_commit = "d" * 40
        runtime_info = {
            "credential_env": "ZAI_API_KEY",
            "provider": "zai",
            "node_root": Path("/node"),
            "package_root": Path("/pi"),
        }
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            cell_dir = campaign / "cells/plain-pi"
            cell_dir.mkdir(parents=True)
            write_json(
                cell_dir / "task.json",
                {"problem_statement": "issue"},
            )
            cell = {
                "task_file": "cells/plain-pi/task.json",
                "patch_file": "cells/plain-pi/model.patch",
                "method": "plain-pi",
                "model": profile["model"],
            }
            manifest = {"campaign_id": "timeout-test", "profile_snapshot": profile}
            docker_commands: list[list[str]] = []

            def docker_checked(command: list[str], *, timeout: int = 120) -> str:
                docker_commands.append(command)
                if "diff" in command:
                    return "diff --git a/a b/a\n"
                if "status" in command:
                    return " M a"
                return "container-id"

            def run_command(command: list[str], **_kwargs):
                if command[:3] == ["docker", "rm", "-f"]:
                    return subprocess.CompletedProcess(command, 0, "", "")
                raise subprocess.TimeoutExpired(command, 300, output="partial output")

            with (
                mock.patch.object(
                    runtime,
                    "_create_agent_container",
                    return_value=("container-id", runtime_info),
                ),
                mock.patch.object(
                    runtime,
                    "_initialize_agent_container",
                    return_value={"patch_base_commit": patch_base_commit},
                ),
                mock.patch.object(
                    runtime, "_agent_command", return_value=["docker", "exec", "container-id"]
                ),
                mock.patch.object(runtime, "_docker_checked", side_effect=docker_checked),
                mock.patch.object(runtime, "_run", side_effect=run_command),
            ):
                result = runtime._run_agent(campaign, manifest, cell)

            stop_index = next(i for i, command in enumerate(docker_commands) if "stop" in command)
            diff_index = next(i for i, command in enumerate(docker_commands) if "diff" in command)
            self.assertLess(stop_index, diff_index)
            self.assertEqual(docker_commands[diff_index][-1], patch_base_commit)
            self.assertTrue(result["timed_out"])
            self.assertTrue(result["patch_exists"])
            self.assertTrue(result["container"]["cleanup"]["removed"])
            self.assertIsNotNone(result["runtime_seconds"])
            self.assertIsNotNone(result["finalization_grace_seconds"])

    def test_goal_plus_state_is_exported_before_container_disposal(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-smoke")
        runtime_info = {
            "credential_env": "ZAI_API_KEY",
            "provider": "zai",
            "node_root": Path("/node"),
            "package_root": Path("/pi"),
            "goal_plus_root": Path("/goal-plus"),
            "goal_plus_dependency_lock": Path("/requirements.lock"),
            "goal_plus_visible_verifier": Path("/visible.py"),
            "goal_plus_controller": Path("/controller.py"),
            "goal_plus_pip_cache": Path("/pip-cache"),
        }
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            cell_dir = campaign / "cells/goal-plus-pi"
            cell_dir.mkdir(parents=True)
            write_json(cell_dir / "task.json", {"problem_statement": "issue"})
            cell = {
                "task_file": "cells/goal-plus-pi/task.json",
                "patch_file": "cells/goal-plus-pi/model.patch",
                "method": "goal-plus-pi",
                "model": profile["model"],
            }
            manifest = {
                "campaign_id": "goal-plus-export-test",
                "profile_snapshot": profile,
                "source": {"goal_plus_commit": "a" * 40},
            }
            sequence: list[str] = []

            def docker_checked(command: list[str], *, timeout: int = 120) -> str:
                del timeout
                if "diff" in command:
                    return "diff --git a/a b/a\n"
                if "status" in command:
                    return " M a"
                return ""

            def export_state(*_args, **_kwargs):
                sequence.append("export")
                return {
                    "actual_subagent_count": 1,
                    "completion": {"passed": True, "reason": None},
                    "worker_usage": {"coverage": "persisted_pi_worker_usage"},
                }

            def dispose(*_args, **_kwargs):
                sequence.append("dispose")
                return {
                    "attempted": True,
                    "removed": True,
                    "retained": False,
                    "stopped": None,
                }

            with (
                mock.patch.object(
                    runtime, "resolve_goal_plus_runtime", return_value=runtime_info
                ),
                mock.patch.object(
                    runtime,
                    "_create_agent_container",
                    return_value=("container-id", runtime_info),
                ),
                mock.patch.object(
                    runtime,
                    "_initialize_agent_container",
                    return_value={"patch_base_commit": "d" * 40},
                ),
                mock.patch.object(runtime, "_agent_command", return_value=["outer"]),
                mock.patch.object(
                    runtime,
                    "_run",
                    return_value=subprocess.CompletedProcess(["outer"], 0, "", ""),
                ),
                mock.patch.object(
                    runtime,
                    "_goal_plus_closeout",
                    return_value={"completed": True},
                ),
                mock.patch.object(
                    runtime, "_export_goal_plus_state", side_effect=export_state
                ),
                mock.patch.object(
                    runtime, "_dispose_agent_container", side_effect=dispose
                ),
                mock.patch.object(
                    runtime, "_docker_checked", side_effect=docker_checked
                ),
            ):
                result = runtime._run_agent(campaign, manifest, cell)

            self.assertEqual(sequence, ["export", "dispose"])
            self.assertEqual(result["state"], "completed")
            self.assertTrue(result["patch_exists"])
            self.assertEqual(result["goal_plus"]["actual_subagent_count"], 1)

    def test_execute_campaign_runs_two_task_cells_concurrently(self) -> None:
        profile = self.profile("sympy-16886-19346-codex-terra-c2")
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            cells = []
            for task in profile["tasks"]:
                cell_id = f"plain-codex--{task['instance_id']}"
                cell = {
                    "cell_id": cell_id,
                    "cell_file": f"cells/{cell_id}/cell.json",
                    "task_id": task["instance_id"],
                    "method": "plain-codex",
                    "state": "prepared",
                    "evaluation": {"state": "pending", "calls": 0},
                }
                cells.append(cell)
                write_json(campaign / cell["cell_file"], cell)
            manifest = {
                "schema_version": 1,
                "campaign_id": "two-cell-concurrency",
                "benchmark_id": "swe-bench-verified",
                "state": "prepared",
                "methods": ["plain-codex"],
                "model": profile["model"],
                "profile_snapshot": profile,
                "budget": {"wall_time_seconds": 1800, "cell_concurrency": 2},
                "cells": cells,
            }
            write_json(campaign / "campaign.json", manifest)

            barrier = threading.Barrier(2)
            lock = threading.Lock()
            active = 0
            max_active = 0

            def run_agent(_campaign, _manifest, cell, **_kwargs):
                nonlocal active, max_active
                with lock:
                    active += 1
                    max_active = max(max_active, active)
                try:
                    barrier.wait(timeout=5)
                    return {
                        "state": "completed",
                        "patch_exists": True,
                        "container": {
                            "id": f"container-{cell['task_id']}",
                            "name": runtime._container_name(
                                "two-cell-concurrency", cell["cell_id"]
                            ),
                            "cleanup": {"removed": True, "retained": False},
                        },
                    }
                finally:
                    with lock:
                        active -= 1

            def evaluate(_campaign, _manifest, cell, **_kwargs):
                return {
                    "state": "completed",
                    "calls": 1,
                    "resolved": False,
                    "patch_applied": True,
                    "task_id": cell["task_id"],
                }

            with (
                mock.patch.object(runtime, "_run_agent", side_effect=run_agent),
                mock.patch.object(
                    runtime, "_official_evaluation", side_effect=evaluate
                ) as official,
            ):
                exit_code = runtime.execute_campaign(campaign)

            saved = read_json(campaign / "campaign.json")
            controller = read_json(campaign / "controller.json")
            self.assertEqual(exit_code, 0)
            self.assertEqual(saved["state"], "completed")
            self.assertEqual(controller["state"], "completed")
            self.assertEqual(controller["returncode"], 0)
            self.assertEqual(max_active, 2)
            self.assertEqual(saved["scheduler"]["max_active_observed"], 2)
            self.assertEqual({cell["state"] for cell in saved["cells"]}, {"completed"})
            self.assertEqual(official.call_count, 2)
            self.assertEqual(
                len(
                    {
                        cell["agent"]["container"]["name"]
                        for cell in saved["cells"]
                    }
                ),
                2,
            )

            with mock.patch.object(runtime, "process_alive", return_value=False):
                status = runtime.status_payload(campaign)
            self.assertEqual(status["controller"]["state"], "completed")
            self.assertFalse(status["controller"]["alive"])

    def test_one_cell_failure_preserves_the_other_completed_score(self) -> None:
        profile = self.profile("sympy-16886-19346-codex-terra-c2")
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            cells = []
            for task in profile["tasks"]:
                cell_id = f"plain-codex--{task['instance_id']}"
                cell = {
                    "cell_id": cell_id,
                    "cell_file": f"cells/{cell_id}/cell.json",
                    "task_id": task["instance_id"],
                    "method": "plain-codex",
                    "state": "prepared",
                    "evaluation": {"state": "pending", "calls": 0},
                }
                cells.append(cell)
                write_json(campaign / cell["cell_file"], cell)
            write_json(
                campaign / "campaign.json",
                {
                    "schema_version": 1,
                    "campaign_id": "one-cell-fails",
                    "benchmark_id": "swe-bench-verified",
                    "state": "prepared",
                    "methods": ["plain-codex"],
                    "model": profile["model"],
                    "profile_snapshot": profile,
                    "budget": {"wall_time_seconds": 1800, "cell_concurrency": 2},
                    "cells": cells,
                },
            )

            def run_agent(_campaign, _manifest, cell, **_kwargs):
                if cell["task_id"].endswith("19346"):
                    raise RuntimeError("synthetic task failure")
                return {
                    "state": "completed",
                    "patch_exists": True,
                    "container": {
                        "cleanup": {"removed": True, "retained": False}
                    },
                }

            with (
                mock.patch.object(runtime, "_run_agent", side_effect=run_agent),
                mock.patch.object(
                    runtime,
                    "_official_evaluation",
                    return_value={
                        "state": "completed",
                        "calls": 1,
                        "resolved": True,
                        "patch_applied": True,
                    },
                ),
            ):
                exit_code = runtime.execute_campaign(campaign)

            saved = read_json(campaign / "campaign.json")
            by_task = {cell["task_id"]: cell for cell in saved["cells"]}
            self.assertEqual(exit_code, 1)
            self.assertEqual(saved["state"], "partial")
            self.assertEqual(by_task["sympy__sympy-16886"]["state"], "completed")
            self.assertTrue(
                by_task["sympy__sympy-16886"]["evaluation"]["resolved"]
            )
            self.assertEqual(by_task["sympy__sympy-19346"]["state"], "failed")
            self.assertIn("synthetic task failure", by_task["sympy__sympy-19346"]["error"])

    def test_unconfirmed_agent_cleanup_blocks_official_evaluator(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            manifest = {
                "schema_version": 1,
                "campaign_id": "cleanup-test",
                "benchmark_id": "swe-bench-verified",
                "state": "prepared",
                "methods": ["plain-codex"],
                "model": "gpt-5.6-sol",
                "budget": {"wall_time_seconds": 300},
                "cells": [
                    {
                        "cell_id": "cell",
                        "task_id": "task",
                        "state": "prepared",
                        "evaluation": {"state": "pending", "calls": 0},
                    }
                ],
            }
            write_json(campaign / "campaign.json", manifest)
            agent_result = {
                "patch_exists": True,
                "container": {"cleanup": {"attempted": True, "removed": False}},
            }
            with (
                mock.patch.object(runtime, "_run_agent", return_value=agent_result),
                mock.patch.object(runtime, "_official_evaluation") as official,
            ):
                exit_code = runtime.execute_campaign(campaign)

            saved = json.loads((campaign / "campaign.json").read_text())
            self.assertEqual(exit_code, 1)
            self.assertEqual(saved["state"], "failed")
            self.assertEqual(saved["cells"][0]["evaluation"]["calls"], 0)
            official.assert_not_called()

    def test_stopped_retained_agent_allows_separate_official_evaluator(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            manifest = {
                "schema_version": 1,
                "campaign_id": "retention-test",
                "benchmark_id": "swe-bench-verified",
                "state": "prepared",
                "methods": ["plain-codex"],
                "model": "gpt-5.6-sol",
                "budget": {"wall_time_seconds": 300},
                "cells": [
                    {
                        "cell_id": "cell",
                        "task_id": "task",
                        "state": "prepared",
                        "evaluation": {"state": "pending", "calls": 0},
                    }
                ],
            }
            write_json(campaign / "campaign.json", manifest)
            agent_result = {
                "state": "completed",
                "patch_exists": True,
                "container": {
                    "id": "container-id",
                    "name": "bgp-swe-agent-test",
                    "cleanup": {
                        "policy": "retain",
                        "removed": False,
                        "retained": True,
                        "stopped": True,
                    },
                },
            }
            official_result = {"state": "completed", "calls": 1, "resolved": False}
            with (
                mock.patch.object(runtime, "_run_agent", return_value=agent_result),
                mock.patch.object(
                    runtime, "_official_evaluation", return_value=official_result
                ) as official,
            ):
                exit_code = runtime.execute_campaign(campaign)

            saved = json.loads((campaign / "campaign.json").read_text())
            self.assertEqual(exit_code, 0)
            self.assertEqual(saved["state"], "completed")
            official.assert_called_once()

    def test_goal_plus_evidence_failure_preserves_official_raw_score(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            manifest = {
                "schema_version": 1,
                "campaign_id": "goal-plus-partial-test",
                "benchmark_id": "swe-bench-verified",
                "state": "prepared",
                "methods": ["goal-plus-pi"],
                "model": "zai/glm-5.2",
                "budget": {"wall_time_seconds": 1800},
                "cells": [
                    {
                        "cell_id": "cell",
                        "task_id": "task",
                        "method": "goal-plus-pi",
                        "state": "prepared",
                        "evaluation": {"state": "pending", "calls": 0},
                    }
                ],
            }
            write_json(campaign / "campaign.json", manifest)
            agent_result = {
                "state": "partial",
                "patch_exists": True,
                "goal_plus": {
                    "actual_subagent_count": 0,
                    "completion": {
                        "passed": False,
                        "reason": "Goal Plus completion evidence failed: bound_pi_worker_sessions",
                    },
                },
                "container": {
                    "cleanup": {
                        "attempted": True,
                        "removed": True,
                        "retained": False,
                    }
                },
            }
            official_result = {
                "state": "completed",
                "calls": 1,
                "resolved": True,
                "patch_applied": True,
            }
            with (
                mock.patch.object(runtime, "_run_agent", return_value=agent_result),
                mock.patch.object(
                    runtime, "_official_evaluation", return_value=official_result
                ),
            ):
                exit_code = runtime.execute_campaign(campaign)

            saved = json.loads((campaign / "campaign.json").read_text())
            self.assertEqual(exit_code, 1)
            self.assertEqual(saved["state"], "partial")
            self.assertTrue(saved["cells"][0]["evaluation"]["resolved"])
            self.assertTrue(saved["cells"][0]["evaluation"]["patch_applied"])
            self.assertIn(
                "bound_pi_worker_sessions",
                saved["cells"][0]["incomplete_reason"],
            )

    def test_debug_disposition_stops_without_removing_agent_container(self) -> None:
        def run(command: list[str], **_kwargs):
            if command[:2] == ["docker", "stop"]:
                return subprocess.CompletedProcess(command, 0, "container-id\n", "")
            return subprocess.CompletedProcess(
                command,
                0,
                json.dumps({"Status": "exited", "Running": False, "ExitCode": 0}),
                "",
            )

        with mock.patch.object(runtime, "_run", side_effect=run) as run_command:
            disposition = runtime._dispose_agent_container(
                "container-id", retain=True
            )

        self.assertEqual(
            run_command.call_args_list[0].args[0],
            ["docker", "stop", "--time", "10", "container-id"],
        )
        self.assertEqual(
            run_command.call_args_list[1].args[0],
            [
                "docker",
                "inspect",
                "--format",
                "{{json .State}}",
                "container-id",
            ],
        )
        self.assertTrue(disposition["retained"])
        self.assertTrue(disposition["stopped"])
        self.assertFalse(disposition["removed"])
        self.assertEqual(disposition["observed_state"]["status"], "exited")

    def test_official_harness_is_recorded_once_and_normalized(self) -> None:
        profile = self.profile("sympy-16886-codex-smoke")
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            evaluator = campaign / "evaluator"
            evaluator.mkdir()
            (evaluator / "instances.json").write_text(
                json.dumps([{"instance_id": "task"}]) + "\n",
                encoding="utf-8",
            )
            patch_path = campaign / "model.patch"
            patch_path.write_text("diff --git a/a b/a\n")
            cell = {
                "task_id": "task",
                "method": "plain-codex",
                "model": "gpt-5.6-sol",
                "patch_file": "model.patch",
                "evaluation": {"state": "pending", "calls": 0},
            }
            manifest = {
                "schema_version": 1,
                "campaign_id": "official-once",
                "profile_snapshot": profile,
                "cells": [cell],
            }
            write_json(campaign / "campaign.json", manifest)
            report_path = (
                evaluator
                / "logs/run_evaluation/official-once"
                / "bench-goal-plus-plain-codex-gpt-5.6-sol"
                / "task/report.json"
            )
            report_path.parent.mkdir(parents=True)
            write_json(
                report_path,
                {
                    "task": {
                        "patch_successfully_applied": True,
                        "resolved": False,
                    }
                },
            )
            completed = subprocess.CompletedProcess([], 0, "official output", "")
            with mock.patch.object(runtime, "_run", return_value=completed):
                result = runtime._official_evaluation(campaign, manifest, cell)

            self.assertEqual(result["state"], "completed")
            self.assertEqual(result["calls"], 1)
            self.assertTrue(result["patch_applied"])
            self.assertFalse(result["resolved"])
            self.assertIn("swebench.harness.run_evaluation", result["command"])
            self.assertEqual(
                result["command"][result["command"].index("--cache_level") + 1],
                "instance",
            )
            self.assertEqual(
                result["command"][result["command"].index("--clean") + 1],
                "false",
            )
            self.assertEqual(
                result["command"][result["command"].index("--force_rebuild") + 1],
                "false",
            )
            persisted = json.loads((campaign / "campaign.json").read_text())
            self.assertEqual(persisted["cells"][0]["evaluation"]["calls"], 1)
            with self.assertRaisesRegex(SweBenchContractError, "already been attempted"):
                runtime._official_evaluation(campaign, manifest, cell)

    def test_official_evaluator_paths_and_run_ids_are_unique_per_cell(self) -> None:
        profile = self.profile("sympy-16886-19346-codex-terra-c2")
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            instances = campaign / "evaluator/instances.json"
            instances.parent.mkdir(parents=True)
            instances.write_text(
                json.dumps(
                    [
                        {"instance_id": task_id}
                        for task_id in profile["task_ids"]
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            cells = []
            for task in profile["tasks"]:
                cell_id = f"plain-codex--{task['instance_id']}"
                cell = {
                    "cell_id": cell_id,
                    "cell_file": f"cells/{cell_id}/cell.json",
                    "task_id": task["instance_id"],
                    "method": "plain-codex",
                    "model": profile["model"],
                    "patch_file": f"cells/{cell_id}/model.patch",
                    "evaluator_dir": f"evaluator/{cell_id}",
                    "evaluation": {"state": "pending", "calls": 0},
                }
                cells.append(cell)
                patch_path = campaign / cell["patch_file"]
                patch_path.parent.mkdir(parents=True, exist_ok=True)
                patch_path.write_text("diff --git a/a b/a\n", encoding="utf-8")
                write_json(campaign / cell["cell_file"], cell)
            manifest = {
                "schema_version": 1,
                "campaign_id": "official-c2",
                "profile_snapshot": profile,
                "dataset": {"evaluator_instances_file": "evaluator/instances.json"},
                "cells": cells,
            }
            write_json(campaign / "campaign.json", manifest)
            observed_run_ids: list[str] = []

            def run(command: list[str], *, cwd: Path, **_kwargs):
                run_id = command[command.index("--run_id") + 1]
                task_id = command[command.index("--instance_ids") + 1]
                model_label = f"bench-goal-plus-plain-codex-{profile['model']}"
                observed_run_ids.append(run_id)
                report_path = (
                    cwd
                    / "logs/run_evaluation"
                    / run_id
                    / model_label
                    / task_id
                    / "report.json"
                )
                write_json(
                    report_path,
                    {
                        task_id: {
                            "patch_successfully_applied": True,
                            "resolved": False,
                        }
                    },
                )
                return subprocess.CompletedProcess(command, 0, "", "")

            with mock.patch.object(runtime, "_run", side_effect=run):
                results = [
                    runtime._official_evaluation(
                        campaign,
                        manifest,
                        cell,
                        persist_manifest=False,
                    )
                    for cell in cells
                ]

            self.assertEqual(len(set(observed_run_ids)), 2)
            self.assertTrue(all(result["calls"] == 1 for result in results))
            self.assertTrue(all(result["state"] == "completed" for result in results))
            self.assertEqual(
                len({result["report_file"].split("/")[1] for result in results}), 2
            )

    def test_finalize_preserves_raw_metric_and_report_kind(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            manifest = {
                "schema_version": 1,
                "campaign_id": "finalize-test",
                "benchmark_id": "swe-bench-verified",
                "state": "completed",
                "budget": {
                    "wall_time_seconds": 300,
                    "live_search_concurrency": 1,
                    "cell_concurrency": 1,
                    "attempts": 1,
                },
                "dataset": {"name": "dataset", "revision": "a" * 40},
                "source": {"swebench_commit": "b" * 40},
                "cells": [
                    {
                        "task_id": "task",
                        "cell_id": "cell",
                        "method": "plain-codex",
                        "model": "gpt-5.6-sol",
                        "reasoning_effort": "medium",
                        "state": "completed",
                        "image": "image:latest",
                        "base_commit": "c" * 40,
                        "patch_file": "model.patch",
                        "agent": {
                            "patch_exists": True,
                            "runtime_seconds": 300.0,
                            "total_runtime_seconds": 312.0,
                            "setup_runtime_seconds": 5.0,
                            "finalization_grace_seconds": 7.0,
                            "usage": {"coverage": "unavailable"},
                            "container": {
                                "id": "container-id",
                                "name": "bgp-swe-agent-test",
                                "cleanup": {
                                    "policy": "retain",
                                    "removed": False,
                                    "retained": True,
                                    "stopped": True,
                                },
                            },
                        },
                        "evaluation": {
                            "state": "completed",
                            "calls": 1,
                            "resolved": False,
                            "patch_applied": True,
                            "runtime_seconds": 10.0,
                        },
                    }
                ],
            }
            write_json(campaign / "campaign.json", manifest)

            summary = reporting.finalize_campaign(campaign)
            kind, loaded, markdown, source_json = benchmark_report.load_campaign(campaign)

            self.assertEqual(summary["report_kind"], "swe-bench-verified")
            self.assertEqual(summary["records"][0]["protocol"]["direction"], "maximize")
            self.assertIs(summary["records"][0]["score"]["raw_metrics"]["resolved"], False)
            self.assertEqual(
                summary["records"][0]["execution"]["finalization_grace_seconds"], 7.0
            )
            self.assertTrue(
                summary["records"][0]["execution"]["agent_container"]["cleanup"][
                    "retained"
                ]
            )
            self.assertEqual(kind, "swe-bench-verified")
            self.assertEqual(loaded["aggregates"]["resolved_count"], 0)
            self.assertEqual(markdown, campaign / "campaign-summary.md")
            self.assertEqual(source_json, campaign / "campaign-summary.json")

    def test_finalize_and_export_include_two_result_rows(self) -> None:
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)

            def cell(task_id: str, resolved: bool) -> dict:
                return {
                    "task_id": task_id,
                    "cell_id": f"plain-codex--{task_id}",
                    "method": "plain-codex",
                    "model": "gpt-5.6-terra",
                    "reasoning_effort": "medium",
                    "state": "completed",
                    "image": f"image-{task_id}:latest",
                    "base_commit": "c" * 40,
                    "patch_file": f"cells/{task_id}/model.patch",
                    "agent": {
                        "state": "completed",
                        "patch_exists": True,
                        "usage": {"coverage": "unavailable"},
                        "container": {
                            "cleanup": {"removed": True, "retained": False}
                        },
                    },
                    "evaluation": {
                        "state": "completed",
                        "calls": 1,
                        "resolved": resolved,
                        "patch_applied": True,
                    },
                }

            manifest = {
                "schema_version": 1,
                "campaign_id": "two-row-report",
                "benchmark_id": "swe-bench-verified",
                "state": "completed",
                "budget": {
                    "wall_time_seconds": 1800,
                    "live_search_concurrency": 1,
                    "cell_concurrency": 2,
                    "attempts": 1,
                },
                "dataset": {"name": "dataset", "revision": "a" * 40},
                "source": {"swebench_commit": "b" * 40},
                "cells": [cell("task-a", True), cell("task-b", False)],
            }
            write_json(campaign / "campaign.json", manifest)

            summary = reporting.finalize_campaign(campaign)
            outputs = benchmark_report.export(campaign, None, None)
            markdown = (campaign / "report.md").read_text(encoding="utf-8")
            workbook = load_workbook(outputs["xlsx"], read_only=True)

            self.assertEqual(len(summary["records"]), 2)
            self.assertEqual(summary["aggregates"]["task_count"], 2)
            self.assertEqual(summary["aggregates"]["official_evaluator_calls"], 2)
            self.assertIn("task-a", markdown)
            self.assertIn("task-b", markdown)
            self.assertEqual(workbook["Results"].max_row, 3)
            workbook.close()

    def test_finalize_recovers_evidence_parser_downgrade(self) -> None:
        profile = self.profile("sympy-16886-goal-plus-pi-luna-high-smoke")
        with self.temporary_directory() as temporary:
            campaign = Path(temporary)
            state_root = campaign / "cells/goal-plus-pi/goal-plus-state"
            self.write_goal_plus_state(state_root)
            patch_file = campaign / "cells/goal-plus-pi/model.patch"
            patch_file.write_text("diff --git a/a b/a\n", encoding="utf-8")
            manifest = {
                "schema_version": 1,
                "campaign_id": "goal-plus-revalidation-test",
                "benchmark_id": "swe-bench-verified",
                "state": "partial",
                "profile_snapshot": profile,
                "budget": {
                    "wall_time_seconds": 1800,
                    "live_search_concurrency": 1,
                    "cell_concurrency": 1,
                    "attempts": 1,
                },
                "dataset": {"name": "dataset", "revision": "a" * 40},
                "source": {"swebench_commit": "b" * 40},
                "cells": [
                    {
                        "task_id": "task",
                        "cell_id": "goal-plus-pi--task",
                        "task_file": "cells/goal-plus-pi/task.json",
                        "patch_file": "cells/goal-plus-pi/model.patch",
                        "method": "goal-plus-pi",
                        "model": profile["model"],
                        "reasoning_effort": "high",
                        "state": "partial",
                        "incomplete_reason": (
                            "Goal Plus completion evidence failed: visible_verifiers"
                        ),
                        "image": "image:latest",
                        "base_commit": "c" * 40,
                        "agent": {
                            "state": "partial",
                            "patch_exists": True,
                            "runtime": {"evidence_annotator": "disabled"},
                            "goal_plus_closeout": {"completed": True, "returncode": 0},
                            "goal_plus": {
                                "completion": {
                                    "passed": False,
                                    "reason": (
                                        "Goal Plus completion evidence failed: "
                                        "visible_verifiers"
                                    ),
                                },
                                "export": {
                                    "completed": True,
                                    "destination": str(state_root),
                                },
                            },
                            "container": {
                                "cleanup": {
                                    "removed": True,
                                    "retained": False,
                                }
                            },
                        },
                        "evaluation": {
                            "state": "completed",
                            "calls": 1,
                            "resolved": True,
                            "patch_applied": True,
                        },
                    }
                ],
            }
            write_json(campaign / "campaign.json", manifest)

            summary = reporting.finalize_campaign(campaign)
            recovered = read_json(campaign / "campaign.json")

            self.assertEqual(summary["state"], "completed")
            self.assertEqual(summary["records"][0]["status"], "succeeded")
            self.assertTrue(
                summary["records"][0]["protocol"]["goal_plus"]["completion"][
                    "passed"
                ]
            )
            self.assertEqual(recovered["state"], "completed")
            self.assertEqual(recovered["cells"][0]["state"], "completed")
            self.assertNotIn("incomplete_reason", recovered["cells"][0])
            revalidation = recovered["cells"][0]["evidence_revalidation"]
            self.assertEqual(revalidation["prior_state"], "partial")
            self.assertIn("visible_verifiers", revalidation["prior_incomplete_reason"])
            self.assertTrue(revalidation["completion_passed"])

            recovered["state"] = "partial"
            recovered_cell = recovered["cells"][0]
            recovered_cell["state"] = "partial"
            recovered_cell["agent"]["state"] = "partial"
            recovered_cell["evaluation"]["calls"] = 2
            recovered_cell.pop("evidence_revalidation", None)
            write_json(campaign / "campaign.json", recovered)

            rejected = reporting.finalize_campaign(campaign)
            self.assertEqual(rejected["state"], "partial")
            self.assertEqual(rejected["records"][0]["status"], "partial")
            self.assertIn(
                "call count is not exactly one",
                rejected["records"][0]["incomplete_reason"],
            )


if __name__ == "__main__":
    unittest.main()
