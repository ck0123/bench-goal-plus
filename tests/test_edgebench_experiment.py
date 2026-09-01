from __future__ import annotations

import importlib.util
import io
import json
import os
import subprocess
import sys
import tarfile
import time
import unittest
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location(
    "edgebench_experiment",
    ROOT / "experiments" / "edgebench" / "experiment.py",
)
assert SPEC and SPEC.loader
EDGE = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = EDGE
SPEC.loader.exec_module(EDGE)

from experiments.edgebench.controller import io as EDGE_IO
from experiments.edgebench.controller import cli as EDGE_CLI
from experiments.edgebench.controller import environment as EDGE_ENV
from experiments.edgebench.controller import evidence as EDGE_EVIDENCE
from experiments.edgebench.controller import profiles as EDGE_PROFILES
from experiments.edgebench.controller import reporting as EDGE_REPORTING
from experiments.edgebench.controller import runtime as EDGE_RUNTIME


class EdgeBenchExperimentTest(unittest.TestCase):
    def test_doctor_accepts_exact_reasoning_override(self) -> None:
        args = EDGE_CLI.build_parser().parse_args(
            [
                "doctor",
                "--profile",
                "vliw-smoke",
                "--method",
                "goal-plus-codex",
                "--model",
                "gpt-5.6-sol",
                "--reasoning-effort",
                "medium",
            ]
        )

        self.assertEqual(args.method, ["goal-plus-codex"])
        self.assertEqual(args.model, "gpt-5.6-sol")
        self.assertEqual(args.reasoning_effort, "medium")

    def test_plan_metadata_accepts_selected_goal_plus_method(self) -> None:
        args = EDGE_CLI.build_parser().parse_args(
            [
                "plan-metadata",
                "--profile",
                "vliw-goal-plus-codex-gpt55-high-local-smoke",
                "--method",
                "goal-plus-codex",
            ]
        )

        self.assertEqual(args.method, ["goal-plus-codex"])

    def test_plan_metadata_reports_native_goal_plus_feature_flags(self) -> None:
        source = {
            "valid": True,
            "source_kind": "managed",
            "source_path": "third_party/muyuan/plugins/goal-plus",
            "expected_ref": "master",
            "branch": "master",
            "commit": "a" * 40,
        }
        stdout = io.StringIO()
        with mock.patch.object(
            EDGE_CLI, "resolve_goal_plus_source", return_value=source
        ), mock.patch("sys.stdout", stdout):
            returncode = EDGE_CLI.main(
                [
                    "plan-metadata",
                    "--profile",
                    "vliw-goal-plus-pi-sol-medium-local-smoke",
                    "--method",
                    "goal-plus-pi",
                ]
            )

        self.assertEqual(returncode, 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(
            payload["runtime_configuration"]["goal_plus"],
            {
                "shared_dir_enabled": False,
                "supplemental_evaluation_enabled": False,
            },
        )

    def test_compatibility_entrypoint_stays_thin(self) -> None:
        entrypoint = ROOT / "experiments" / "edgebench" / "experiment.py"
        source = entrypoint.read_text(encoding="utf-8")

        self.assertLessEqual(len(source.splitlines()), 200)
        for implementation in (
            "def doctor_payload(",
            "def execute_campaign(",
            "def finalize_campaign(",
            "def prepare(",
        ):
            self.assertNotIn(implementation, source)
        for module in (
            "cli.py",
            "context.py",
            "asset_issues.py",
            "environment.py",
            "evidence.py",
            "io.py",
            "preparation.py",
            "profiles.py",
            "reporting.py",
            "runtime.py",
        ):
            self.assertTrue(
                (ROOT / "experiments" / "edgebench" / "controller" / module).is_file()
            )

    def test_live_goal_plus_status_uses_pi_runtime_snapshot(self) -> None:
        task_run = self.temp / "task-run"
        task_run.mkdir(parents=True)
        (task_run / "goal-plus-live-status.json").write_text(
            json.dumps(
                {
                    "captured_at": "2026-07-31T10:40:00Z",
                    "candidate_ids": ["c001", "c002"],
                    "candidate_count": 2,
                    "worker_sessions": [
                        {
                            "agent_session_id": "agent_001",
                            "candidate_id": "c001",
                            "host": "pi",
                            "verifier_runs": 4,
                        },
                        {
                            "agent_session_id": "agent_002",
                            "candidate_id": "c002",
                            "host": "pi",
                            "verifier_runs": 7,
                        },
                    ],
                    "agent_session_count": 2,
                    "bound_worker_handles": [
                        {
                            "agent_session_id": "agent_001",
                            "host": "pi-rpc",
                            "external_id": "agent_001",
                        },
                        {
                            "agent_session_id": "agent_002",
                            "host": "pi-rpc",
                            "external_id": "agent_002",
                        },
                    ],
                    "actual_worker_launch_count": 2,
                    "verifier_ledger": [
                        {"candidate_id": "c001", "iteration": index}
                        for index in range(1, 5)
                    ]
                    + [
                        {"candidate_id": "c002", "iteration": index}
                        for index in range(1, 8)
                    ],
                    "worker_verifier_runs": 11,
                    "verifier_candidate_ids": ["c001", "c002"],
                    "selected_candidate_ids": ["c001"],
                    "promoted_candidate_ids": ["c001"],
                    "evidence_annotations": {
                        "tasks": 3,
                        "attempts": 2,
                        "views_published": 1,
                        "states": {"completed": 1, "pending": 2},
                        "active_attempts": [
                            {
                                "candidate_id": "c002",
                                "iteration": 1,
                                "attempt": 1,
                                "state": "running",
                                "json_lines": 4,
                                "event_type_counts": {"message_update": 4},
                            }
                        ],
                        "recent_attempts": [],
                        "monitor_files": 1,
                    },
                    "goal_statuses": [
                        {"goal_plus_id": "gp_0001", "status": "complete"}
                    ],
                    "terminal_ready": True,
                }
            ),
            encoding="utf-8",
        )

        status = EDGE_EVIDENCE.live_goal_plus_status(
            self.temp,
            {
                "state": "running",
                "started_at": "2026-07-31T10:35:00Z",
                "wall_time_seconds": 600,
                "goal_plus_finalization_grace_seconds": 120,
                "task_id": "vliw_kernel_optimization",
                "sforge_run_id": "pi-live-status-test",
            },
            task_run,
        )

        self.assertEqual(status["candidate_count"], 2)
        self.assertEqual(status["agent_session_count"], 2)
        self.assertEqual(status["actual_worker_launch_count"], 2)
        self.assertEqual(status["worker_verifier_runs"], 11)
        self.assertEqual(status["promoted_candidate_ids"], ["c001"])
        self.assertEqual(status["evidence_annotations"]["views_published"], 1)
        self.assertEqual(
            status["evidence_annotations"]["active_attempts"][0]["state"],
            "running",
        )
        self.assertEqual(
            status["evidence_annotations"]["active_attempts"][0]["event_type_counts"],
            {"message_update": 4},
        )
        self.assertTrue(status["terminal_ready"])
        self.assertEqual(status["state_sources"], ["goal-plus-live-status.json"])

    def test_live_goal_plus_codex_does_not_count_allocated_sessions_as_workers(
        self,
    ) -> None:
        task_run = self.temp / "task-run"
        task_run.mkdir(parents=True)
        (task_run / "goal-plus-live-status.json").write_text(
            json.dumps(
                {
                    "candidate_ids": ["c001", "c002"],
                    "candidate_count": 2,
                    "agent_session_count": 2,
                    "actual_worker_launch_count": 2,
                    "bound_worker_handles": [
                        {
                            "agent_session_id": "agent_001",
                            "host": "codex",
                            "external_id": "allocated-task-name-001",
                        },
                        {
                            "agent_session_id": "agent_002",
                            "host": "codex",
                            "external_id": "allocated-task-name-002",
                        },
                    ],
                }
            ),
            encoding="utf-8",
        )

        status = EDGE_EVIDENCE.live_goal_plus_status(
            self.temp,
            {
                "method": "goal-plus-codex",
                "state": "running",
                "started_at": "2026-07-31T10:35:00Z",
                "wall_time_seconds": 600,
                "goal_plus_finalization_grace_seconds": 120,
                "task_id": "vliw_kernel_optimization",
                "sforge_run_id": "codex-live-status-test",
            },
            task_run,
        )

        self.assertEqual(status["candidate_count"], 2)
        self.assertEqual(status["agent_session_count"], 2)
        self.assertEqual(status["actual_worker_launch_count"], 0)
        self.assertEqual(status["spawn_agent_completed_count"], 0)

    def test_remaining_time_uses_agent_start_after_install(self) -> None:
        task_run = self.temp / "task-run"
        task_run.mkdir(parents=True)
        agent_started_at = time.time() - 120
        (task_run / "started_at").write_text(
            f"2026-07-31T10:35:00\n{agent_started_at}\n",
            encoding="utf-8",
        )

        remaining = EDGE_EVIDENCE.remaining_time(
            {
                "state": "running",
                "started_at": "2000-01-01T00:00:00Z",
                "wall_time_seconds": 600,
                "goal_plus_finalization_grace_seconds": 120,
            },
            task_run,
        )

        self.assertGreaterEqual(remaining["exploration_seconds"], 479)
        self.assertLessEqual(remaining["exploration_seconds"], 480)
        self.assertGreaterEqual(remaining["finalization_seconds"], 599)
        self.assertLessEqual(remaining["finalization_seconds"], 600)

    def setUp(self) -> None:
        self.temp = (
            EDGE.ensure_temp_root("test-edgebench-experiment")
            / f"{self._testMethodName}-{time.time_ns()}"
        )
        self.temp.mkdir(parents=True)
        self.original_paths = EDGE.current_paths()
        self.test_paths = replace(
            self.original_paths,
            edge_root=self.temp / "edgebench",
            goal_plus_root=self.temp / "goal-plus",
            tasks_dir=self.temp / "edgebench" / "tasks",
            runs_root=self.temp / "runs",
        )
        EDGE.set_paths(self.test_paths)
        self.test_paths.tasks_dir.mkdir(parents=True)
        self.test_paths.goal_plus_root.mkdir(parents=True)
        for relative in EDGE_ENV.GOAL_PLUS_REQUIRED_ASSETS:
            asset = self.test_paths.goal_plus_root / relative
            asset.parent.mkdir(parents=True, exist_ok=True)
            asset.write_text("fixture\n", encoding="utf-8")
        plugin_hooks = self.test_paths.goal_plus_root / "hooks" / "hooks.json"
        plugin_hooks.parent.mkdir(parents=True, exist_ok=True)
        plugin_hooks.write_text('{"hooks": {}}\n', encoding="utf-8")
        subprocess.run(
            [
                "git",
                "init",
                "--initial-branch=master",
                str(self.test_paths.goal_plus_root),
            ],
            check=True,
            capture_output=True,
        )
        subprocess.run(
            ["git", "-C", str(self.test_paths.goal_plus_root), "add", "."],
            check=True,
            capture_output=True,
        )
        subprocess.run(
            [
                "git",
                "-C",
                str(self.test_paths.goal_plus_root),
                "-c",
                "user.name=EdgeBench Test",
                "-c",
                "user.email=edgebench@example.invalid",
                "commit",
                "-m",
                "fixture",
            ],
            check=True,
            capture_output=True,
        )
        (self.test_paths.tasks_dir / "vliw_kernel_optimization.json").write_text(
            json.dumps(
                {
                    "task_id": "vliw_kernel_optimization",
                    "internet": False,
                    "work": {
                        "agent_query": "Optimize solution.py.",
                        "image_tag": "work123",
                    },
                    "judge": {
                        "image_tag": "judge123",
                        "score_direction": "minimize",
                    },
                }
            ),
            encoding="utf-8",
        )

    def tearDown(self) -> None:
        EDGE.set_paths(self.original_paths)

    def _commit_goal_plus_fixture(self, message: str) -> None:
        subprocess.run(
            ["git", "-C", str(self.test_paths.goal_plus_root), "add", "-A"],
            check=True,
            capture_output=True,
        )
        subprocess.run(
            [
                "git",
                "-C",
                str(self.test_paths.goal_plus_root),
                "-c",
                "user.name=EdgeBench Test",
                "-c",
                "user.email=edgebench@example.invalid",
                "commit",
                "-m",
                message,
            ],
            check=True,
            capture_output=True,
        )

    def profile(self) -> dict:
        _, profile = EDGE.load_profile("vliw-smoke")
        return profile

    def write_dataset_revision(self, revision: str) -> None:
        metadata = (
            self.test_paths.tasks_dir
            / ".cache"
            / "huggingface"
            / "download"
            / "vliw_kernel_optimization.json.metadata"
        )
        metadata.parent.mkdir(parents=True, exist_ok=True)
        metadata.write_text(revision + "\n", encoding="utf-8")

    def test_local_asset_inventory_lists_exact_images_and_containers(self) -> None:
        profile = self.profile()
        self.write_dataset_revision(profile["dataset_revision"])
        work_ref = "edgebench.work.vliw_kernel_optimization:work123"
        judge_ref = "edgebench.judge.vliw_kernel_optimization:judge123"
        commands = []
        original = EDGE_IO.run_capture

        def fake_run_capture(command, *, env=None):
            commands.append(command)
            if command[:3] == ["docker", "ps", "-a"]:
                return {
                    "returncode": 0,
                    "stdout": json.dumps(
                        {
                            "ID": "container-work",
                            "Names": "edgebench-work",
                            "Image": work_ref,
                            "ImageID": "sha256:work-id",
                            "State": "running",
                            "Status": "Up 5 minutes",
                        }
                    ),
                    "stderr": "",
                }
            if command[:3] == ["docker", "image", "inspect"]:
                reference = command[-1]
                image_id = (
                    "sha256:work-id" if reference == work_ref else "sha256:judge-id"
                )
                return {
                    "returncode": 0,
                    "stdout": json.dumps(
                        [
                            {
                                "Id": image_id,
                                "RepoTags": [reference],
                                "RepoDigests": [
                                    reference.split(":", 1)[0] + "@sha256:digest"
                                ],
                                "Size": 123456,
                                "Architecture": "amd64",
                                "Os": "linux",
                            }
                        ]
                    ),
                    "stderr": "",
                }
            self.fail(f"unexpected inventory command: {command}")

        EDGE_IO.run_capture = fake_run_capture
        try:
            payload = EDGE.local_asset_inventory(profile)
        finally:
            EDGE_IO.run_capture = original

        self.assertTrue(payload["ok"])
        self.assertTrue(payload["read_only"])
        self.assertFalse(payload["acquisition_attempted"])
        task = payload["tasks"][0]
        self.assertTrue(task["task_file_present"])
        self.assertTrue(task["dataset_revision_matches"])
        self.assertEqual(
            [image["reference"] for image in task["images"]],
            [work_ref, judge_ref],
        )
        self.assertEqual(
            [image["image_id"] for image in task["images"]],
            ["sha256:work-id", "sha256:judge-id"],
        )
        self.assertEqual(task["images"][0]["containers"][0]["id"], "container-work")
        self.assertEqual(payload["summary"]["images_present"], 2)
        self.assertEqual(payload["docker_commands"], commands)
        for command in commands:
            self.assertTrue(
                command[:3]
                in (
                    ["docker", "ps", "-a"],
                    ["docker", "image", "inspect"],
                )
            )
            self.assertNotIn("pull", command)
            self.assertNotIn("run", command)

    def test_local_asset_inventory_reports_missing_exact_image(self) -> None:
        profile = self.profile()
        self.write_dataset_revision(profile["dataset_revision"])
        missing_ref = "edgebench.judge.vliw_kernel_optimization:judge123"
        original = EDGE_IO.run_capture

        def fake_run_capture(command, *, env=None):
            if command[:3] == ["docker", "ps", "-a"]:
                return {"returncode": 0, "stdout": "", "stderr": ""}
            if command[-1] == missing_ref:
                return {
                    "returncode": 1,
                    "stdout": "",
                    "stderr": "No such image",
                }
            reference = command[-1]
            return {
                "returncode": 0,
                "stdout": json.dumps(
                    [
                        {
                            "Id": "sha256:work-id",
                            "RepoTags": [reference],
                            "RepoDigests": [],
                            "Size": 123456,
                            "Architecture": "amd64",
                            "Os": "linux",
                        }
                    ]
                ),
                "stderr": "",
            }

        EDGE_IO.run_capture = fake_run_capture
        try:
            payload = EDGE.local_asset_inventory(profile)
        finally:
            EDGE_IO.run_capture = original

        self.assertFalse(payload["ok"])
        self.assertEqual(payload["summary"]["images_missing"], 1)
        self.assertEqual(payload["missing_image_references"], [missing_ref])
        self.assertEqual(payload["tasks"][0]["images"][1]["error"], "No such image")

    def test_local_asset_inventory_blocks_known_broken_judge_image(self) -> None:
        task_id = "order_addition_permutation_optimization"
        work_ref = f"edgebench.work.{task_id}:f723a1d13d8e"
        judge_ref = f"edgebench.judge.{task_id}:f6f385925889"
        judge_image_id = (
            "sha256:97b871cd558d3e7eacc22f5a85ac50a85ffe2e7333512de8cfdf3fc2cadd4f09"
        )
        profile = {
            **self.profile(),
            "task_ids": [task_id],
        }
        (self.test_paths.tasks_dir / f"{task_id}.json").write_text(
            json.dumps(
                {
                    "task_id": task_id,
                    "work": {"image_tag": "f723a1d13d8e"},
                    "judge": {"image_tag": "f6f385925889"},
                }
            ),
            encoding="utf-8",
        )
        metadata = (
            self.test_paths.tasks_dir
            / ".cache"
            / "huggingface"
            / "download"
            / f"{task_id}.json.metadata"
        )
        metadata.parent.mkdir(parents=True, exist_ok=True)
        metadata.write_text(profile["dataset_revision"] + "\n", encoding="utf-8")

        def fake_run_capture(command, *, env=None):
            if command[:3] == ["docker", "ps", "-a"]:
                return {"returncode": 0, "stdout": "", "stderr": ""}
            reference = command[-1]
            return {
                "returncode": 0,
                "stdout": json.dumps(
                    [
                        {
                            "Id": (
                                judge_image_id
                                if reference == judge_ref
                                else "sha256:work-id"
                            ),
                            "RepoTags": [reference],
                            "RepoDigests": [],
                            "Size": 123456,
                            "Architecture": "amd64",
                            "Os": "linux",
                        }
                    ]
                ),
                "stderr": "",
            }

        with mock.patch.object(EDGE_IO, "run_capture", side_effect=fake_run_capture):
            payload = EDGE.local_asset_inventory(profile)

        self.assertFalse(payload["ok"])
        self.assertEqual(payload["summary"]["images_present"], 2)
        self.assertEqual(payload["summary"]["blocking_known_asset_issues"], 1)
        self.assertEqual(
            payload["blocking_known_asset_issues"][0]["id"],
            "order-addition-judge-score-helper-sha-mismatch",
        )
        judge = payload["tasks"][0]["images"][1]
        self.assertEqual(judge["reference"], judge_ref)
        self.assertTrue(judge["known_issue_image_id_matches"])
        self.assertEqual(judge["known_issue"]["severity"], "blocking")
        self.assertEqual(
            [work_ref, judge_ref],
            [image["reference"] for image in payload["tasks"][0]["images"]],
        )

    def test_known_judge_asset_issue_qualifies_existing_campaign_score(self) -> None:
        issue = EDGE.asset_protocol_issue(
            "order_addition_permutation_optimization",
            "47846a4c3669ad447e0ea984833b0d352460c5f9",
        )

        self.assertIsNotNone(issue)
        self.assertIn("score-helper-sha-mismatch", issue)
        self.assertIsNotNone(
            EDGE.asset_protocol_issue(
                "order_addition_permutation_optimization",
                "6cc5a7f1b3288ce52484ad828177b8cc86b05b75",
            )
        )
        self.assertIsNone(
            EDGE.asset_protocol_issue(
                "order_addition_permutation_optimization",
                "future-corrected-dataset-revision",
            )
        )

    def test_known_asset_issue_requires_the_broken_image_id(self) -> None:
        issues = EDGE_ENV.known_asset_issues()
        reference = (
            "edgebench.judge.order_addition_permutation_optimization:f6f385925889"
        )

        self.assertIsNotNone(
            EDGE_ENV._known_asset_issue(
                issues,
                "Judge",
                reference,
                "6cc5a7f1b3288ce52484ad828177b8cc86b05b75",
                "sha256:97b871cd558d3e7eacc22f5a85ac50a85ffe2e7333512de8cfdf3fc2cadd4f09",
            )
        )
        self.assertIsNone(
            EDGE_ENV._known_asset_issue(
                issues,
                "Judge",
                reference,
                "6cc5a7f1b3288ce52484ad828177b8cc86b05b75",
                "sha256:corrected-image",
            )
        )

    def test_full_codex_profile_covers_all_runnable_public_tasks(self) -> None:
        _, profile = EDGE.load_profile("full-codex-2h")
        _, terra_profile = EDGE.load_profile("full-codex-terra-high-2h-k1-c4")

        self.assertEqual(len(profile["task_ids"]), 50)
        self.assertEqual(len(set(profile["task_ids"])), 50)
        self.assertEqual(set(terra_profile["task_ids"]), set(profile["task_ids"]))
        self.assertNotIn("order_addition_permutation_optimization", profile["task_ids"])
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 2)
        self.assertNotIn("work_cpu_limit", profile)
        self.assertNotIn("judge_cpu_limit", profile)

    def test_profile_rejects_campaign_exclusion_for_affected_revision(self) -> None:
        _, profile = EDGE.load_profile("vliw-smoke")
        profile["task_ids"] = ["order_addition_permutation_optimization"]
        path = self.temp / "excluded-task.json"
        path.write_text(json.dumps(profile), encoding="utf-8")

        with self.assertRaisesRegex(
            ValueError,
            "order_addition_permutation_optimization.*excluded_from_campaigns",
        ):
            EDGE.load_profile(path)

        profile["dataset_revision"] = "future-corrected-dataset-revision"
        path.write_text(json.dumps(profile), encoding="utf-8")
        _, loaded = EDGE.load_profile(path)
        self.assertEqual(
            loaded["task_ids"], ["order_addition_permutation_optimization"]
        )

    def test_vliw_codex_local_smoke_is_explicit_and_reproducible(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 300)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 1)
        self.assertEqual(
            profile["protocol_overrides"],
            {"eval_interval": 60, "internet": False},
        )
        self.assertTrue(profile["protocol_override_reasons"]["eval_interval"])
        self.assertTrue(profile["protocol_override_reasons"]["internet"])

    def test_registered_profiles_never_enable_agent_internet(self) -> None:
        for path in sorted((ROOT / "experiments/edgebench/profiles").glob("*.json")):
            profile = json.loads(path.read_text())
            self.assertIsNot(
                (profile.get("protocol_overrides") or {}).get("internet"),
                True,
                path.name,
            )

    def test_vliw_goal_plus_codex_smoke_bounds_local_verifiers(self) -> None:
        _, profile = EDGE.load_profile(
            "vliw-goal-plus-codex-gpt55-high-local-smoke"
        )

        self.assertEqual(profile["wall_time_seconds"], 900)
        self.assertEqual(profile["worker_runtime_seconds"], 240)
        self.assertEqual(profile["goal_plus_verifier_timeout_seconds"], 30)

    def test_pi_profiles_use_canonical_method_names_and_explicit_budgets(self) -> None:
        _, plain = EDGE.load_profile("vliw-pi-sol-medium-local-smoke")
        _, goal_plus = EDGE.load_profile("vliw-goal-plus-pi-sol-medium-local-smoke")
        _, api_provider = EDGE.load_profile(
            "vliw-goal-plus-pi-glm-5-2-provider-1h-k2-c1"
        )
        _, zai_provider = EDGE.load_profile("vliw-goal-plus-pi-zai-glm-5-2-1h-k2-c1")

        self.assertEqual(EDGE.METHODS["plain-pi"]["agent"], "pi")
        self.assertEqual(EDGE.METHODS["plain-pi-provider"]["agent"], "pi-provider")
        self.assertEqual(
            EDGE.METHODS["plain-pi-provider"]["api_protocol"], "pi-provider"
        )
        self.assertEqual(EDGE.METHODS["goal-plus-pi"]["agent"], "pi-goal-plus")
        self.assertFalse(EDGE_PROFILES.methods_require_codex(["goal-plus-pi"]))
        self.assertTrue(EDGE_PROFILES.methods_require_codex(["goal-plus-codex"]))
        self.assertEqual(
            EDGE.METHODS["goal-plus-pi-provider"]["agent"],
            "pi-goal-plus-provider",
        )
        self.assertEqual(plain["methods"], ["plain-pi"])
        self.assertEqual(goal_plus["methods"], ["goal-plus-pi"])
        self.assertEqual(goal_plus["concurrency"], 2)
        self.assertEqual(goal_plus["worker_runtime_seconds"], 240)
        self.assertEqual(goal_plus["worker_min_verifier_runs"], 1)
        self.assertEqual(goal_plus["goal_plus_verifier_timeout_seconds"], 30)
        self.assertEqual(goal_plus["goal_plus_finalization_grace_seconds"], 300)
        self.assertEqual(goal_plus["global_evidence_mode"], "manual")
        self.assertEqual(api_provider["methods"], ["goal-plus-pi-provider"])
        self.assertEqual(api_provider["model"], "glm-proxy/GLM-5.2")
        self.assertEqual(api_provider["pi_package_version"], "0.83.0")
        self.assertEqual(api_provider["wall_time_seconds"], 3600)
        self.assertEqual(api_provider["concurrency"], 2)
        self.assertEqual(api_provider["cell_concurrency"], 1)
        self.assertEqual(zai_provider["methods"], ["goal-plus-pi-provider"])
        self.assertEqual(zai_provider["model"], "zai/glm-5.2")
        self.assertEqual(zai_provider["pi_package_version"], "0.83.0")
        self.assertEqual(zai_provider["wall_time_seconds"], 3600)
        self.assertEqual(zai_provider["concurrency"], 2)
        self.assertEqual(zai_provider["cell_concurrency"], 1)

    def test_goal_plus_profile_validates_global_evidence_mode(self) -> None:
        _, profile = EDGE.load_profile("vliw-goal-plus-pi-zai-glm-5-2-1h-k2-c1")
        profile["global_evidence_mode"] = "independent"
        path = self.temp / "independent-evidence.json"
        path.write_text(json.dumps(profile), encoding="utf-8")

        _, loaded = EDGE.load_profile(path)

        self.assertEqual(loaded["global_evidence_mode"], "independent")
        profile["global_evidence_mode"] = "sometimes"
        path.write_text(json.dumps(profile), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "global_evidence_mode"):
            EDGE.load_profile(path)

    def test_goal_plus_profile_validates_runtime_feature_flags(self) -> None:
        _, profile = EDGE.load_profile(
            "vliw-goal-plus-pi-zai-glm-5-2-1h-k2-c1"
        )
        self.assertFalse(profile["shared_dir_enabled"])
        self.assertFalse(profile["supplemental_evaluation_enabled"])

        for field in (
            "shared_dir_enabled",
            "supplemental_evaluation_enabled",
        ):
            with self.subTest(field=field):
                invalid = {**profile, field: "false"}
                path = self.temp / f"invalid-{field}.json"
                path.write_text(json.dumps(invalid), encoding="utf-8")
                with self.assertRaisesRegex(ValueError, f"{field} must be boolean"):
                    EDGE.load_profile(path)

        _, plain = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")
        plain["shared_dir_enabled"] = False
        path = self.temp / "plain-with-goal-plus-field.json"
        path.write_text(json.dumps(plain), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "require a Goal Plus method"):
            EDGE.load_profile(path)

    def test_pi_provider_profile_requires_qualified_role_models(self) -> None:
        _, profile = EDGE.load_profile("vliw-goal-plus-pi-glm-5-2-provider-1h-k2-c1")
        for field in ("model", "worker_model", "evidence_annotator_model"):
            with self.subTest(field=field):
                invalid = {**profile, field: "unqualified-model"}
                path = self.temp / f"invalid-{field}.json"
                path.write_text(json.dumps(invalid), encoding="utf-8")
                with self.assertRaisesRegex(ValueError, "PROVIDER/MODEL"):
                    EDGE.load_profile(path)

    def test_profile_rejects_invalid_eval_interval_override(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")
        for value in (0, "60"):
            with self.subTest(value=value):
                profile["protocol_overrides"]["eval_interval"] = value
                path = self.temp / f"invalid-eval-interval-{value}.json"
                path.write_text(json.dumps(profile), encoding="utf-8")

                with self.assertRaisesRegex(
                    ValueError, "eval_interval override must be a positive integer"
                ):
                    EDGE.load_profile(path)

    def test_vliw_glm_profile_pins_claude_thinking_effort_and_budget(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-2-high-20m-k1")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-claude"])
        self.assertEqual(profile["model"], "glm-5.2")
        self.assertEqual(profile["thinking"], {"type": "enabled"})
        self.assertEqual(profile["reasoning_effort"], "high")
        self.assertEqual(profile["wall_time_seconds"], 1200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 1)

        _, none_profile = EDGE.load_profile("vliw-glm-5-2-none-20m-k1")
        self.assertEqual(none_profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(none_profile["methods"], ["plain-claude"])
        self.assertEqual(none_profile["model"], "glm-5.2")
        self.assertEqual(none_profile["thinking"], {"type": "disabled"})
        self.assertEqual(none_profile["reasoning_effort"], "none")
        self.assertEqual(none_profile["wall_time_seconds"], 1200)
        self.assertEqual(none_profile["concurrency"], 1)

    def test_vliw_glm_51_profile_uses_official_adaptive_defaults(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-1-adaptive-2h-k1")

        self.assertEqual(profile["task_ids"], ["vliw_kernel_optimization"])
        self.assertEqual(profile["methods"], ["plain-claude"])
        self.assertEqual(profile["model"], "glm-5.1")
        self.assertEqual(profile["thinking"], {"type": "adaptive"})
        self.assertNotIn("reasoning_effort", profile)
        self.assertEqual(profile["claude_context_window_tokens"], 200000)
        self.assertEqual(profile["claude_autocompact_percent"], 80)
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)

    def test_protocol_regression_profile_targets_prior_failures(self) -> None:
        _, profile = EDGE.load_profile("protocol-regression-codex-2h")

        self.assertEqual(
            profile["task_ids"],
            [
                "borden_source_inversion",
                "exchange_core_throughput",
                "schemathesis_config_modernization",
                "schemathesis_datagen_pipeline",
                "schemathesis_reporting_observability",
                "anchorhead_text_adventure",
                "trinity_text_adventure",
                "tryst_text_adventure",
            ],
        )
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 2)

    def test_validation_regression_profile_targets_suspicious_legacy_cells(
        self,
    ) -> None:
        _, profile = EDGE.load_profile("validation-regression-codex-2h-c4")

        self.assertEqual(len(profile["task_ids"]), 17)
        self.assertEqual(len(set(profile["task_ids"])), 17)
        self.assertEqual(profile["task_ids"][-1], "integer_compression_codec")
        self.assertEqual(profile["methods"], ["plain-codex"])
        self.assertEqual(profile["model"], "gpt-5.6-sol")
        self.assertEqual(profile["reasoning_effort"], "medium")
        self.assertEqual(profile["wall_time_seconds"], 7200)
        self.assertEqual(profile["concurrency"], 1)
        self.assertEqual(profile["cell_concurrency"], 4)
        self.assertEqual(profile["judge_concurrency"], 1)

    def test_official_codex_protocol_covers_all_tasks_and_overrides(self) -> None:
        protocol = EDGE.load_official_codex_protocol()
        _, full_profile = EDGE.load_profile("full-codex-2h")

        self.assertEqual(len(protocol["tasks"]), 51)
        self.assertEqual(
            set(protocol["tasks"]) - {"order_addition_permutation_optimization"},
            set(full_profile["task_ids"]),
        )
        self.assertEqual(protocol["official_model"], "gpt-5.5")
        self.assertEqual(protocol["stagger_seconds"], 600)
        self.assertEqual(
            protocol["defaults"],
            {
                "backend": "k8s",
                "agent": "codex",
                "timeout": 43200,
                "eval_interval": 1800,
                "submission_cooldown": 120,
                "work_cpu_limit": 4,
                "work_mem_limit": "16g",
                "judge_cpu_limit": 4,
                "judge_mem_limit": "8g",
            },
        )
        self.assertEqual(
            protocol["tasks"]["dabic_gravity_inversion"],
            {"submission_cooldown": 2160},
        )
        self.assertEqual(
            protocol["tasks"]["schemathesis_config_modernization"],
            {"submission_cooldown": 216},
        )
        self.assertEqual(
            protocol["tasks"]["anchorhead_text_adventure"],
            {"submission_cooldown": 0},
        )
        self.assertEqual(
            protocol["tasks"]["graph_node_classification"],
            {"judge_mem_limit": "16g"},
        )
        self.assertEqual(
            protocol["tasks"]["lean_analysis_proofs"],
            {
                "work_cpu_limit": 8,
                "work_mem_limit": "16g",
                "judge_cpu_limit": 8,
                "judge_mem_limit": "16g",
            },
        )
        self.assertEqual(protocol["tasks"]["smt_solver"]["work_cpu_limit"], 16)
        self.assertEqual(protocol["tasks"]["smt_solver"]["judge_cpu_limit"], 16)
        serialized = json.dumps(protocol)
        self.assertNotIn("sk-xxxx", serialized)
        self.assertNotIn("SFORGE_K8S_IMAGE_REGISTRY", serialized)
        self.assertNotIn("api_key", serialized)

    def test_official_task_protocol_uses_task_owned_internet(self) -> None:
        protocol = EDGE.load_official_codex_protocol()

        isolated = EDGE.official_task_protocol(
            protocol,
            "vliw_kernel_optimization",
            {"internet": False},
        )
        connected = EDGE.official_task_protocol(
            protocol,
            "college_english_exam_bank",
            {"internet": True},
        )

        self.assertFalse(isolated["internet"])
        self.assertTrue(connected["internet"])
        self.assertEqual(isolated["eval_interval"], 1800)
        self.assertEqual(isolated["submission_cooldown"], 120)
        self.assertEqual(isolated["work_cpu_limit"], 4)
        self.assertEqual(isolated["work_mem_limit"], "16g")
        self.assertEqual(isolated["judge_cpu_limit"], 4)
        self.assertEqual(isolated["judge_mem_limit"], "8g")
        self.assertFalse(isolated["disable_auto_eval"])
        self.assertFalse(isolated["disable_auto_resume"])
        self.assertFalse(isolated["disable_stop_hook"])

        explicit_lifecycle = {
            **protocol,
            "defaults": {
                **protocol["defaults"],
                "disable_auto_eval": True,
                "max_submissions": 5,
            },
        }
        explicit = EDGE.official_task_protocol(
            explicit_lifecycle,
            "vliw_kernel_optimization",
            {"internet": False},
        )
        self.assertTrue(explicit["disable_auto_eval"])
        self.assertEqual(explicit["max_submissions"], 5)

    def test_api_only_preflight_rejects_open_network_with_source(self) -> None:
        profile = self.profile()
        profile["protocol_overrides"] = {"internet": True}
        profile["protocol_override_reasons"]["internet"] = "invalid test override"

        with self.assertRaisesRegex(
            ValueError,
            r"internet=true.*profiles/vliw-smoke\.protocol_overrides\.internet",
        ):
            EDGE_ENV.require_api_only_network(profile)

        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="must-not-exist",
        )
        with self.assertRaisesRegex(ValueError, "API-only"):
            EDGE.prepare(args, profile)
        self.assertFalse((self.test_paths.runs_root / "must-not-exist").exists())

    def test_doctor_api_only_check_lists_model_calling_roles(self) -> None:
        profile = self.profile()
        report = EDGE_ENV.DoctorReport(profile["id"])
        EDGE_ENV._check_api_only_network(
            report, profile, EDGE.load_official_codex_protocol()
        )

        check = report.payload()["checks"][0]
        self.assertTrue(check["passed"])
        self.assertEqual(check["policy"], "api-only")
        self.assertEqual(check["allowed_classes"], ["judge", "llm-api"])
        self.assertEqual(
            {(item["method"], item["role"]) for item in check["api_endpoints"]},
            {
                ("plain-codex", "main"),
                ("goal-plus-codex", "main"),
                ("goal-plus-codex", "worker"),
                ("goal-plus-codex", "evidence_annotator"),
            },
        )

    def test_known_protocol_marker_depends_on_effective_cell_config(self) -> None:
        aligned = {
            "task_id": "schemathesis_config_modernization",
            "internet": False,
            "submission_cooldown": 216,
            "work_cpu_limit": 4,
            "work_mem_limit": "16g",
            "judge_cpu_limit": 4,
            "judge_mem_limit": "8g",
        }
        legacy = {**aligned, "internet": True, "submission_cooldown": None}

        self.assertIsNone(EDGE.paper_protocol_issue(aligned))
        self.assertIn("Internet access used", EDGE.paper_protocol_issue(legacy))

    def test_protocol_diff_rejects_resource_or_network_overrides(self) -> None:
        for field, official, effective in (
            ("work_cpu_limit", 4, 8),
            ("internet", False, True),
            ("submission_cooldown", 120, 0),
        ):
            with self.subTest(field=field):
                with self.assertRaisesRegex(
                    ValueError, "unsupported EdgeBench protocol override"
                ):
                    EDGE._protocol_diff(
                        official={field: official},
                        effective={field: effective},
                        reasons={field: "not permitted"},
                    )

    def test_protocol_diff_accepts_only_explicitly_allowed_network_override(
        self,
    ) -> None:
        diff = EDGE._protocol_diff(
            official={"internet": False},
            effective={"internet": True},
            reasons={"internet": "local development smoke"},
            allowed_fields=(EDGE.ALLOWED_PROTOCOL_OVERRIDE_FIELDS | {"internet"}),
        )

        self.assertEqual(
            diff,
            [
                {
                    "field": "internet",
                    "official": False,
                    "effective": True,
                    "reason": "local development smoke",
                }
            ],
        )

    def test_paper_gpt55_reference_covers_profile_and_records_provenance(self) -> None:
        _, profile = EDGE.load_profile("full-codex-2h")
        paper = EDGE.load_paper_reference()

        self.assertEqual(
            set(paper["tasks"]) - {"order_addition_permutation_optimization"},
            set(profile["task_ids"]),
        )
        self.assertEqual(paper["reference"]["agent"], "Codex")
        self.assertEqual(paper["reference"]["model"], "GPT-5.5")
        self.assertEqual(paper["reference"]["budget_hours"], 12)
        self.assertEqual(paper["reference"]["scheduled_runs"], 3)
        self.assertEqual(paper["tasks"]["borden_source_inversion"]["mean"], 38.5)
        self.assertEqual(
            paper["tasks"]["vliw_kernel_optimization"]["sample_stddev"],
            1.9,
        )
        self.assertEqual(
            paper["source"]["source_archive_sha256"],
            "8193aeb41a3474690a40fac82e2ecbd53e651ab6b4759984b4c6845c04fbfd29",
        )

    def test_paper_opus_headroom_reference_keeps_official_public_set(self) -> None:
        _, profile = EDGE.load_profile("full-codex-2h")
        reference_path = (
            ROOT
            / "experiments"
            / "edgebench"
            / "references"
            / "paper-opus-4.8-vs-gpt-5.5-headroom.json"
        )
        reference = json.loads(reference_path.read_text(encoding="utf-8"))
        groups = reference["candidate_groups"]

        partition = [task_id for task_ids in groups.values() for task_id in task_ids]
        material = (
            groups["material_at_both_2h_and_12h"]
            + groups["material_at_2h_only"]
            + groups["material_at_12h_only"]
        )

        self.assertEqual(len(partition), 51)
        self.assertEqual(len(partition), len(set(partition)))
        self.assertEqual(reference["scope"]["task_count"], 51)
        self.assertEqual(reference["scope"]["profile_task_count"], 50)
        self.assertEqual(
            reference["scope"]["excluded_from_campaigns"],
            ["order_addition_permutation_optimization"],
        )
        self.assertEqual(set(partition), set(EDGE.load_paper_reference()["tasks"]))
        self.assertEqual(
            set(partition) - {"order_addition_permutation_optimization"},
            set(profile["task_ids"]),
        )
        self.assertEqual(set(material), set(reference["candidates"]))
        self.assertEqual(reference["summary"]["material_at_2h_or_12h"], 20)
        self.assertEqual(reference["summary"]["material_at_both_2h_and_12h"], 8)
        self.assertEqual(
            reference["candidates"]["order_addition_permutation_optimization"][
                "delta_12h"
            ],
            13.1,
        )
        self.assertEqual(
            reference["candidates"]["schemathesis_datagen_pipeline"]["delta_2h"],
            13.4,
        )

    def test_comparison_workbook_uses_same_budget_gap_for_issue_marker(self) -> None:
        paper = EDGE.load_paper_reference()
        payload = {
            "campaign_id": "comparison-test",
            "matched_protocol": True,
            "paper_reference": paper,
            "finalized_at": "2026-07-29T00:00:00+00:00",
            "local_fast_reference": {
                "schema_version": 2,
                "reference": {
                    "label": "Local Codex + gpt-5.6-sol inclusive checkpoints",
                    "selection": "strict local checkpoint",
                    "official_comparison": False,
                },
                "task_count": 2,
                "checkpoints": {
                    "0.5h": {
                        "boundary_hours": 0.5,
                        "boundary_seconds": 1800,
                        "available_count": 1,
                        "tasks": {
                            "portfolio_risk_calibration": {
                                "task_id": "portfolio_risk_calibration",
                                "checkpoint_hours": 0.5,
                                "checkpoint_seconds": 1800,
                                "raw_score": 19.83,
                                "edgebench_score": 19.83,
                                "model": "gpt-5.6-sol",
                                "reasoning_effort": "medium",
                                "campaign_id": "fast-campaign",
                                "source": "runs/evidence.json",
                            }
                        },
                        "missing_tasks": {
                            "borden_source_inversion": [
                                {"status": "no_scored_submission"}
                            ]
                        },
                    },
                    "1h": {
                        "boundary_hours": 1,
                        "boundary_seconds": 3600,
                        "available_count": 1,
                        "tasks": {
                            "portfolio_risk_calibration": {
                                "task_id": "portfolio_risk_calibration",
                                "checkpoint_hours": 1,
                                "checkpoint_seconds": 3600,
                                "raw_score": 30,
                                "edgebench_score": 30,
                                "model": "gpt-5.6-sol",
                                "reasoning_effort": "medium",
                                "campaign_id": "fast-campaign",
                                "source": "runs/evidence-1h.json",
                            }
                        },
                        "missing_tasks": {
                            "borden_source_inversion": [
                                {"status": "no_scored_submission"}
                            ]
                        },
                    },
                },
            },
            "cells": [
                {
                    "task_id": "borden_source_inversion",
                    "method": "plain-codex",
                    "model": "gpt-5.6-sol",
                    "reasoning_effort": "medium",
                    "wall_time_seconds": 7200,
                    "live_search_concurrency": 1,
                    "completed_trajectories": 1,
                    "valid_trajectories": 1,
                    "observations": [],
                    "best": {
                        "raw_score": 78.502,
                        "edgebench_score": 78.502,
                        "official_comparison": {
                            "checkpoint_hours": 2,
                            "references": {"GPT-5.5": 38.5},
                        },
                    },
                },
                {
                    "task_id": "portfolio_risk_calibration",
                    "method": "plain-codex",
                    "model": "gpt-5.6-sol",
                    "reasoning_effort": "medium",
                    "wall_time_seconds": 7200,
                    "live_search_concurrency": 1,
                    "completed_trajectories": 1,
                    "valid_trajectories": 1,
                    "observations": [],
                    "best": {
                        "raw_score": 44.97,
                        "edgebench_score": 44.97,
                        "official_comparison": {
                            "checkpoint_hours": 2,
                            "references": {"GPT-5.5": 17.3},
                        },
                    },
                },
            ],
        }
        destination = self.temp / "comparison.xlsx"

        EDGE.write_comparison_workbook(payload, destination)

        workbook = load_workbook(destination, data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Overview", "Results", "Local Fast", "Protocol"],
        )
        results = workbook["Results"]
        headers = [cell.value for cell in results[1]]
        rows = {
            row[headers.index("Task")].value: {
                header: row[index].value for index, header in enumerate(headers)
            }
            for row in results.iter_rows(min_row=2)
        }
        portfolio = rows["portfolio_risk_calibration"]
        self.assertEqual(portfolio["Current budget (h)"], 2)
        self.assertEqual(portfolio["T (s)"], 7200)
        self.assertEqual(portfolio["Current EdgeBench 0-100"], 44.97)
        self.assertEqual(portfolio["Local <=0.5h best"], 19.83)
        self.assertAlmostEqual(portfolio["Delta vs local <=0.5h (pp)"], 25.14)
        self.assertEqual(portfolio["Local <=1h best"], 30)
        self.assertAlmostEqual(portfolio["Delta vs local <=1h (pp)"], 14.97)
        self.assertEqual(portfolio["GPT-5.5 checkpoint (h)"], 2)
        self.assertEqual(portfolio["GPT-5.5 same-budget"], 17.3)
        self.assertAlmostEqual(portfolio["Delta vs same-budget (pp)"], 27.67)
        self.assertEqual(portfolio["Paper Codex + GPT-5.5 @12h mean"], 25.0)
        self.assertEqual(portfolio["Paper sample stddev"], 6.5)
        self.assertAlmostEqual(portfolio["Delta vs paper 12h (pp)"], 19.97)
        self.assertEqual(portfolio["Issue marker"], "REVIEW_HIGH")
        self.assertIn("KNOWN_PROTOCOL", rows["borden_source_inversion"]["Issue marker"])
        self.assertEqual(results.freeze_panes, "A2")
        self.assertEqual(len(results.tables), 1)
        overview_values = {row[0].value: row[1].value for row in workbook["Overview"]}
        self.assertIn(
            "not an apples-to-apples", overview_values["Paper reference role"]
        )
        self.assertEqual(
            overview_values["Local fast coverage"],
            "<=0.5h: 1/2; <=1h: 1/2",
        )
        local_fast_rows = list(workbook["Local Fast"].iter_rows(values_only=True))
        self.assertEqual(
            local_fast_rows[1][0:3],
            (0.5, "portfolio_risk_calibration", "available"),
        )
        self.assertEqual(
            local_fast_rows[2][0:3],
            (0.5, "borden_source_inversion", "missing"),
        )

    def test_score_task_run_counts_game_sessions_without_run_history(self) -> None:
        task_run = self.temp / "game-run"
        task_run.mkdir()
        EDGE.write_json(
            task_run / "final_result.json",
            {
                "runtime_seconds": 120,
                "total_rounds": 2,
                "agent_submissions": 2,
                "auto_submissions": 0,
                "resume_count": 0,
                "timed_out": True,
            },
        )
        EDGE.write_json(
            task_run / "game_history.json",
            {
                "entries": [
                    {"type": "game", "round": "game-1"},
                    {"type": "game", "round": "game-2"},
                ]
            },
        )
        original_run_capture = EDGE_IO.run_capture
        EDGE_IO.run_capture = lambda *_args, **_kwargs: {
            "returncode": 0,
            "stdout": json.dumps(
                {
                    "source": str(task_run / "final_result.json"),
                    "edgebench_score": 25,
                }
            ),
            "stderr": "",
        }
        try:
            observation = EDGE.score_task_run(
                task_run,
                {"model": "gpt-test", "wall_time_seconds": 120},
            )
        finally:
            EDGE_IO.run_capture = original_run_capture

        self.assertEqual(observation["evaluator_calls"], 2)

    def test_prepare_encodes_plain_outer_and_goal_plus_inner_concurrency(self) -> None:
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=180,
            concurrency=2,
            model="gpt-test",
            reasoning_effort="high",
            campaign_id="unit-campaign",
        )

        profile = self.profile()
        profile.update(
            worker_runtime_seconds=120,
            worker_min_runtime_seconds=90,
            worker_min_verifier_runs=1,
            closeout_reserve_seconds=30,
            goal_plus_verifier_timeout_seconds=30,
        )
        destination = EDGE.prepare(args, profile)

        plain = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-codex"
                / "cell.json"
            ).read_text()
        )
        goal_plus = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--goal-plus-codex"
                / "cell.json"
            ).read_text()
        )
        self.assertEqual(plain["outer_replicas"], 2)
        self.assertEqual(plain["inner_search_concurrency"], 0)
        self.assertEqual(goal_plus["outer_replicas"], 1)
        self.assertEqual(goal_plus["inner_search_concurrency"], 2)
        self.assertEqual(goal_plus["worker_runtime_seconds"], 120)
        self.assertEqual(goal_plus["worker_min_runtime_seconds"], 90)
        self.assertEqual(goal_plus["worker_min_verifier_runs"], 1)
        self.assertEqual(goal_plus["closeout_reserve_seconds"], 30)
        self.assertEqual(goal_plus["goal_plus_verifier_timeout_seconds"], 30)
        self.assertFalse(goal_plus["shared_dir_enabled"])
        self.assertFalse(goal_plus["supplemental_evaluation_enabled"])
        self.assertNotIn("shared_dir_enabled", plain)
        self.assertNotIn("supplemental_evaluation_enabled", plain)
        self.assertEqual(
            goal_plus["goal_plus_config"],
            {
                "entrypoint": "$goal-plus",
                "command_config": {
                    "mode": "autonomous",
                    "max_parallel": 2,
                    "workspace_backend": "git_worktree",
                    "promotion_mode": "artifact_only",
                    "strategy": "agent_guided",
                    "workers": "gpt-test*2",
                    "annotator": "gpt-test",
                },
            },
        )
        self.assertFalse(plain["internet"])
        self.assertEqual(plain["eval_interval_seconds"], 1800)
        self.assertEqual(plain["submission_cooldown"], 120)
        self.assertEqual(plain["work_cpu_limit"], 4)
        self.assertEqual(plain["work_mem_limit"], "16g")
        self.assertEqual(plain["judge_cpu_limit"], 4)
        self.assertEqual(plain["judge_mem_limit"], "8g")
        self.assertTrue(plain["auto_eval_enabled"])
        self.assertTrue(plain["auto_resume_enabled"])
        self.assertTrue(plain["stop_hook_enabled"])
        self.assertEqual(
            {item["field"] for item in plain["protocol_diff"]},
            {
                "attempts_per_task",
                "backend",
                "cell_concurrency",
                "judge_concurrency",
                "model",
                "reasoning_effort",
                "timeout",
            },
        )
        self.assertEqual(
            {item["field"] for item in goal_plus["protocol_diff"]},
            {
                "agent",
                "attempts_per_task",
                "backend",
                "cell_concurrency",
                "judge_concurrency",
                "model",
                "reasoning_effort",
                "timeout",
            },
        )
        self.assertEqual(
            plain["protocol_source"]["sha256"],
            EDGE.sha256_file(EDGE.OFFICIAL_CODEX_PROTOCOL_PATH),
        )
        self.assertFalse(plain["official_edgebench_comparable"])
        self.assertNotIn("sk-xxxx", json.dumps(plain))
        self.assertEqual(
            json.loads((destination / "profile.json").read_text())["cell_concurrency"],
            1,
        )
        campaign = json.loads((destination / "campaign.json").read_text())
        self.assertFalse(campaign["shared_dir_enabled"])
        self.assertFalse(campaign["supplemental_evaluation_enabled"])
        self.assertFalse(
            any(path.name in {".gp", ".goal-plus"} for path in destination.rglob("*"))
        )

    def test_prepare_recomputes_overrides_and_routes_pi_role_models(self) -> None:
        _, profile = EDGE.load_profile("vliw-goal-plus-pi-zai-glm-5-2-1h-k2-c1")
        role_config = {
            "worker_model": "worker-provider/worker-model",
            "worker_reasoning_effort": "medium",
            "evidence_annotator_model": "annotation-provider/annotation-model",
            "evidence_annotator_reasoning_effort": "low",
            "evidence_annotator_timeout_seconds": 900,
        }
        profile.update(role_config)
        profile["global_evidence_mode"] = "independent"
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=1200,
            concurrency=1,
            cell_concurrency=1,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-dynamic-protocol-reasons",
        )

        models = self.temp / "role-models.json"
        models.write_text(
            json.dumps(
                {
                    "providers": {
                        "worker-provider": {
                            "baseUrl": "http://127.0.0.1:4101/v1",
                            "apiKey": "$WORKER_API_KEY",
                            "models": [{"id": "worker-model"}],
                        },
                        "annotation-provider": {
                            "baseUrl": "http://127.0.0.1:4102/v1",
                            "apiKey": "$ANNOTATION_API_KEY",
                            "models": [{"id": "annotation-model"}],
                        },
                    }
                }
            )
        )
        with mock.patch.dict(
            EDGE.os.environ,
            {
                "SFORGE_PI_MODELS_FILE": str(models),
                "ZAI_API_KEY": "zai-secret",
                "WORKER_API_KEY": "worker-secret",
                "ANNOTATION_API_KEY": "annotation-secret",
            },
            clear=False,
        ):
            destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--goal-plus-pi-provider"
                / "cell.json"
            ).read_text()
        )
        reasons = {item["field"]: item["reason"] for item in cell["protocol_diff"]}

        self.assertIn("T=1200", reasons["timeout"])
        self.assertIn("K=1", reasons["attempts_per_task"])
        self.assertIn("C=1", reasons["cell_concurrency"])
        self.assertIn("eval_interval=300", reasons["eval_interval"])
        serialized = json.dumps(cell)
        self.assertNotIn("one hour", serialized)
        self.assertNotIn("one-hour", serialized)
        self.assertNotIn("K=2", serialized)
        self.assertEqual(
            cell["goal_plus_config"],
            {
                "entrypoint": "/goal-plus",
                "command_config": {
                    "mode": "autonomous",
                    "max_parallel": 1,
                    "workspace_backend": "git_worktree",
                    "promotion_mode": "artifact_only",
                    "strategy": "agent_guided",
                    "workers": "worker-provider/worker-model*1",
                    "annotator": "annotation-provider/annotation-model",
                },
            },
        )
        campaign = EDGE.read_json(destination / "campaign.json")
        for payload in (cell, campaign):
            self.assertEqual(
                {key: payload[key] for key in role_config},
                role_config,
            )
            self.assertEqual(payload["global_evidence_mode"], "independent")
        snapshot = EDGE.read_json(destination / "profile.json")
        self.assertEqual(snapshot["global_evidence_mode"], "independent")

        models_path = self.temp / "runtime-models.json"
        env = EDGE.cell_environment(
            cell,
            api_base_urls=[
                "https://api.z.ai/api/coding/paas/v4",
                "http://127.0.0.1:4101/v1",
                "http://127.0.0.1:4102/v1",
            ],
            pi_models_file=models_path,
            pi_provider_credentials={"TEST_API_KEY": "test-secret"},
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        field_map = {
            "SFORGE_GOAL_PLUS_WORKER_MODEL": "worker_model",
            "SFORGE_GOAL_PLUS_WORKER_REASONING_EFFORT": "worker_reasoning_effort",
            "GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL": "evidence_annotator_model",
            "GOAL_PLUS_EVIDENCE_ANNOTATOR_REASONING_EFFORT": (
                "evidence_annotator_reasoning_effort"
            ),
            "SFORGE_GOAL_PLUS_EVIDENCE_ANNOTATOR_TIMEOUT_SECONDS": (
                "evidence_annotator_timeout_seconds"
            ),
        }
        for env_name, profile_name in field_map.items():
            self.assertEqual(extra[env_name], str(role_config[profile_name]))
        self.assertEqual(extra["GOAL_PLUS_GLOBAL_EVIDENCE_MODE"], "independent")
        self.assertEqual(
            extra["SFORGE_PI_AUX_MODELS"],
            "worker-provider/worker-model annotation-provider/annotation-model",
        )
        self.assertEqual(env["SFORGE_PI_MODELS_FILE"], str(models_path))
        self.assertEqual(env["TEST_API_KEY"], "test-secret")

    def test_stop_default_wait_covers_normal_controller_closeout(self) -> None:
        parsed = EDGE.build_parser().parse_args(["stop", "--campaign", "campaign"])
        self.assertEqual(parsed.wait_seconds, 60)

        campaign = self.temp / "stop-campaign"
        campaign.mkdir()
        EDGE.write_json(
            campaign / "controller.json",
            {"pid": 12345, "pgid": 12345, "state": "running"},
        )
        clock = {"now": 0.0}

        def process_alive(_pid):
            return clock["now"] < 25.0

        def advance(seconds):
            clock["now"] += seconds

        with (
            mock.patch.object(EDGE_RUNTIME, "process_alive", side_effect=process_alive),
            mock.patch.object(EDGE_RUNTIME.os, "kill"),
            mock.patch.object(
                EDGE_RUNTIME.time,
                "monotonic",
                side_effect=lambda: clock["now"],
            ),
            mock.patch.object(EDGE_RUNTIME.time, "sleep", side_effect=advance),
        ):
            result = EDGE_RUNTIME.stop_campaign(campaign, wait_seconds=60)

        self.assertEqual(result, 0)
        self.assertGreaterEqual(clock["now"], 25.0)
        self.assertLess(clock["now"], 60.0)

    def test_stop_still_reports_timeout_when_controller_exceeds_wait(self) -> None:
        campaign = self.temp / "slow-stop-campaign"
        campaign.mkdir()
        EDGE.write_json(
            campaign / "controller.json",
            {"pid": 54321, "pgid": 54321, "state": "running"},
        )
        clock = {"now": 0.0}

        def advance(seconds):
            clock["now"] += seconds

        with (
            mock.patch.object(EDGE_RUNTIME, "process_alive", return_value=True),
            mock.patch.object(EDGE_RUNTIME.os, "kill"),
            mock.patch.object(
                EDGE_RUNTIME.time,
                "monotonic",
                side_effect=lambda: clock["now"],
            ),
            mock.patch.object(EDGE_RUNTIME.time, "sleep", side_effect=advance),
        ):
            result = EDGE_RUNTIME.stop_campaign(campaign, wait_seconds=60)

        self.assertEqual(result, 2)
        self.assertGreaterEqual(clock["now"], 60.0)

    def test_prepare_applies_local_smoke_network_override_with_provenance(self) -> None:
        _, profile = EDGE.load_profile("vliw-codex-sol-medium-local-smoke")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-local-codex-smoke",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-codex"
                / "cell.json"
            ).read_text()
        )
        command = EDGE.build_sforge_command(destination, cell)

        self.assertFalse(cell["internet"])
        self.assertEqual(cell["eval_interval_seconds"], 60)
        self.assertEqual(
            cell["internet_source"],
            "profiles/vliw-codex-sol-medium-local-smoke.protocol_overrides.internet",
        )
        self.assertIn(
            "eval_interval", {item["field"] for item in cell["protocol_diff"]}
        )
        self.assertFalse(cell["official_edgebench_comparable"])
        self.assertEqual(
            command[command.index("--eval-interval") + 1],
            "60",
        )
        self.assertIn("--disable-internet", command)
        self.assertNotIn("--enable-internet", command)

    def test_prepare_encodes_plain_claude_api_and_thinking_contract(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-2-high-20m-k1")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-claude-campaign",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-claude"
                / "cell.json"
            ).read_text()
        )

        self.assertEqual(cell["sforge_agent"], "claude-code")
        self.assertEqual(cell["api_protocol"], "anthropic")
        self.assertEqual(cell["thinking"], {"type": "enabled"})
        self.assertEqual(cell["reasoning_effort"], "high")
        self.assertEqual(cell["outer_replicas"], 1)
        self.assertEqual(cell["inner_search_concurrency"], 0)

    def test_prepare_preserves_adaptive_claude_without_effort(self) -> None:
        _, profile = EDGE.load_profile("vliw-glm-5-1-adaptive-2h-k1")
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=None,
            concurrency=None,
            cell_concurrency=None,
            model=None,
            reasoning_effort=None,
            campaign_id="unit-claude-adaptive-campaign",
        )

        destination = EDGE.prepare(args, profile)
        cell = json.loads(
            (
                destination
                / "cells"
                / "vliw_kernel_optimization--plain-claude"
                / "cell.json"
            ).read_text()
        )

        self.assertEqual(cell["thinking"], {"type": "adaptive"})
        self.assertIsNone(cell["reasoning_effort"])
        self.assertEqual(cell["claude_context_window_tokens"], 200000)
        self.assertEqual(cell["claude_autocompact_percent"], 80)
        self.assertNotIn(
            "reasoning_effort",
            {item["field"] for item in cell["protocol_diff"]},
        )

    def test_cell_queue_limits_parallel_cells_and_continues_after_failure(self) -> None:
        destination = self.temp / "campaign"
        destination.mkdir()
        EDGE.write_json(destination / "controller.json", {"state": "running"})
        campaign = {
            "cells": [
                {"cell_id": cell_id, "task_id": cell_id} for cell_id in ("a", "b", "c")
            ]
        }
        started = []
        processes = []
        live = 0
        max_live = 0

        class FakeProcess:
            def __init__(self, pid):
                self.pid = pid
                self.done = False

            def poll(self):
                return 0 if self.done else None

        def fake_start(_destination, summary, **_kwargs):
            nonlocal live, max_live
            process = FakeProcess(100 + len(processes))
            processes.append(process)
            started.append(summary["cell_id"])
            live += 1
            max_live = max(max_live, live)
            return {
                "cell": {
                    "cell_id": summary["cell_id"],
                    "task_id": summary["task_id"],
                    "started_at": "now",
                },
                "process": process,
            }

        def fake_finish(_destination, running, *, stop_requested):
            nonlocal live
            self.assertFalse(stop_requested)
            live -= 1
            return 1 if running["cell"]["cell_id"] == "b" else 0

        sleeps = 0

        def fake_sleep(_seconds):
            nonlocal sleeps
            sleeps += 1
            if sleeps == 1:
                processes[0].done = True
            else:
                for process in processes:
                    process.done = True

        original_start = EDGE_RUNTIME.start_campaign_cell
        original_finish = EDGE_RUNTIME.finish_campaign_cell
        original_sleep = EDGE.time.sleep
        EDGE_RUNTIME.start_campaign_cell = fake_start
        EDGE_RUNTIME.finish_campaign_cell = fake_finish
        EDGE.time.sleep = fake_sleep
        try:
            returncode = EDGE.execute_cell_queue(
                destination,
                campaign,
                {"state": "running"},
                cell_concurrency=2,
                judge_container_url="http://judge",
                api_config={"api_key_source": None, "api_base_url_source": None},
                api_key=None,
                runtime_api_base_url=None,
                runtime_api_base_urls=["https://api.openai.com"],
                bridge_host=None,
                stop_requested=lambda: False,
            )
        finally:
            EDGE_RUNTIME.start_campaign_cell = original_start
            EDGE_RUNTIME.finish_campaign_cell = original_finish
            EDGE.time.sleep = original_sleep

        self.assertEqual(returncode, 1)
        self.assertEqual(started, ["a", "b", "c"])
        self.assertEqual(max_live, 2)
        self.assertEqual(
            json.loads((destination / "controller.json").read_text())[
                "active_children"
            ],
            {},
        )

    def test_cell_queue_stop_interrupts_active_cells_without_starting_more(
        self,
    ) -> None:
        destination = self.temp / "campaign"
        destination.mkdir()
        EDGE.write_json(destination / "controller.json", {"state": "running"})
        campaign = {
            "cells": [
                {"cell_id": cell_id, "task_id": cell_id} for cell_id in ("a", "b", "c")
            ]
        }
        started = []
        processes = []
        requested = False

        class FakeProcess:
            def __init__(self, pid):
                self.pid = pid
                self.done = False
                self.signals = []

            def poll(self):
                return 130 if self.done else None

            def send_signal(self, value):
                self.signals.append(value)
                self.done = True

        def fake_start(_destination, summary, **_kwargs):
            process = FakeProcess(200 + len(processes))
            processes.append(process)
            started.append(summary["cell_id"])
            return {
                "cell": {
                    "cell_id": summary["cell_id"],
                    "task_id": summary["task_id"],
                    "started_at": "now",
                },
                "process": process,
            }

        def fake_finish(_destination, _running, *, stop_requested):
            self.assertTrue(stop_requested)
            return 130

        def fake_sleep(_seconds):
            nonlocal requested
            requested = True

        original_start = EDGE_RUNTIME.start_campaign_cell
        original_finish = EDGE_RUNTIME.finish_campaign_cell
        original_sleep = EDGE.time.sleep
        EDGE_RUNTIME.start_campaign_cell = fake_start
        EDGE_RUNTIME.finish_campaign_cell = fake_finish
        EDGE.time.sleep = fake_sleep
        try:
            returncode = EDGE.execute_cell_queue(
                destination,
                campaign,
                {"state": "running"},
                cell_concurrency=2,
                judge_container_url="http://judge",
                api_config={"api_key_source": None, "api_base_url_source": None},
                api_key=None,
                runtime_api_base_url=None,
                runtime_api_base_urls=["https://api.openai.com"],
                bridge_host=None,
                stop_requested=lambda: requested,
            )
        finally:
            EDGE_RUNTIME.start_campaign_cell = original_start
            EDGE_RUNTIME.finish_campaign_cell = original_finish
            EDGE.time.sleep = original_sleep

        self.assertEqual(returncode, 130)
        self.assertEqual(started, ["a", "b"])
        self.assertEqual(
            [process.signals for process in processes],
            [[EDGE.signal.SIGINT], [EDGE.signal.SIGINT]],
        )

    def test_goal_plus_environment_uses_pinned_source_and_configured_k(self) -> None:
        cell = {
            "method": "goal-plus-codex",
            "model": "gpt-5.6-sol",
            "reasoning_effort": "xhigh",
            "internet": False,
            "inner_search_concurrency": 4,
            "worker_runtime_seconds": 600,
            "goal_plus_verifier_timeout_seconds": 45,
            "goal_plus_finalization_grace_seconds": 90,
        }

        env = EDGE.cell_environment(
            cell,
            api_key="runtime-key",
            api_base_url="http://192.0.2.10:45678/v1",
        )

        self.assertEqual(
            env["SFORGE_GOAL_PLUS_SOURCE_DIR"],
            str(EDGE.current_paths().goal_plus_root),
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "4")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "600")
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_VERIFIER_TIMEOUT_SECONDS"], "45"
        )
        self.assertEqual(extra["GOAL_PLUS_GLOBAL_EVIDENCE_MODE"], "manual")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_SHARED_DIR_ENABLED"], "false")
        self.assertEqual(
            extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_ENABLED"], "0"
        )
        self.assertEqual(
            extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_REQUIRED"], "0"
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_FINALIZATION_GRACE_SECONDS"], "90")
        self.assertEqual(extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL"], "gpt-5.6-sol")
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_REASONING_EFFORT"], "xhigh"
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_API_KEY_ENV"],
            "SFORGE_AGENT_API_KEY",
        )
        for key in ("TMPDIR", "TMP", "TEMP"):
            self.assertTrue(Path(env[key]).is_relative_to(ROOT))

    def test_goal_plus_environment_overrides_ambient_feature_flags(self) -> None:
        ambient = {
            "SFORGE_AGENT_EXTRA_ENV": ",".join(
                (
                    "SFORGE_GOAL_PLUS_SHARED_DIR_ENABLED=true",
                    "GOAL_PLUS_SUPPLEMENTAL_EVALUATION_ENABLED=1",
                    "GOAL_PLUS_SUPPLEMENTAL_EVALUATION_REQUIRED=1",
                )
            )
        }
        cell = {
            "method": "goal-plus-pi-provider",
            "sforge_agent": "pi-goal-plus-provider",
            "model": "test-provider/test-model",
            "reasoning_effort": "medium",
            "internet": False,
            "inner_search_concurrency": 2,
            "worker_runtime_seconds": 240,
            "shared_dir_enabled": False,
            "supplemental_evaluation_enabled": False,
        }

        with mock.patch.dict(os.environ, ambient, clear=False):
            env = EDGE.cell_environment(
                cell,
                api_base_urls=["https://api.example.invalid/v1"],
            )

        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_SHARED_DIR_ENABLED"], "false")
        self.assertEqual(
            extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_ENABLED"], "0"
        )
        self.assertEqual(
            extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_REQUIRED"], "0"
        )

        enabled_cell = {
            **cell,
            "shared_dir_enabled": True,
            "supplemental_evaluation_enabled": True,
        }
        enabled_env = EDGE.cell_environment(
            enabled_cell,
            api_base_urls=["https://api.example.invalid/v1"],
        )
        enabled_extra = dict(
            item.split("=", 1)
            for item in enabled_env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(
            enabled_extra["SFORGE_GOAL_PLUS_SHARED_DIR_ENABLED"], "true"
        )
        self.assertEqual(
            enabled_extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_ENABLED"], "1"
        )
        self.assertEqual(
            enabled_extra["GOAL_PLUS_SUPPLEMENTAL_EVALUATION_REQUIRED"], "1"
        )

    def test_goal_plus_pi_environment_uses_the_same_runtime_contract(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "goal-plus-pi",
                "sforge_agent": "pi-goal-plus",
                "model": "gpt-5.6-sol",
                "reasoning_effort": "medium",
                "pi_package_version": "0.83.0",
                "internet": False,
                "inner_search_concurrency": 2,
                "worker_runtime_seconds": 240,
                "worker_min_runtime_seconds": 180,
                "worker_min_verifier_runs": 1,
                "closeout_reserve_seconds": 60,
                "goal_plus_finalization_grace_seconds": 120,
            }
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )

        self.assertEqual(
            env["SFORGE_GOAL_PLUS_SOURCE_DIR"],
            str(EDGE.current_paths().goal_plus_root),
        )
        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "2")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "240")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_MIN_RUNTIME_SECONDS"], "180")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_MIN_VERIFIER_RUNS"], "1")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_CLOSEOUT_RESERVE_SECONDS"], "60")
        self.assertEqual(extra["SFORGE_PI_REASONING_EFFORT"], "medium")
        self.assertEqual(extra["SFORGE_PI_PACKAGE_VERSION"], "0.83.0")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_FINALIZATION_GRACE_SECONDS"], "120")

    def test_goal_plus_pi_provider_environment_uses_goal_plus_contract(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "goal-plus-pi-provider",
                "sforge_agent": "pi-goal-plus-provider",
                "model": "glm-proxy/GLM-5.2",
                "reasoning_effort": "high",
                "internet": False,
                "inner_search_concurrency": 2,
                "worker_runtime_seconds": 3300,
                "goal_plus_finalization_grace_seconds": 300,
            },
            api_base_urls=["https://api.z.ai/api/coding/paas/v4"],
        )
        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )

        self.assertEqual(extra["SFORGE_GOAL_PLUS_PARALLEL_NUM"], "2")
        self.assertEqual(extra["SFORGE_GOAL_PLUS_WORKER_RUNTIME_SECONDS"], "3300")
        self.assertEqual(extra["SFORGE_PI_REASONING_EFFORT"], "high")
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_WORKER_MODEL"],
            "glm-proxy/GLM-5.2",
        )
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_WORKER_REASONING_EFFORT"],
            "high",
        )
        self.assertEqual(
            extra["SFORGE_GOAL_PLUS_EVIDENCE_ANNOTATOR_TIMEOUT_SECONDS"],
            "1800",
        )
        self.assertEqual(
            extra["GOAL_PLUS_EVIDENCE_ANNOTATOR_MODEL"],
            "GLM-5.2",
        )
        self.assertNotIn("SFORGE_PI_AUX_MODELS", extra)

    def test_api_config_prefers_sforge_then_openai_then_codex(self) -> None:
        config = EDGE.resolve_agent_api_config(
            {
                "SFORGE_AGENT_API_KEY": "sforge-key",
                "SFORGE_AGENT_API_BASE_URL": "https://sforge.example/v1",
                "OPENAI_API_KEY": "openai-key",
                "OPENAI_BASE_URL": "https://openai.example/v1",
                "CODEX_API_KEY": "codex-key",
            }
        )

        self.assertEqual(config["api_key"], "sforge-key")
        self.assertEqual(config["api_key_source"], "SFORGE_AGENT_API_KEY")
        self.assertEqual(config["api_base_url"], "https://sforge.example/v1")

        fallback = EDGE.resolve_agent_api_config(
            {
                "OPENAI_API_KEY": "openai-key",
                "OPENAI_BASE_URL": "http://127.0.0.1:3788/v1",
                "CODEX_API_KEY": "codex-key",
            }
        )
        self.assertEqual(fallback["api_key"], "openai-key")
        self.assertEqual(fallback["api_key_source"], "OPENAI_API_KEY")
        self.assertEqual(fallback["api_base_url_source"], "OPENAI_BASE_URL")

        anthropic = EDGE.resolve_agent_api_config(
            {
                "ANTHROPIC_AUTH_TOKEN": "anthropic-token",
                "ANTHROPIC_BASE_URL": "https://anthropic.example",
                "OPENAI_API_KEY": "wrong-protocol-key",
            },
            protocol="anthropic",
        )
        self.assertEqual(anthropic["api_key"], "anthropic-token")
        self.assertEqual(anthropic["api_key_source"], "ANTHROPIC_AUTH_TOKEN")
        self.assertEqual(anthropic["api_base_url"], "https://anthropic.example")
        self.assertEqual(
            EDGE.agent_api_probe_url(
                "https://anthropic.example/api/anthropic", "anthropic"
            ),
            "https://anthropic.example/api/anthropic/v1/messages",
        )

        provider = EDGE.resolve_agent_api_config(
            {
                "OPENAI_API_KEY": "must-not-be-used",
                "ANTHROPIC_API_KEY": "must-not-be-used",
            },
            protocol="pi-provider",
        )
        self.assertIsNone(provider["api_key"])
        self.assertIsNone(provider["api_base_url"])

    def test_pi_auth_requires_an_openai_codex_login(self) -> None:
        auth = self.temp / "pi-auth.json"
        auth.write_text(json.dumps({"other-provider": {}}))
        self.assertFalse(
            EDGE.resolve_pi_auth({"SFORGE_PI_AUTH_FILE": str(auth)})["valid"]
        )

        auth.write_text(json.dumps({"openai-codex": {"type": "oauth"}}))
        status = EDGE.resolve_pi_auth({"SFORGE_PI_AUTH_FILE": str(auth)})
        self.assertTrue(status["valid"])
        self.assertEqual(status["path"], auth)

    def test_pi_oauth_accepts_an_api_base_override_without_api_key(self) -> None:
        report = EDGE_ENV.DoctorReport("pi-oauth-proxy")
        profile = {"methods": ["goal-plus-pi"], "model": "gpt-5.6-sol"}

        with mock.patch.dict(
            os.environ,
            {
                "SFORGE_AGENT_API_BASE_URL": (
                    "http://host.docker.internal:19090/backend-api"
                )
            },
            clear=True,
        ), mock.patch.object(
            EDGE_ENV,
            "resolve_pi_auth",
            return_value={"valid": True, "path": "/tmp/pi-auth.json"},
        ):
            protocol, api_config, host_preflight_ready = EDGE_ENV._check_auth(
                report, profile
            )

        auth_check = next(
            check for check in report.checks if check["name"] == "auth:agent"
        )
        self.assertTrue(auth_check["passed"])
        self.assertEqual(auth_check["mode"], "pi-oauth")
        self.assertEqual(protocol, "openai")
        self.assertEqual(
            api_config["api_base_url"],
            "http://host.docker.internal:19090/backend-api",
        )
        self.assertTrue(host_preflight_ready)

    def test_pi_provider_validates_registry_model_and_credential_env(self) -> None:
        models = self.temp / "models.json"
        models.write_text(
            json.dumps(
                {
                    "providers": {
                        "glm-anthropic": {
                            "api": "anthropic-messages",
                            "apiKey": "${GLM_PROXY_API_KEY}",
                            "models": [{"id": "GLM-5.2"}],
                        },
                        "glm-openai": {
                            "api": "openai-completions",
                            "apiKey": "$GLM_PROXY_API_KEY",
                            "models": [{"id": "GLM-5.2"}],
                        },
                    }
                }
            )
        )
        env = {
            "SFORGE_PI_MODELS_FILE": str(models),
            "GLM_PROXY_API_KEY": "secret-value",
        }

        for provider in ("glm-anthropic", "glm-openai"):
            with self.subTest(provider=provider):
                status = EDGE.resolve_pi_provider(f"{provider}/GLM-5.2", env)
                self.assertTrue(status["valid"])
                self.assertEqual(status["provider"], provider)
                self.assertEqual(status["model"], "GLM-5.2")
                self.assertEqual(status["credential_env"], "GLM_PROXY_API_KEY")
                self.assertNotIn("secret-value", json.dumps(status))

        missing = EDGE.resolve_pi_provider(
            "glm-anthropic/GLM-5.2",
            {"SFORGE_PI_MODELS_FILE": str(models)},
        )
        self.assertFalse(missing["valid"])
        self.assertEqual(missing["error"], "missing GLM_PROXY_API_KEY")

    def test_pi_provider_rejects_literal_or_bare_api_key(self) -> None:
        for api_key in ("GLM_PROXY_API_KEY", "literal-secret"):
            with self.subTest(api_key=api_key):
                models = self.temp / f"models-{len(api_key)}.json"
                models.write_text(
                    json.dumps(
                        {
                            "providers": {
                                "custom": {
                                    "apiKey": api_key,
                                    "models": [{"id": "model"}],
                                }
                            }
                        }
                    )
                )
                status = EDGE.resolve_pi_provider(
                    "custom/model", {"SFORGE_PI_MODELS_FILE": str(models)}
                )
                self.assertFalse(status["valid"])
                self.assertIn("$NAME", status["error"])
                self.assertNotIn("literal-secret", json.dumps(status))

    def test_pi_provider_requires_custom_api_key_reference(self) -> None:
        models = self.temp / "models-missing-api-key.json"
        models.write_text(
            json.dumps({"providers": {"custom": {"models": [{"id": "model"}]}}})
        )

        status = EDGE.resolve_pi_provider(
            "custom/model", {"SFORGE_PI_MODELS_FILE": str(models)}
        )

        self.assertFalse(status["valid"])
        self.assertIn("apiKey as $NAME", status["error"])

    def test_pi_provider_uses_builtin_deepseek_environment(self) -> None:
        status = EDGE.resolve_pi_provider(
            "deepseek/deepseek-chat",
            {"DEEPSEEK_API_KEY": "secret-value"},
        )

        self.assertTrue(status["valid"])
        self.assertEqual(status["credential_env"], "DEEPSEEK_API_KEY")
        self.assertEqual(status["api_base_url"], "https://api.deepseek.com")
        self.assertNotIn("secret-value", json.dumps(status))

    def test_pi_provider_uses_builtin_zai_endpoint_for_offline_allowlist(self) -> None:
        status = EDGE.resolve_pi_provider(
            "zai/glm-5.2",
            {"ZAI_API_KEY": "secret-value"},
        )

        self.assertTrue(status["valid"])
        self.assertEqual(status["credential_env"], "ZAI_API_KEY")
        self.assertEqual(
            status["api_base_url"],
            "https://api.z.ai/api/coding/paas/v4",
        )
        self.assertNotIn("secret-value", json.dumps(status))

    def test_pi_builtin_openai_and_anthropic_endpoints_are_resolvable(self) -> None:
        cases = (
            ("openai/gpt-5.6-sol", "OPENAI_API_KEY", "api.openai.com"),
            ("anthropic/claude-sonnet-5", "ANTHROPIC_API_KEY", "api.anthropic.com"),
        )
        for model_ref, key_name, host in cases:
            with self.subTest(model_ref=model_ref):
                status = EDGE.resolve_pi_provider(model_ref, {key_name: "secret"})
                self.assertTrue(status["valid"])
                self.assertEqual(
                    EDGE_ENV.sanitized_api_endpoint(status["api_base_url"])["host"],
                    host,
                )

    def test_pi_host_probe_uses_external_registry_and_real_tool_roundtrip(self) -> None:
        models = self.temp / "external-models.json"
        registry = {
            "providers": {
                "dynamic-provider": {
                    "baseUrl": "http://127.0.0.1:43123/custom/base",
                    "api": "openai-responses",
                    "apiKey": "$DYNAMIC_API_KEY",
                    "models": [{"id": "dynamic-model", "reasoning": True}],
                }
            }
        }
        EDGE.write_json(models, registry)
        captured: dict[str, object] = {}

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            if command[-1] == "--version":
                return {"returncode": 0, "stdout": "0.84.1", "stderr": ""}
            captured["command"] = command
            captured["env"] = dict(env or {})
            copied = Path(str(env["PI_CODING_AGENT_DIR"])) / "models.json"
            captured["registry"] = json.loads(copied.read_text(encoding="utf-8"))
            stdout = "\n".join(
                [
                    json.dumps(
                        {
                            "type": "message_end",
                            "message": {
                                "role": "assistant",
                                "provider": "dynamic-provider",
                                "model": "dynamic-model",
                                "api": "openai-responses",
                                "content": [{"type": "toolCall", "name": "read"}],
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "message_end",
                            "message": {"role": "toolResult", "content": []},
                        }
                    ),
                    json.dumps(
                        {
                            "type": "message_end",
                            "message": {
                                "role": "assistant",
                                "provider": "dynamic-provider",
                                "model": "dynamic-model",
                                "api": "openai-responses",
                                "content": [{"type": "text", "text": "PI_HOST_API_OK"}],
                            },
                        }
                    ),
                ]
            )
            return {"returncode": 0, "stdout": stdout, "stderr": ""}

        probe_env = {
            "SFORGE_PI_MODELS_FILE": str(models),
            "SFORGE_PI_HOST_EXECUTABLE": "/opt/pi/bin/pi",
            "DYNAMIC_API_KEY": "rotating-secret-value",
        }
        with mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.pi_host_provider_probe(
                "dynamic-provider/dynamic-model",
                reasoning_effort="medium",
                expected_pi_version="0.84.1",
                env=probe_env,
            )

        command = captured["command"]
        self.assertTrue(result["passed"])
        self.assertEqual(captured["registry"], registry)
        self.assertEqual(command[command.index("--provider") + 1], "dynamic-provider")
        self.assertEqual(command[command.index("--model") + 1], "dynamic-model")
        self.assertEqual(command[command.index("--thinking") + 1], "medium")
        self.assertEqual(result["wire_apis"], ["openai-responses"])
        self.assertTrue(result["tool_roundtrip"])
        self.assertNotIn("rotating-secret-value", json.dumps(result))
        self.assertNotIn("rotating-secret-value", json.dumps(command))

    def test_codex_host_probe_uses_dynamic_provider_and_real_tool_roundtrip(
        self,
    ) -> None:
        captured: dict[str, object] = {}

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            if command[-1] == "--version":
                return {
                    "returncode": 0,
                    "stdout": "codex-cli 0.150.1",
                    "stderr": "",
                }
            captured["command"] = command
            captured["env"] = dict(env or {})
            config_path = Path(str(env["CODEX_HOME"])) / "config.toml"
            captured["config"] = config_path.read_text(encoding="utf-8")
            stdout = "\n".join(
                [
                    json.dumps({"type": "turn.started"}),
                    json.dumps(
                        {
                            "type": "item.started",
                            "item": {
                                "id": "item-1",
                                "type": "command_execution",
                                "status": "in_progress",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "item-1",
                                "type": "command_execution",
                                "status": "completed",
                                "exit_code": 0,
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "item-2",
                                "type": "agent_message",
                                "text": "CODEX_HOST_API_OK",
                            },
                        }
                    ),
                    json.dumps({"type": "turn.completed", "usage": {}}),
                ]
            )
            return {"returncode": 0, "stdout": stdout, "stderr": ""}

        probe_env = {
            "SFORGE_CODEX_HOST_EXECUTABLE": "/opt/codex/bin/codex",
            "OPENAI_API_KEY": "rotating-secret-value",
            "OPENAI_BASE_URL": "http://127.0.0.1:43123/changing/base",
        }
        with mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.codex_host_provider_probe(
                {
                    "model": "gpt-5.6-sol",
                    "reasoning_effort": "medium",
                },
                env=probe_env,
            )

        command = captured["command"]
        config = str(captured["config"])
        self.assertTrue(result["passed"])
        self.assertTrue(result["tool_roundtrip"])
        self.assertTrue(result["turn_completed"])
        self.assertEqual(result["contract"]["provider"], "sforge-proxy")
        self.assertIn(
            'base_url = "http://127.0.0.1:43123/changing/base"', config
        )
        self.assertIn('env_key = "OPENAI_API_KEY"', config)
        self.assertIn("stream_idle_timeout_ms = 60000", config)
        self.assertIn("stream_max_retries = 2", config)
        self.assertIn("request_max_retries = 2", config)
        self.assertEqual(command[command.index("--model") + 1], "gpt-5.6-sol")
        self.assertEqual(command[command.index("--disable") + 1], "plugins")
        self.assertNotIn("http://127.0.0.1:43123/changing/base", command)
        self.assertNotIn("rotating-secret-value", config)
        self.assertNotIn("rotating-secret-value", json.dumps(result))
        self.assertNotIn("rotating-secret-value", json.dumps(command))

    def test_codex_goal_plus_host_probe_requires_real_mcp_roundtrip(self) -> None:
        captured: dict[str, object] = {}
        source_dir = self.test_paths.goal_plus_root
        resolved_source = {
            "valid": True,
            "source_kind": "external",
            "source_dir": str(source_dir),
            "source_path": "goal-plus",
            "expected_ref": "experiment/test",
            "branch": "experiment/test",
            "commit": "a" * 40,
            "error": None,
        }

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            if command[-1] == "--version":
                return {
                    "returncode": 0,
                    "stdout": "codex-cli 0.150.1",
                    "stderr": "",
                }
            captured["command"] = command
            captured["env"] = dict(env or {})
            config_path = Path(str(env["CODEX_HOME"])) / "config.toml"
            captured["config"] = config_path.read_text(encoding="utf-8")
            stdout = "\n".join(
                [
                    json.dumps({"type": "turn.started"}),
                    json.dumps(
                        {
                            "type": "item.started",
                            "item": {
                                "id": "mcp-1",
                                "type": "mcp_tool_call",
                                "server": "goal-plus",
                                "tool": "goal_plus_monitor_snapshot",
                                "status": "in_progress",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "mcp-1",
                                "type": "mcp_tool_call",
                                "server": "goal-plus",
                                "tool": "goal_plus_monitor_snapshot",
                                "status": "completed",
                                "result": {"ok": True},
                                "error": None,
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.started",
                            "item": {
                                "id": "mcp-2",
                                "type": "mcp_tool_call",
                                "server": "goal-plus",
                                "tool": "goal_plus_monitor_snapshot",
                                "status": "in_progress",
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "mcp-2",
                                "type": "mcp_tool_call",
                                "server": "goal-plus",
                                "tool": "goal_plus_monitor_snapshot",
                                "status": "completed",
                                "result": {"ok": True},
                                "error": None,
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "message-1",
                                "type": "agent_message",
                                "text": "GOAL_PLUS_MCP_HOST_OK",
                            },
                        }
                    ),
                    json.dumps({"type": "turn.completed", "usage": {}}),
                ]
            )
            return {"returncode": 0, "stdout": stdout, "stderr": ""}

        probe_env = {
            "SFORGE_CODEX_HOST_EXECUTABLE": "/opt/codex/bin/codex",
            "SFORGE_GOAL_PLUS_HOST_EXECUTABLE": "/opt/goal-plus/bin/goal-plus",
            "OPENAI_API_KEY": "rotating-secret-value",
            "OPENAI_BASE_URL": "https://changing.example/v1",
        }
        with mock.patch.object(
            EDGE_ENV, "resolve_goal_plus_source", return_value=resolved_source
        ), mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.codex_host_provider_probe(
                {
                    "methods": ["goal-plus-codex"],
                    "model": "gpt-5.5",
                    "reasoning_effort": "high",
                    "goal_plus_source": {
                        "source_dir": str(source_dir),
                        "expected_ref": "experiment/test",
                        "commit": "a" * 40,
                    },
                },
                env=probe_env,
            )

        command = captured["command"]
        config = str(captured["config"])
        probe_python_path = str(captured["env"]["PYTHONPATH"])
        self.assertTrue(result["passed"])
        self.assertTrue(result["mcp_tool_roundtrip"])
        self.assertEqual(
            result["mcp_tools"], ["goal-plus:goal_plus_monitor_snapshot"]
        )
        self.assertEqual(
            result["mcp_tool_completion_counts"],
            {"goal-plus:goal_plus_monitor_snapshot": 2},
        )
        self.assertTrue(result["goal_plus_mcp_required"])
        self.assertEqual(result["goal_plus_source"]["commit"], "a" * 40)
        self.assertIn("[mcp_servers.goal-plus]", config)
        self.assertIn(
            'command = "/opt/goal-plus/bin/goal-plus"', config
        )
        self.assertIn(str(source_dir / "src"), probe_python_path.split(os.pathsep))
        self.assertIn("goal_plus_monitor_snapshot exactly twice", command[-1])
        self.assertNotIn("rotating-secret-value", json.dumps(result))

    def test_codex_goal_plus_host_probe_rejects_generic_tool_roundtrip(self) -> None:
        resolved_source = {
            "valid": True,
            "source_kind": "external",
            "source_dir": str(self.test_paths.goal_plus_root),
            "source_path": "goal-plus",
            "expected_ref": "experiment/test",
            "branch": "experiment/test",
            "commit": "b" * 40,
            "error": None,
        }

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            if command[-1] == "--version":
                return {
                    "returncode": 0,
                    "stdout": "codex-cli 0.150.1",
                    "stderr": "",
                }
            stdout = "\n".join(
                [
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "shell-1",
                                "type": "command_execution",
                                "status": "completed",
                                "exit_code": 0,
                            },
                        }
                    ),
                    json.dumps(
                        {
                            "type": "item.completed",
                            "item": {
                                "id": "message-1",
                                "type": "agent_message",
                                "text": "GOAL_PLUS_MCP_HOST_OK",
                            },
                        }
                    ),
                    json.dumps({"type": "turn.completed", "usage": {}}),
                ]
            )
            return {"returncode": 0, "stdout": stdout, "stderr": ""}

        with mock.patch.object(
            EDGE_ENV, "resolve_goal_plus_source", return_value=resolved_source
        ), mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.codex_host_provider_probe(
                {
                    "methods": ["goal-plus-codex"],
                    "model": "gpt-5.5",
                    "reasoning_effort": "high",
                },
                env={
                    "SFORGE_CODEX_HOST_EXECUTABLE": "/opt/codex/bin/codex",
                    "SFORGE_GOAL_PLUS_HOST_EXECUTABLE": (
                        "/opt/goal-plus/bin/goal-plus"
                    ),
                    "OPENAI_API_KEY": "secret",
                    "OPENAI_BASE_URL": "https://changing.example/v1",
                },
            )

        self.assertFalse(result["passed"])
        self.assertTrue(result["tool_roundtrip"])
        self.assertFalse(result["mcp_tool_roundtrip"])
        self.assertIn("required Goal Plus MCP", result["error"])

    def test_codex_provider_contract_changes_with_external_api_config(self) -> None:
        profile = {"model": "gpt-5.6-sol", "reasoning_effort": "medium"}
        first = EDGE_ENV.codex_provider_contract(
            profile,
            {
                "OPENAI_API_KEY": "secret",
                "OPENAI_BASE_URL": "https://first.example/v1",
            },
        )
        second = EDGE_ENV.codex_provider_contract(
            profile,
            {
                "OPENAI_API_KEY": "secret",
                "OPENAI_BASE_URL": "https://second.example/new-base",
            },
        )

        self.assertTrue(first["valid"])
        self.assertTrue(second["valid"])
        self.assertNotEqual(
            first["provider_config_sha256"], second["provider_config_sha256"]
        )
        self.assertNotIn("secret", json.dumps(first))
        self.assertNotIn("/v1", json.dumps(first))

    def test_pi_host_probe_requires_tool_result_and_final_marker(self) -> None:
        models = self.temp / "incomplete-models.json"
        EDGE.write_json(
            models,
            {
                "providers": {
                    "dynamic": {
                        "baseUrl": "https://one.example/api",
                        "api": "openai-completions",
                        "apiKey": "$DYNAMIC_KEY",
                        "models": [{"id": "model"}],
                    }
                }
            },
        )

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            if command[-1] == "--version":
                return {"returncode": 0, "stdout": "0.84.1", "stderr": ""}
            return {
                "returncode": 0,
                "stdout": json.dumps(
                    {
                        "type": "message_end",
                        "message": {
                            "role": "assistant",
                            "provider": "dynamic",
                            "model": "model",
                            "api": "openai-completions",
                            "content": [{"type": "toolCall", "name": "read"}],
                        },
                    }
                ),
                "stderr": "",
            }

        with mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.pi_host_provider_probe(
                "dynamic/model",
                env={
                    "SFORGE_PI_MODELS_FILE": str(models),
                    "SFORGE_PI_HOST_EXECUTABLE": "/opt/pi/bin/pi",
                    "DYNAMIC_KEY": "secret",
                },
            )

        self.assertFalse(result["passed"])
        self.assertIn("tool call/result", result["error"])
        self.assertIn("expected final response", result["error"])

    def test_external_goal_plus_source_requires_and_matches_explicit_ref(self) -> None:
        subprocess.run(
            [
                "git",
                "-C",
                str(self.test_paths.goal_plus_root),
                "switch",
                "-c",
                "experiment/test-goal-plus-ref",
            ],
            check=True,
            capture_output=True,
        )
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(self.test_paths.goal_plus_root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "experiment/test-goal-plus-ref",
        }

        source = EDGE_ENV.resolve_goal_plus_source(env)

        self.assertTrue(source["valid"])
        self.assertEqual(source["source_kind"], "external")
        self.assertEqual(source["branch"], "experiment/test-goal-plus-ref")
        self.assertEqual(source["commit"], source["expected_ref_commit"])
        self.assertEqual(
            EDGE.upstream_entry("goal_plus")["tracking_branch"],
            "master",
        )

        missing_ref = EDGE_ENV.resolve_goal_plus_source(
            {"SFORGE_GOAL_PLUS_SOURCE_DIR": str(self.test_paths.goal_plus_root)}
        )
        self.assertFalse(missing_ref["valid"])
        self.assertIn("SFORGE_GOAL_PLUS_EXPECTED_REF", missing_ref["error"])

        (self.test_paths.goal_plus_root / "dirty.txt").write_text("dirty\n")
        dirty = EDGE_ENV.resolve_goal_plus_source(env)
        self.assertFalse(dirty["valid"])
        self.assertIn("must be clean", dirty["error"])

    def test_codex_goal_plus_source_requires_explicit_mcp_server_assets(self) -> None:
        root = self.test_paths.goal_plus_root
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }
        self.assertTrue(
            EDGE_ENV.resolve_goal_plus_source(
                env, methods=["goal-plus-codex"]
            )["valid"]
        )

        server = root / "src" / "goal_plus" / "server.py"
        server.unlink()
        self._commit_goal_plus_fixture("remove MCP server")
        missing = EDGE_ENV.resolve_goal_plus_source(
            env, methods=["goal-plus-codex"]
        )
        self.assertFalse(missing["valid"])
        self.assertIn("src/goal_plus/server.py", missing["missing_assets"])

    def test_codex_goal_plus_source_requires_host_authorization_assets(self) -> None:
        root = self.test_paths.goal_plus_root
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }

        metadata = root / ".codex" / "skills" / "goal-plus" / "agents" / "openai.yaml"
        metadata.unlink()
        self._commit_goal_plus_fixture("remove explicit-only skill metadata")
        missing_metadata = EDGE_ENV.resolve_goal_plus_source(
            env, methods=["goal-plus-codex"]
        )
        self.assertFalse(missing_metadata["valid"])
        self.assertIn(
            ".codex/skills/goal-plus/agents/openai.yaml",
            missing_metadata["missing_assets"],
        )

    def test_codex_goal_plus_source_requires_a_project_hook_layout(self) -> None:
        root = self.test_paths.goal_plus_root
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }

        (root / "hooks" / "hooks.json").unlink()
        self._commit_goal_plus_fixture("remove project hooks")
        missing_hooks = EDGE_ENV.resolve_goal_plus_source(
            env, methods=["goal-plus-codex"]
        )
        self.assertFalse(missing_hooks["valid"])
        self.assertEqual(
            missing_hooks["missing_asset_alternatives"],
            [list(EDGE_ENV.GOAL_PLUS_CODEX_HOOK_ASSETS)],
        )

    def test_active_codex_goal_plus_adapter_requires_exact_host_commands(self) -> None:
        contract = EDGE_ENV.active_sforge_codex_runtime_contract()

        self.assertTrue(contract["valid"], contract)
        self.assertEqual(contract["mode"], "host-command-hooks-explicit-mcp")
        self.assertTrue(contract["project_hooks_enabled"])
        self.assertTrue(contract["exact_start"])
        self.assertTrue(contract["typed_command_config"])
        self.assertTrue(contract["exact_resume"])
        self.assertFalse(contract["plugin_install"])

    def test_active_pi_goal_plus_adapter_requires_exact_host_commands(self) -> None:
        contract = EDGE_ENV.active_sforge_pi_runtime_contract()

        self.assertTrue(contract["valid"], contract)
        self.assertEqual(contract["mode"], "pi-extension-exact-host-command")
        self.assertTrue(contract["exact_start"])
        self.assertTrue(contract["typed_command_config"])
        self.assertTrue(contract["extension_loaded"])
        self.assertTrue(contract["reasoning_explicit"])
        self.assertTrue(contract["promotion_sync_persisted"])
        self.assertTrue(contract["exact_resume"])

    def test_goal_plus_source_checks_active_sforge_runtime_compatibility(self) -> None:
        root = self.test_paths.goal_plus_root
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }
        incompatible_adapter = {
            "valid": False,
            "mode": "plugin",
            "error": "plugin runtime enabled",
        }

        with mock.patch.object(
            EDGE_ENV,
            "active_sforge_codex_runtime_contract",
            return_value=incompatible_adapter,
        ):
            incompatible = EDGE_ENV.resolve_goal_plus_source(
                env, methods=["goal-plus-codex"]
            )
        self.assertFalse(incompatible["valid"])
        self.assertEqual(
            incompatible["codex_runtime_compatibility"]["mode"],
            "plugin",
        )

        with mock.patch.object(
            EDGE_ENV,
            "active_sforge_pi_runtime_contract",
            return_value=incompatible_adapter,
        ):
            incompatible = EDGE_ENV.resolve_goal_plus_source(
                env, methods=["goal-plus-pi"]
            )
        self.assertFalse(incompatible["valid"])
        self.assertEqual(
            incompatible["pi_runtime_compatibility"]["mode"],
            "plugin",
        )

    def test_goal_plus_runtime_assets_are_selected_by_method(self) -> None:
        root = self.test_paths.goal_plus_root
        server = root / "src" / "goal_plus" / "server.py"
        server.unlink()
        self._commit_goal_plus_fixture("remove Codex MCP server")
        env = {
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }

        codex = EDGE_ENV.resolve_goal_plus_source(
            env, methods=["goal-plus-codex"]
        )
        pi = EDGE_ENV.resolve_goal_plus_source(
            env, methods=["goal-plus-pi"]
        )

        self.assertFalse(codex["valid"])
        self.assertIn("src/goal_plus/server.py", codex["missing_assets"])
        self.assertTrue(pi["valid"])

    def test_prepare_freezes_external_api_and_goal_plus_inputs(self) -> None:
        models = self.temp / "prepare-models.json"
        EDGE.write_json(
            models,
            {
                "providers": {
                    "dynamic": {
                        "baseUrl": "http://127.0.0.1:43123/replaceable",
                        "api": "openai-responses",
                        "apiKey": "$DYNAMIC_KEY",
                        "models": [{"id": "model"}],
                    }
                }
            },
        )
        profile = {
            **self.profile(),
            "methods": ["goal-plus-pi-provider"],
            "model": "dynamic/model",
            "reasoning_effort": "medium",
            "pi_package_version": "0.84.1",
        }
        args = SimpleNamespace(
            method=None,
            wall_time_seconds=60,
            concurrency=2,
            cell_concurrency=1,
            model=None,
            reasoning_effort=None,
            campaign_id="freeze-external-inputs",
        )
        selected_env = {
            "SFORGE_PI_MODELS_FILE": str(models),
            "DYNAMIC_KEY": "secret",
            "SFORGE_GOAL_PLUS_SOURCE_DIR": str(self.test_paths.goal_plus_root),
            "SFORGE_GOAL_PLUS_EXPECTED_REF": "master",
        }
        with mock.patch.dict(EDGE_ENV.os.environ, selected_env, clear=False):
            destination = EDGE.prepare(args, profile)

        campaign = EDGE.read_json(destination / "campaign.json")
        snapshot = EDGE.read_json(destination / "profile.json")
        cell = EDGE.read_json(
            destination
            / "cells"
            / "vliw_kernel_optimization--goal-plus-pi-provider"
            / "cell.json"
        )
        self.assertEqual(campaign["goal_plus_source"]["source_kind"], "external")
        self.assertIsNone(campaign["goal_plus_tracking_branch"])
        self.assertEqual(
            campaign["goal_plus_commit"], EDGE.git_head(self.test_paths.goal_plus_root)
        )
        self.assertEqual(cell["goal_plus_source"], campaign["goal_plus_source"])
        self.assertEqual(
            snapshot["pi_provider_contract"]["model_refs"], ["dynamic/model"]
        )
        self.assertNotIn("secret", json.dumps(campaign))

    def test_goal_plus_pi_roles_resolve_multiple_provider_endpoints(self) -> None:
        models = self.temp / "multi-role-models.json"
        models.write_text(
            json.dumps(
                {
                    "providers": {
                        "local-worker": {
                            "baseUrl": "http://127.0.0.1:4101/v1",
                            "apiKey": "$LOCAL_WORKER_KEY",
                            "models": [{"id": "worker"}],
                        },
                        "local-view": {
                            "baseUrl": "http://127.0.0.1:4102/v1",
                            "apiKey": "$LOCAL_VIEW_KEY",
                            "models": [{"id": "annotator"}],
                        },
                    }
                }
            )
        )
        profile = {
            **self.profile(),
            "methods": ["goal-plus-pi-provider"],
            "model": "zai/glm-5.2",
            "worker_model": "local-worker/worker",
            "evidence_annotator_model": "local-view/annotator",
        }
        endpoints = EDGE_ENV.resolve_profile_api_endpoints(
            profile, {"SFORGE_PI_MODELS_FILE": str(models)}
        )

        self.assertEqual(
            [item["role"] for item in endpoints],
            ["main", "worker", "evidence_annotator"],
        )
        self.assertEqual(
            {(item["host"], item["port"]) for item in endpoints},
            {
                ("api.z.ai", 443),
                ("127.0.0.1", 4101),
                ("127.0.0.1", 4102),
            },
        )

    def test_pi_provider_prefers_anthropic_oauth_environment(self) -> None:
        status = EDGE.resolve_pi_provider(
            "anthropic/claude-sonnet-4-20250514",
            {
                "ANTHROPIC_OAUTH_TOKEN": "oauth-secret",
                "ANTHROPIC_API_KEY": "api-secret",
            },
        )

        self.assertTrue(status["valid"])
        self.assertEqual(status["credential_env"], "ANTHROPIC_OAUTH_TOKEN")
        self.assertNotIn("oauth-secret", json.dumps(status))
        self.assertNotIn("api-secret", json.dumps(status))

    def test_claude_environment_pins_effort_and_preserves_extra_env(self) -> None:
        previous = EDGE.os.environ.get("SFORGE_AGENT_EXTRA_ENV")
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = "EXISTING=value"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "reasoning_effort": "high",
                    "internet": False,
                }
            )
        finally:
            if previous is None:
                EDGE.os.environ.pop("SFORGE_AGENT_EXTRA_ENV", None)
            else:
                EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = previous

        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        self.assertEqual(extra["CLAUDE_CODE_EFFORT_LEVEL"], "high")
        self.assertEqual(extra["CLAUDE_CODE_ALWAYS_ENABLE_EFFORT"], "1")
        self.assertEqual(env["SFORGE_CLAUDE_CACHE_OPT"], "1")
        self.assertNotIn("SFORGE_CODEX_REASONING_EFFORT", env)

    def test_claude_none_environment_disables_thinking_and_removes_effort(self) -> None:
        keys = (
            "SFORGE_AGENT_EXTRA_ENV",
            "CLAUDE_CODE_EFFORT_LEVEL",
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT",
        )
        previous = {key: EDGE.os.environ.get(key) for key in keys}
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = (
            "EXISTING=value,CLAUDE_CODE_EFFORT_LEVEL=high,"
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT=1"
        )
        EDGE.os.environ["CLAUDE_CODE_EFFORT_LEVEL"] = "high"
        EDGE.os.environ["CLAUDE_CODE_ALWAYS_ENABLE_EFFORT"] = "1"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "reasoning_effort": "none",
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        self.assertEqual(extra["MAX_THINKING_TOKENS"], "0")
        self.assertEqual(extra["CLAUDE_CODE_DISABLE_THINKING"], "1")
        self.assertNotIn("CLAUDE_CODE_EFFORT_LEVEL", extra)
        self.assertNotIn("CLAUDE_CODE_ALWAYS_ENABLE_EFFORT", extra)
        self.assertNotIn("CLAUDE_CODE_EFFORT_LEVEL", env)
        self.assertNotIn("CLAUDE_CODE_ALWAYS_ENABLE_EFFORT", env)

    def test_claude_adaptive_environment_removes_fixed_thinking_controls(self) -> None:
        keys = (
            "SFORGE_AGENT_EXTRA_ENV",
            "MAX_THINKING_TOKENS",
            "CLAUDE_CODE_DISABLE_THINKING",
            "CLAUDE_CODE_DISABLE_ADAPTIVE_THINKING",
            "CLAUDE_CODE_EFFORT_LEVEL",
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT",
        )
        previous = {key: EDGE.os.environ.get(key) for key in keys}
        EDGE.os.environ["SFORGE_AGENT_EXTRA_ENV"] = (
            "EXISTING=value,MAX_THINKING_TOKENS=0,"
            "CLAUDE_CODE_DISABLE_THINKING=1,CLAUDE_CODE_EFFORT_LEVEL=high,"
            "CLAUDE_CODE_ALWAYS_ENABLE_EFFORT=1"
        )
        for key in keys[1:]:
            EDGE.os.environ[key] = "1"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-claude",
                    "sforge_agent": "claude-code",
                    "model": "glm-5.1",
                    "thinking": {"type": "adaptive"},
                    "reasoning_effort": None,
                    "claude_context_window_tokens": 200000,
                    "claude_autocompact_percent": 80,
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        extra = dict(
            item.split("=", 1) for item in env["SFORGE_AGENT_EXTRA_ENV"].split(",")
        )
        self.assertEqual(extra["EXISTING"], "value")
        for key in (
            "ANTHROPIC_MODEL",
            "ANTHROPIC_DEFAULT_OPUS_MODEL",
            "ANTHROPIC_DEFAULT_SONNET_MODEL",
            "ANTHROPIC_DEFAULT_HAIKU_MODEL",
            "CLAUDE_CODE_SUBAGENT_MODEL",
        ):
            self.assertEqual(extra[key], "glm-5.1")
        self.assertEqual(extra["CLAUDE_CODE_AUTO_COMPACT_WINDOW"], "200000")
        self.assertEqual(extra["CLAUDE_AUTOCOMPACT_PCT_OVERRIDE"], "80")
        for key in keys[1:]:
            self.assertNotIn(key, extra)
            self.assertNotIn(key, env)

    def test_loopback_api_bridge_preserves_base_path(self) -> None:
        self.assertEqual(
            EDGE.loopback_api_target("http://127.0.0.1:3788/v1"),
            ("127.0.0.1", 3788),
        )
        self.assertEqual(
            EDGE.bridged_base_url("http://127.0.0.1:3788/v1", "192.0.2.10", 45678),
            "http://192.0.2.10:45678/v1",
        )
        self.assertIsNone(EDGE.loopback_api_target("https://api.example.com/v1"))

    def test_pi_role_models_get_independent_loopback_bridges(self) -> None:
        def provider(port: int, key: str, *model_ids: str) -> dict:
            return {
                "baseUrl": f"http://127.0.0.1:{port}/v1",
                "apiKey": f"${key}",
                "models": [{"id": model_id} for model_id in model_ids],
            }

        models = self.temp / "models.json"
        EDGE.write_json(
            models,
            {
                "providers": {
                    "main-provider": provider(
                        18080,
                        "MAIN_API_KEY",
                        "main-model",
                        "annotator-model",
                        "unused-model",
                    ),
                    "worker-provider": provider(
                        18081, "WORKER_API_KEY", "worker-model"
                    ),
                }
            },
        )
        resources = EDGE.RuntimeResources()
        controller = {"bridges": []}

        def bridge(pid: int, port: int) -> tuple:
            return SimpleNamespace(pid=pid), {"listen_port": port}, mock.Mock()

        bridges = mock.Mock(side_effect=[bridge(101, 28080), bridge(102, 28081)])
        probes = mock.Mock(
            return_value={"passed": True, "status": "200", "stderr": None}
        )
        profile = {
            "model": "main-provider/main-model",
            "worker_model": "worker-provider/worker-model",
            "evidence_annotator_model": "main-provider/annotator-model",
            "task_ids": ["vliw_kernel_optimization"],
        }
        with mock.patch.dict(
            EDGE_RUNTIME.os.environ,
            {
                "SFORGE_PI_MODELS_FILE": str(models),
                "MAIN_API_KEY": "main-secret",
                "WORKER_API_KEY": "worker-secret",
            },
            clear=False,
        ), mock.patch.object(
            EDGE_RUNTIME, "default_route_ipv4", return_value="192.0.2.10"
        ), mock.patch.object(
            EDGE_RUNTIME, "start_socket_bridge", bridges
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_endpoint_reachability_probe", probes
        ), mock.patch.object(
            EDGE_RUNTIME, "task_images", return_value=("example:work", "example:judge")
        ):
            EDGE_RUNTIME.prepare_pi_provider_runtime(
                resources, self.temp, profile, controller
            )

        runtime_registry = json.loads(resources.pi_models_file.read_text())
        self.assertEqual(bridges.call_count, 2)
        self.assertEqual(probes.call_count, 2)
        self.assertCountEqual(
            [call.args[1] for call in probes.call_args_list],
            [
                "http://192.0.2.10:28080/v1",
                "http://192.0.2.10:28081/v1",
            ],
        )
        self.assertEqual(
            {
                provider: config["baseUrl"]
                for provider, config in runtime_registry["providers"].items()
            },
            {
                "main-provider": "http://192.0.2.10:28080/v1",
                "worker-provider": "http://192.0.2.10:28081/v1",
            },
        )
        self.assertEqual(
            [
                model["id"]
                for model in runtime_registry["providers"]["main-provider"]["models"]
            ],
            ["main-model", "annotator-model"],
        )
        self.assertEqual(
            resources.pi_provider_credentials,
            {"MAIN_API_KEY": "main-secret", "WORKER_API_KEY": "worker-secret"},
        )
        self.assertEqual(
            resources.runtime_api_base_url,
            "http://192.0.2.10:28080/v1",
        )
        self.assertEqual(
            resources.runtime_api_base_urls,
            [
                "http://192.0.2.10:28080/v1",
                "http://192.0.2.10:28081/v1",
            ],
        )
        self.assertEqual(
            controller["pi_provider_roles"],
            {
                "main": "main-provider/main-model",
                "worker": "worker-provider/worker-model",
                "evidence_annotator": "main-provider/annotator-model",
                "credential_envs": ["MAIN_API_KEY", "WORKER_API_KEY"],
            },
        )
        self.assertNotIn("secret", json.dumps(runtime_registry))

    def test_pi_remote_provider_does_not_require_loopback_route(self) -> None:
        models = self.temp / "models.json"
        EDGE.write_json(
            models,
            {
                "providers": {
                    "remote-provider": {
                        "baseUrl": "http://192.0.2.20:18081/v1",
                        "api": "openai-responses",
                        "apiKey": "$REMOTE_API_KEY",
                        "models": [{"id": "remote-model"}],
                    }
                }
            },
        )
        resources = EDGE.RuntimeResources()
        controller = {"bridges": []}
        profile = {
            "model": "remote-provider/remote-model",
            "task_ids": ["vliw_kernel_optimization"],
        }
        probes = mock.Mock(
            return_value={"passed": True, "status": "401", "stderr": None}
        )

        with mock.patch.dict(
            EDGE_RUNTIME.os.environ,
            {
                "SFORGE_PI_MODELS_FILE": str(models),
                "REMOTE_API_KEY": "remote-secret",
            },
            clear=False,
        ), mock.patch.object(
            EDGE_RUNTIME, "default_route_ipv4"
        ) as default_route, mock.patch.object(
            EDGE_RUNTIME, "start_socket_bridge"
        ) as start_bridge, mock.patch.object(
            EDGE_RUNTIME, "docker_endpoint_reachability_probe", probes
        ), mock.patch.object(
            EDGE_RUNTIME, "task_images", return_value=("example:work", "example:judge")
        ):
            EDGE_RUNTIME.prepare_pi_provider_runtime(
                resources, self.temp, profile, controller
            )

        default_route.assert_not_called()
        start_bridge.assert_not_called()
        probes.assert_called_once_with(
            "example:work", "http://192.0.2.20:18081/v1"
        )
        self.assertIsNone(resources.bridge_host)
        self.assertEqual(
            resources.runtime_api_base_url, "http://192.0.2.20:18081/v1"
        )
        self.assertEqual(controller["bridges"], [])

    def test_cell_environment_maps_api_key_and_bridge(self) -> None:
        env = EDGE.cell_environment(
            {
                "method": "plain-codex",
                "reasoning_effort": "medium",
                "internet": False,
            },
            api_key="runtime-key",
            api_base_url="http://192.0.2.10:45678/v1",
            bridge_host="192.0.2.10",
        )

        self.assertEqual(env["SFORGE_AGENT_API_KEY"], "runtime-key")
        self.assertEqual(
            env["SFORGE_AGENT_API_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertIn("192.0.2.10", env["SFORGE_NO_PROXY"].split(","))
        self.assertEqual(
            json.loads(env["SFORGE_AGENT_API_BASE_URLS"]),
            ["http://192.0.2.10:45678/v1"],
        )

    def test_command_rejects_open_network_cell(self) -> None:
        with self.assertRaisesRegex(ValueError, "API-only"):
            EDGE.build_sforge_command(
                self.temp,
                {
                    "cell_id": "open-network",
                    "task_id": "vliw_kernel_optimization",
                    "sforge_agent": "codex",
                    "model": "gpt-5.6-sol",
                    "wall_time_seconds": 60,
                    "eval_interval_seconds": 60,
                    "sforge_run_id": "open-network",
                    "outer_replicas": 1,
                    "outer_replica_concurrency": 1,
                    "judge_concurrency": 1,
                    "internet": True,
                },
            )

    def test_judge_environment_uses_fixed_model_and_runtime_api(self) -> None:
        previous = EDGE.os.environ.pop("SFORGE_JUDGE_EXTRA_ENV", None)
        try:
            env = EDGE.judge_server_environment(
                api_key="judge-key",
                api_base_url="http://192.0.2.10:45678/v1",
                bridge_host="192.0.2.10",
            )
        finally:
            if previous is not None:
                EDGE.os.environ["SFORGE_JUDGE_EXTRA_ENV"] = previous

        values = dict(
            item.split("=", 1) for item in env["SFORGE_JUDGE_EXTRA_ENV"].split(",")
        )
        self.assertEqual(values["SFORGE_JUDGE_API_KEY"], "judge-key")
        self.assertEqual(
            values["SFORGE_JUDGE_API_BASE_URL"],
            "http://192.0.2.10:45678/v1",
        )
        self.assertEqual(values["SFORGE_JUDGE_MODEL"], "gpt-5.5")

    def test_open_network_cell_environment_is_rejected(self) -> None:
        previous = {
            key: EDGE.os.environ.get(key)
            for key in ("SFORGE_HTTPS_PROXY", "HTTPS_PROXY")
        }
        EDGE.os.environ.pop("SFORGE_HTTPS_PROXY", None)
        EDGE.os.environ["HTTPS_PROXY"] = "http://127.0.0.1:3128"
        try:
            with self.assertRaisesRegex(ValueError, "API-only"):
                EDGE.cell_environment(
                    {
                        "method": "plain-codex",
                        "reasoning_effort": "high",
                        "internet": True,
                    }
                )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

    def test_isolated_cell_environment_removes_all_proxy_variables(self) -> None:
        proxy_keys = (
            "ALL_PROXY",
            "HTTP_PROXY",
            "HTTPS_PROXY",
            "SFORGE_HTTP_PROXY",
            "SFORGE_HTTPS_PROXY",
            "all_proxy",
            "http_proxy",
            "https_proxy",
        )
        previous = {key: EDGE.os.environ.get(key) for key in proxy_keys}
        for key in proxy_keys:
            EDGE.os.environ[key] = "http://proxy.example:3128"
        try:
            env = EDGE.cell_environment(
                {
                    "method": "plain-codex",
                    "reasoning_effort": "high",
                    "internet": False,
                }
            )
        finally:
            for key, value in previous.items():
                if value is None:
                    EDGE.os.environ.pop(key, None)
                else:
                    EDGE.os.environ[key] = value

        for key in proxy_keys:
            self.assertNotIn(key, env)

    def test_docker_resource_probe_verifies_applied_host_config(self) -> None:
        original = EDGE_IO.run_capture
        commands = []

        def fake_run_capture(command, *, env=None):
            commands.append(command)
            if command[1] == "run":
                return {"returncode": 0, "stdout": "container-id\n", "stderr": ""}
            if command[1] == "inspect":
                return {
                    "returncode": 0,
                    "stdout": json.dumps(
                        {"NanoCpus": 4_000_000_000, "Memory": 16 * 1024**3}
                    ),
                    "stderr": "",
                }
            return {"returncode": 0, "stdout": "", "stderr": ""}

        EDGE_IO.run_capture = fake_run_capture
        try:
            result = EDGE.docker_resource_limit_probe(
                "example:work", cpu_limit=4, mem_limit="16g"
            )
        finally:
            EDGE_IO.run_capture = original

        self.assertTrue(result["passed"])
        self.assertEqual(commands[0][2:4], ["--pull", "never"])
        self.assertIn("--cpus", commands[0])
        self.assertIn("--memory", commands[0])
        self.assertEqual(commands[-1][1:3], ["rm", "--force"])

    def test_rust_runtime_probe_preserves_image_environment(self) -> None:
        original = EDGE_IO.run_capture
        captured = []

        def fake_run_capture(command, *, env=None):
            captured.append(command)
            return {"returncode": 0, "stdout": "rustc 1.88.0", "stderr": ""}

        EDGE_IO.run_capture = fake_run_capture
        try:
            result = EDGE.rust_image_runtime_probe("example:rust", "1.88.0")
        finally:
            EDGE_IO.run_capture = original

        self.assertEqual(result["returncode"], 0)
        self.assertEqual(captured[0][2:4], ["--pull", "never"])
        self.assertIn("-c", captured[0])
        self.assertNotIn("-lc", captured[0])
        self.assertIn("command -v cargo", captured[0][-1])

    def test_docker_http_probe_never_pulls_diagnostic_image(self) -> None:
        original = EDGE_IO.run_capture
        captured = []

        def fake_run_capture(command, *, env=None):
            captured.append(command)
            return {"returncode": 0, "stdout": "200", "stderr": ""}

        EDGE_IO.run_capture = fake_run_capture
        try:
            result = EDGE.docker_http_probe(
                "example:work", "http://host.docker.internal:3788/v1/models"
            )
        finally:
            EDGE_IO.run_capture = original

        self.assertTrue(result["passed"])
        self.assertEqual(captured[0][2:4], ["--pull", "never"])

    def test_pi_container_probe_only_checks_dynamic_endpoint_reachability(self) -> None:
        captured: dict[str, object] = {}

        def fake_run_capture(command, *, env=None, timeout_seconds=None):
            captured["command"] = command
            captured["env"] = dict(env or {})
            return {"returncode": 0, "stdout": "404", "stderr": ""}

        with mock.patch.object(EDGE_ENV.io, "run_capture", fake_run_capture):
            result = EDGE_ENV.docker_endpoint_reachability_probe(
                "example:work", "https://changing.example/custom/base"
            )

        self.assertTrue(result["passed"])
        self.assertEqual(result["status"], "404")
        self.assertEqual(captured["command"][2:4], ["--pull", "never"])
        self.assertEqual(
            captured["env"]["SFORGE_PROBE_URL"],
            "https://changing.example/custom/base",
        )
        serialized = json.dumps(captured)
        self.assertNotIn("SFORGE_PROBE_API_KEY", serialized)
        self.assertNotIn("/responses", serialized)
        self.assertNotIn("/chat/completions", serialized)

    def test_host_pi_failure_stops_launch_before_all_docker_calls(self) -> None:
        profile = {
            **self.profile(),
            "methods": ["goal-plus-pi-provider"],
            "api_protocol": "pi-provider",
            "model": "dynamic/model",
            "reasoning_effort": "medium",
            "goal_plus_source": {
                "source_dir": str(self.test_paths.goal_plus_root),
                "expected_ref": "master",
                "commit": EDGE.git_head(self.test_paths.goal_plus_root),
            },
        }
        bundle = {
            "valid": True,
            "error": None,
            "model_refs": ["dynamic/model"],
            "models": [],
            "credential_envs": ["DYNAMIC_KEY"],
            "registry": {"providers": {}},
        }
        docker_probe = mock.Mock()
        with mock.patch.object(
            EDGE_RUNTIME,
            "resolve_goal_plus_source",
            return_value={
                "valid": True,
                "source_kind": "external",
                "source_path": "goal-plus",
                "checkout_root": "checkout",
                "expected_ref": "master",
                "branch": "master",
                "commit": profile["goal_plus_source"]["commit"],
                "dirty": False,
                "missing_assets": [],
                "missing_asset_alternatives": [],
                "codex_runtime_compatibility": None,
                "pi_runtime_compatibility": {"valid": True},
                "error": None,
            },
        ), mock.patch.object(
            EDGE_RUNTIME, "resolve_pi_provider_bundle", return_value=bundle
        ), mock.patch.object(
            EDGE_RUNTIME,
            "pi_provider_host_preflight",
            return_value={
                "passed": False,
                "contract": None,
                "probes": [{"passed": False, "error": "tool roundtrip failed"}],
                "error": None,
            },
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_endpoint_reachability_probe", docker_probe
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_http_probe", docker_probe
        ):
            with self.assertRaisesRegex(RuntimeError, "before Docker"):
                EDGE_RUNTIME.prepare_runtime_resources(
                    self.temp / "campaign", profile, {}
                )

        docker_probe.assert_not_called()

    def test_host_codex_failure_stops_launch_before_all_docker_calls(self) -> None:
        profile = {
            **self.profile(),
            "methods": ["goal-plus-codex"],
            "api_protocol": "openai",
            "model": "gpt-5.6-sol",
            "reasoning_effort": "medium",
            "goal_plus_source": {
                "source_dir": str(self.test_paths.goal_plus_root),
                "expected_ref": "master",
                "commit": EDGE.git_head(self.test_paths.goal_plus_root),
            },
        }
        contract = {
            "valid": True,
            "provider": "sforge-proxy",
            "model": "gpt-5.6-sol",
        }
        docker_probe = mock.Mock()
        with mock.patch.object(
            EDGE_RUNTIME,
            "resolve_goal_plus_source",
            return_value={
                "valid": True,
                "source_kind": "external",
                "source_path": "goal-plus",
                "checkout_root": "checkout",
                "expected_ref": "master",
                "branch": "master",
                "commit": profile["goal_plus_source"]["commit"],
                "dirty": False,
                "missing_assets": [],
                "missing_asset_alternatives": [],
                "codex_runtime_compatibility": None,
                "pi_runtime_compatibility": None,
                "error": None,
            },
        ), mock.patch.object(
            EDGE_RUNTIME, "codex_provider_contract", return_value=contract
        ), mock.patch.object(
            EDGE_RUNTIME,
            "codex_host_provider_probe",
            return_value={
                "passed": False,
                "contract": contract,
                "error": "tool roundtrip failed",
            },
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_endpoint_reachability_probe", docker_probe
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_http_probe", docker_probe
        ):
            with self.assertRaisesRegex(RuntimeError, "before Docker"):
                EDGE_RUNTIME.prepare_runtime_resources(
                    self.temp / "campaign", profile, {}
                )

        docker_probe.assert_not_called()

    def test_pi_oauth_runtime_does_not_run_codex_host_preflight(self) -> None:
        profile = {
            **self.profile(),
            "methods": ["goal-plus-pi"],
            "api_protocol": "openai",
            "model": "gpt-5.6-sol",
            "reasoning_effort": "medium",
            "goal_plus_source": {
                "source_dir": str(self.test_paths.goal_plus_root),
                "expected_ref": "master",
                "commit": EDGE.git_head(self.test_paths.goal_plus_root),
            },
        }
        resolved_source = {
            "valid": True,
            "source_kind": "external",
            "source_path": "goal-plus",
            "checkout_root": "checkout",
            "expected_ref": "master",
            "branch": "master",
            "commit": profile["goal_plus_source"]["commit"],
            "dirty": False,
            "missing_assets": [],
            "missing_asset_alternatives": [],
            "codex_runtime_compatibility": None,
            "pi_runtime_compatibility": {"valid": True},
            "error": None,
        }
        codex_contract = mock.Mock(side_effect=AssertionError("Codex contract used"))
        codex_probe = mock.Mock(side_effect=AssertionError("Codex probe used"))

        with mock.patch.object(
            EDGE_RUNTIME, "resolve_goal_plus_source", return_value=resolved_source
        ), mock.patch.object(
            EDGE_RUNTIME, "codex_provider_contract", codex_contract
        ), mock.patch.object(
            EDGE_RUNTIME, "codex_host_provider_probe", codex_probe
        ), mock.patch.object(
            EDGE_RUNTIME,
            "start_or_reuse_judge",
            side_effect=RuntimeError("reached Judge startup"),
        ):
            with self.assertRaisesRegex(RuntimeError, "reached Judge startup"):
                EDGE_RUNTIME.prepare_runtime_resources(
                    self.temp / "campaign", profile, {}
                )

        codex_contract.assert_not_called()
        codex_probe.assert_not_called()

    def test_changed_external_provider_config_stops_before_host_pi_or_docker(
        self,
    ) -> None:
        profile = {
            **self.profile(),
            "methods": ["plain-pi-provider"],
            "api_protocol": "pi-provider",
            "model": "dynamic/model",
            "reasoning_effort": "medium",
            "pi_provider_contract": {"frozen": True},
        }
        bundle = {
            "valid": True,
            "error": None,
            "model_refs": ["dynamic/model"],
            "models": [],
            "credential_envs": [],
            "registry": {"providers": {}},
        }
        host_preflight = mock.Mock()
        docker_probe = mock.Mock()
        with mock.patch.object(
            EDGE_RUNTIME, "resolve_pi_provider_bundle", return_value=bundle
        ), mock.patch.object(
            EDGE_RUNTIME,
            "pi_provider_bundle_contract",
            return_value={"frozen": False},
        ), mock.patch.object(
            EDGE_RUNTIME, "pi_provider_host_preflight", host_preflight
        ), mock.patch.object(
            EDGE_RUNTIME, "docker_endpoint_reachability_probe", docker_probe
        ):
            with self.assertRaisesRegex(RuntimeError, "changed after campaign prepare"):
                EDGE_RUNTIME.prepare_runtime_resources(
                    self.temp / "campaign", profile, {}
                )

        host_preflight.assert_not_called()
        docker_probe.assert_not_called()

    def test_codex_usage_reads_jsonl_agent_output(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        (run / "agent_output.txt").write_text(
            "\n".join(
                [
                    '{"type":"thread.started","thread_id":"thread-1"}',
                    '{"type":"turn.completed","usage":{"input_tokens":11,"cached_input_tokens":3,"output_tokens":5}}',
                    "non-json status line",
                    '{"type":"turn.completed","usage":{"input_tokens":7,"output_tokens":2}}',
                ]
            ),
            encoding="utf-8",
        )

        usage = EDGE.codex_usage(run)

        self.assertEqual(usage["coverage"], "agent_output_only")
        self.assertEqual(usage["session_count"], 1)
        self.assertEqual(usage["tokens"]["input_tokens"], 18)
        self.assertEqual(usage["tokens"]["output_tokens"], 7)
        self.assertEqual(usage["tokens"]["cached_input_tokens"], 3)

    def test_codex_usage_reads_cumulative_rollout_tokens_once(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        events = "\n".join(
            [
                '{"type":"session_meta","payload":{"id":"session-1"}}',
                '{"type":"event_msg","payload":{"type":"token_count","info":{"total_token_usage":{"input_tokens":11,"cached_input_tokens":3,"output_tokens":5,"total_tokens":16}}}}',
                '{"type":"event_msg","payload":{"type":"token_count","info":{"total_token_usage":{"input_tokens":18,"cached_input_tokens":7,"output_tokens":9,"total_tokens":27}}}}',
            ]
        ).encode()
        member = tarfile.TarInfo("sessions/rollout.jsonl")
        member.size = len(events)
        with tarfile.open(run / "codex-sessions.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(events))

        usage = EDGE.codex_usage(run)

        self.assertEqual(usage["coverage"], "all_collected_codex_sessions")
        self.assertEqual(usage["session_count"], 1)
        self.assertEqual(usage["tokens"]["input_tokens"], 18)
        self.assertEqual(usage["tokens"]["cached_input_tokens"], 7)
        self.assertEqual(usage["tokens"]["output_tokens"], 9)
        self.assertEqual(usage["tokens"]["total_tokens"], 27)

    def test_agent_usage_reads_pi_message_usage_once(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        usage = {"input": 11, "output": 5, "cacheRead": 3, "cacheWrite": 2}
        (run / "agent_output.txt").write_text(
            "\n".join(
                [
                    '{"type":"session","id":"pi-session"}',
                    json.dumps({"type": "message_end", "message": {"usage": usage}}),
                    json.dumps({"type": "turn_end", "usage": usage}),
                ]
            ),
            encoding="utf-8",
        )

        observed = EDGE.codex_usage(run)

        self.assertEqual(observed["coverage"], "pi_agent_output")
        self.assertEqual(observed["session_count"], 1)
        self.assertEqual(observed["tokens"]["input_tokens"], 11)
        self.assertEqual(observed["tokens"]["cached_input_tokens"], 3)
        self.assertEqual(observed["tokens"]["output_tokens"], 5)
        self.assertEqual(observed["tokens"]["processed_tokens"], 21)

    def test_goal_plus_stats_counts_empty_search_run(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        payload = b'{"run_id":"run-1","state":"running"}'
        member = tarfile.TarInfo(".goal-plus/runs/run-1/run.json")
        member.size = len(payload)
        annotation = json.dumps(
            {
                "state": "completed",
                "attempts": 1,
                "usage": {
                    "input_tokens": 13,
                    "output_tokens": 5,
                    "cost_usd": 0.001,
                },
            }
        ).encode()
        annotation_member = tarfile.TarInfo(
            ".goal-plus/runs/run-1/candidates/c001/"
            "evidence-annotations/iteration-0001.json"
        )
        annotation_member.size = len(annotation)
        with tarfile.open(run / "goal-plus-state.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(payload))
            archive.addfile(annotation_member, io.BytesIO(annotation))

        stats = EDGE.goal_plus_stats(run)

        self.assertIsNotNone(stats)
        assert stats is not None
        self.assertEqual(stats["search_runs"], 1)
        self.assertEqual(stats["candidates"], 0)
        self.assertEqual(stats["search_run_states"], {"running": 1})
        self.assertEqual(stats["selected_candidate_ids"], [])
        self.assertEqual(stats["promoted_candidate_ids"], [])
        self.assertEqual(
            stats["evidence_annotator_usage"],
            {
                "input_tokens": 13,
                "output_tokens": 5,
                "cost_usd": 0.001,
                "tasks": 1,
                "attempts": 1,
                "states": {"completed": 1},
                "coverage": "persisted Goal Plus Evidence annotator turns",
            },
        )

    def test_goal_plus_stats_and_report_count_pi_worker_usage(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        events = (
            '{"type":"message_end","usage":'
            '{"input":7,"output":3,"cacheRead":5,"cacheWrite":0}}\n'
            '{"type":"turn_end","usage":'
            '{"input":7,"output":3,"cacheRead":5,"cacheWrite":0}}'
        ).encode()
        member = tarfile.TarInfo(".goal-plus/host-logs/pi-rpc-agent-session.jsonl")
        member.size = len(events)
        with tarfile.open(run / "goal-plus-state.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(events))

        stats = EDGE.goal_plus_stats(run)

        assert stats is not None
        self.assertEqual(stats["worker_usage"]["input_tokens"], 7)
        self.assertEqual(stats["worker_usage"]["output_tokens"], 3)
        self.assertEqual(stats["worker_usage"]["sessions"], 1)
        record = EDGE.comparison_record(
            {
                "task_id": "vliw_kernel_optimization",
                "method": "goal-plus-pi",
                "wall_time_seconds": 7200,
                "live_search_concurrency": 4,
                "completed_trajectories": 1,
                "valid_trajectories": 1,
                "observations": [{"goal_plus": stats}],
            },
            {"vliw_kernel_optimization": {"mean": 0}},
        )
        self.assertEqual(record["Input tokens"], 7)
        self.assertEqual(record["Output tokens"], 3)
        self.assertEqual(record["Usage coverage"], "persisted Pi worker message usage")

    def test_goal_plus_stats_recovers_archived_promotion(self) -> None:
        run = self.temp / "task-run"
        run.mkdir()
        payload = json.dumps(
            {
                "run_id": "run-1",
                "state": "promoted",
                "selected_candidate_id": "c001",
            }
        ).encode()
        member = tarfile.TarInfo(".goal-plus/runs/run-1/run.json")
        member.size = len(payload)
        with tarfile.open(run / "goal-plus-state.tar", "w") as archive:
            archive.addfile(member, io.BytesIO(payload))

        stats = EDGE.goal_plus_stats(run)

        self.assertIsNotNone(stats)
        assert stats is not None
        self.assertEqual(stats["search_run_states"], {"promoted": 1})
        self.assertEqual(stats["selected_candidate_ids"], ["c001"])
        self.assertEqual(stats["promoted_candidate_ids"], ["c001"])

    def test_provision_excludes_downloaded_tasks_from_git_status(self) -> None:
        exclude = EDGE.current_paths().edge_root / ".git" / "info" / "exclude"
        exclude.parent.mkdir(parents=True)
        exclude.write_text("# local excludes\n", encoding="utf-8")

        EDGE.ensure_local_task_exclude()
        EDGE.ensure_local_task_exclude()

        self.assertEqual(
            exclude.read_text(encoding="utf-8").splitlines().count("tasks/"),
            1,
        )

    def test_command_keeps_plain_replicas_distinct_from_goal_plus_workers(self) -> None:
        destination = self.temp / "campaign"
        plain = {
            "cell_id": "plain",
            "task_id": "vliw_kernel_optimization",
            "sforge_agent": "codex",
            "model": "gpt-test",
            "wall_time_seconds": 300,
            "eval_interval_seconds": 120,
            "sforge_run_id": "run-plain",
            "outer_replicas": 4,
            "outer_replica_concurrency": 4,
            "judge_concurrency": 1,
            "judge_port": 8080,
            "work_cpu_limit": 4,
            "work_mem_limit": "16g",
            "judge_cpu_limit": 4,
            "judge_mem_limit": "8g",
            "submission_cooldown": 0,
            "max_submissions": 7,
            "auto_eval_enabled": False,
            "auto_resume_enabled": False,
            "stop_hook_enabled": False,
            "internet": False,
        }

        command = EDGE.build_sforge_command(destination, plain)

        self.assertLess(command.index("--silent"), command.index("run"))
        self.assertEqual(command[command.index("--replicas") + 1], "4")
        self.assertEqual(command[command.index("--replica-concurrency") + 1], "4")
        self.assertEqual(
            command[command.index("--judge-url") + 1],
            "http://host.docker.internal:8080",
        )
        self.assertEqual(command[command.index("--work-cpu-limit") + 1], "4")
        self.assertEqual(command[command.index("--work-mem-limit") + 1], "16g")
        self.assertEqual(command[command.index("--judge-cpu-limit") + 1], "4")
        self.assertEqual(command[command.index("--judge-mem-limit") + 1], "8g")
        self.assertEqual(command[command.index("--submission-cooldown") + 1], "0")
        self.assertEqual(command[command.index("--max-submissions") + 1], "7")
        self.assertIn("--disable-auto-eval", command)
        self.assertIn("--disable-auto-resume", command)
        self.assertIn("--disable-stop-hook", command)
        self.assertIn("--disable-internet", command)

    def test_goal_plus_codex_completion_requires_real_spawn_and_verifier_evidence(
        self,
    ) -> None:
        cell = {
            "method": "goal-plus-codex",
            "outer_replicas": 1,
            "inner_search_concurrency": 2,
        }
        complete = {
            "edgebench_score": 50.0,
            "goal_plus": {
                "candidates": 2,
                "agent_sessions": 2,
                "worker_verifier_runs": 2,
            },
            "agent_events": {
                "spawn_agent_completed_count": 2,
                "spawned_agent_thread_count": 2,
                "goal_plus": {
                    "candidate_ids": ["c001", "c002"],
                    "agent_session_ids": ["a001", "a002"],
                    "verifier_ledger": [
                        {"candidate_id": "c001"},
                        {"candidate_id": "c002"},
                    ],
                    "selected_candidate_ids": ["c001"],
                    "promoted_candidate_ids": ["c001"],
                },
            },
        }

        passed = EDGE.goal_plus_completion_evidence(
            cell, [complete], valid_trajectories=1
        )
        missing_spawn = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "agent_events": {
                        **complete["agent_events"],
                        "spawn_agent_completed_count": 0,
                        "spawned_agent_thread_count": 0,
                    },
                }
            ],
            valid_trajectories=1,
        )
        too_many_spawns = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "agent_events": {
                        **complete["agent_events"],
                        "spawn_agent_completed_count": 3,
                        "spawned_agent_thread_count": 3,
                    },
                }
            ],
            valid_trajectories=1,
        )

        self.assertTrue(passed["passed"])
        self.assertFalse(missing_spawn["passed"])
        self.assertFalse(too_many_spawns["passed"])
        self.assertEqual(
            missing_spawn["checks"]["actual_worker_launches"],
            {"expected": 2, "actual": 0},
        )
        self.assertEqual(
            missing_spawn["checks"]["spawn_agent_event_coverage"],
            {"expected": 2, "actual": 0},
        )
        self.assertEqual(
            too_many_spawns["checks"]["actual_worker_launches"],
            {"expected": 2, "actual": 3},
        )

    def test_goal_plus_pi_completion_uses_persisted_session_evidence(self) -> None:
        cell = {
            "method": "goal-plus-pi",
            "outer_replicas": 1,
            "inner_search_concurrency": 2,
        }
        complete = {
            "edgebench_score": 40.0,
            "goal_plus": {
                "candidates": 2,
                "agent_sessions": 2,
                "worker_verifier_runs": 3,
                "verifier_candidate_ids": ["c001", "c002"],
                "selected_candidate_ids": ["c001"],
                "promoted_candidate_ids": ["c001"],
            },
            "agent_events": {
                "spawn_agent_completed_count": 0,
                "goal_plus": {
                    "candidate_ids": [],
                    "agent_session_ids": [],
                    "verifier_ledger": [],
                    "selected_candidate_ids": [],
                    "promoted_candidate_ids": [],
                },
            },
        }
        evidence = EDGE.goal_plus_completion_evidence(
            cell,
            [complete],
            valid_trajectories=1,
        )
        too_many_sessions = EDGE.goal_plus_completion_evidence(
            cell,
            [
                {
                    **complete,
                    "goal_plus": {
                        **complete["goal_plus"],
                        "agent_sessions": 3,
                    },
                }
            ],
            valid_trajectories=1,
        )

        self.assertTrue(evidence["passed"])
        self.assertFalse(too_many_sessions["passed"])
        self.assertNotIn("actual_worker_launches", evidence["checks"])
        self.assertEqual(
            too_many_sessions["checks"]["agent_sessions"],
            {"expected": 2, "actual": 3},
        )

    def test_finalize_downgrades_missing_goal_plus_evidence_to_partial(self) -> None:
        destination = self.temp / "campaign-finalize"
        cell_dir = destination / "cells" / "vliw--goal-plus-codex"
        cell_dir.mkdir(parents=True)
        campaign = {
            "campaign_id": "campaign-finalize",
            "state": "completed",
            "task_ids": ["vliw_kernel_optimization"],
            "cells": [
                {
                    "cell_id": "vliw--goal-plus-codex",
                    "task_id": "vliw_kernel_optimization",
                    "method": "goal-plus-codex",
                    "state": "completed",
                }
            ],
        }
        cell = {
            "cell_id": "vliw--goal-plus-codex",
            "task_id": "vliw_kernel_optimization",
            "method": "goal-plus-codex",
            "state": "completed",
        }
        (destination / "campaign.json").write_text(json.dumps(campaign))
        (cell_dir / "cell.json").write_text(json.dumps(cell))
        original_summary = EDGE_REPORTING.summarize_cell
        original_reference = EDGE_REPORTING.load_paper_reference
        original_workbook = EDGE_REPORTING.write_comparison_workbook
        EDGE_REPORTING.summarize_cell = lambda *_args, **_kwargs: {
            "cell_id": "vliw--goal-plus-codex",
            "task_id": "vliw_kernel_optimization",
            "model": "gpt-test",
            "reasoning_effort": "medium",
            "wall_time_seconds": 60,
            "live_search_concurrency": 2,
            "completion_evidence": {"passed": False},
            "incomplete_reason": "missing worker evidence",
        }
        EDGE_REPORTING.load_paper_reference = lambda: {
            "tasks": {"vliw_kernel_optimization": {}}
        }
        EDGE_REPORTING.write_comparison_workbook = lambda *_args, **_kwargs: None
        try:
            payload = EDGE.finalize_campaign(destination)
        finally:
            EDGE_REPORTING.summarize_cell = original_summary
            EDGE_REPORTING.load_paper_reference = original_reference
            EDGE_REPORTING.write_comparison_workbook = original_workbook

        self.assertFalse(payload["completion_evidence_passed"])
        self.assertEqual(
            json.loads((destination / "campaign.json").read_text())["state"],
            "partial",
        )
        self.assertEqual(
            json.loads((cell_dir / "cell.json").read_text())["state"],
            "partial",
        )

    def test_finalize_recovers_prior_evidence_downgrade(self) -> None:
        destination = self.temp / "campaign-recover"
        cell_id = "vliw--goal-plus-pi"
        cell_dir = destination / "cells" / cell_id
        cell_dir.mkdir(parents=True)
        incomplete_reason = "missing promotion evidence"
        campaign = {
            "campaign_id": "campaign-recover",
            "state": "partial",
            "completion_evidence_passed": True,
            "incomplete_cells": {cell_id: incomplete_reason},
            "task_ids": ["vliw_kernel_optimization"],
            "cells": [
                {
                    "cell_id": cell_id,
                    "task_id": "vliw_kernel_optimization",
                    "method": "goal-plus-pi",
                    "state": "partial",
                    "incomplete_reason": incomplete_reason,
                }
            ],
        }
        cell = {
            "cell_id": cell_id,
            "task_id": "vliw_kernel_optimization",
            "method": "goal-plus-pi",
            "state": "partial",
            "incomplete_reason": incomplete_reason,
        }
        controller = {
            "state": "partial",
            "returncode": 2,
            "completion_evidence_passed": False,
        }
        (destination / "campaign.json").write_text(json.dumps(campaign))
        (destination / "controller.json").write_text(json.dumps(controller))
        (cell_dir / "cell.json").write_text(json.dumps(cell))
        original_summary = EDGE_REPORTING.summarize_cell
        original_reference = EDGE_REPORTING.load_paper_reference
        original_workbook = EDGE_REPORTING.write_comparison_workbook
        EDGE_REPORTING.summarize_cell = lambda *_args, **_kwargs: {
            "cell_id": cell_id,
            "task_id": "vliw_kernel_optimization",
            "model": "gpt-test",
            "reasoning_effort": "medium",
            "wall_time_seconds": 60,
            "live_search_concurrency": 2,
            "completion_evidence": {"passed": True},
            "incomplete_reason": None,
        }
        EDGE_REPORTING.load_paper_reference = lambda: {
            "tasks": {"vliw_kernel_optimization": {}}
        }
        EDGE_REPORTING.write_comparison_workbook = lambda *_args, **_kwargs: None
        try:
            payload = EDGE.finalize_campaign(destination)
        finally:
            EDGE_REPORTING.summarize_cell = original_summary
            EDGE_REPORTING.load_paper_reference = original_reference
            EDGE_REPORTING.write_comparison_workbook = original_workbook

        recovered_campaign = json.loads((destination / "campaign.json").read_text())
        recovered_cell = json.loads((cell_dir / "cell.json").read_text())
        recovered_controller = json.loads((destination / "controller.json").read_text())
        self.assertTrue(payload["completion_evidence_passed"])
        self.assertEqual(recovered_campaign["state"], "completed")
        self.assertTrue(recovered_campaign["completion_evidence_passed"])
        self.assertNotIn("incomplete_cells", recovered_campaign)
        self.assertEqual(recovered_campaign["cells"][0]["state"], "completed")
        self.assertNotIn("incomplete_reason", recovered_campaign["cells"][0])
        self.assertEqual(recovered_cell["state"], "completed")
        self.assertNotIn("incomplete_reason", recovered_cell)
        self.assertEqual(recovered_controller["state"], "completed")
        self.assertEqual(recovered_controller["returncode"], 0)
        self.assertTrue(recovered_controller["completion_evidence_passed"])


if __name__ == "__main__":
    unittest.main()
